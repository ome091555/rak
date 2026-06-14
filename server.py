import csv
import io
import os
import sqlite3
import uuid
import urllib.request
import json
from datetime import datetime, timezone, timedelta
from flask import Flask, redirect, render_template_string, request, session, url_for, jsonify, Response, send_file

from legal import terms_html, privacy_html, tokushoho_html

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

JST = timezone(timedelta(hours=9))
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'rak-secret-2026')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
DATABASE = os.environ.get('DATABASE', 'rak.db')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(DATABASE)), 'uploads')
STRIPE_SECRET_KEY = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_PRICE_ID_PRO = os.environ.get('STRIPE_PRICE_ID_PRO', '')
STRIPE_PRICE_ID_PRO_YEARLY = os.environ.get('STRIPE_PRICE_ID_PRO_YEARLY', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
NOTIFY_EMAIL = 'm.ome.091555@gmail.com'
FREE_MEMBER_LIMIT = 20
BASIC_AUTH_USER = os.environ.get('BASIC_AUTH_USER', '')
BASIC_AUTH_PASS = os.environ.get('BASIC_AUTH_PASS', '')
PROMO_CODES = [c.strip() for c in os.environ.get('PROMO_CODES', '').split(',') if c.strip()]
VAPID_PUBLIC_KEY  = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_PRIVATE_KEY = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_EMAIL       = os.environ.get('VAPID_EMAIL', 'mailto:m.ome.091555@gmail.com')
GOOGLE_SITE_VERIFICATION = os.environ.get('GOOGLE_SITE_VERIFICATION', '')
FIREBASE_SERVICE_ACCOUNT_JSON = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON', '')


def base_url():
    """Railway等のリバースプロキシ環境でも常にhttpsを返す"""
    url = request.host_url
    if not request.host.startswith('localhost') and not request.host.startswith('127.0.0.1'):
        url = url.replace('http://', 'https://')
    return url


_BOT_UA_WORDS = ('bot', 'crawler', 'spider', 'facebookexternalhit', 'preview', 'curl', 'python-requests')

def log_lp_event(event, src=''):
    """LPファネル計測。失敗してもページ表示には影響させない"""
    try:
        ua = (request.headers.get('User-Agent') or '').lower()
        if any(w in ua for w in _BOT_UA_WORDS):
            return
        is_mobile = 1 if ('mobile' in ua or 'iphone' in ua or 'android' in ua) else 0
        ref = (request.referrer or '')[:200]
        conn = get_db()
        conn.execute('INSERT INTO lp_events (id,event,src,ref,is_mobile,created_at) VALUES (?,?,?,?,?,?)',
                     (new_id(), event, src[:40], ref, is_mobile, now_str()))
        conn.commit()
        conn.close()
    except Exception:
        pass


@app.before_request
def redirect_apex_to_www():
    """wwwなし（rakapp.jp）はwwwありへ301リダイレクトして正規URLに統一する"""
    if request.host == 'rakapp.jp':
        return redirect('https://www.rakapp.jp' + request.full_path.rstrip('?'), code=301)


# ── Web Push 通知 ────────────────────────────────────────────────────────

def _vapid_headers(endpoint):
    """VAPIDのAuthorizationヘッダーを生成する"""
    try:
        import time, base64
        import jwt as pyjwt
        from cryptography.hazmat.primitives.serialization import load_pem_private_key
        aud = '/'.join(endpoint.split('/')[:3])
        token = pyjwt.encode(
            {'sub': VAPID_EMAIL, 'aud': aud, 'exp': int(time.time()) + 43200},
            VAPID_PRIVATE_KEY,
            algorithm='ES256'
        )
        return {
            'Authorization': f'vapid t={token},k={VAPID_PUBLIC_KEY}',
            'Content-Type': 'application/json',
            'TTL': '86400'
        }
    except Exception:
        return None

_fcm_token_cache = {'token': None, 'exp': 0}

def _fcm_access_token():
    """FirebaseサービスアカウントからFCM用OAuth2アクセストークンを取得する（キャッシュ付き）"""
    import time
    if _fcm_token_cache['token'] and time.time() < _fcm_token_cache['exp'] - 60:
        return _fcm_token_cache['token']
    try:
        import jwt as pyjwt
        import requests as _req
        sa = json.loads(FIREBASE_SERVICE_ACCOUNT_JSON)
        now = int(time.time())
        assertion = pyjwt.encode(
            {
                'iss': sa['client_email'],
                'scope': 'https://www.googleapis.com/auth/firebase.messaging',
                'aud': 'https://oauth2.googleapis.com/token',
                'iat': now,
                'exp': now + 3600,
            },
            sa['private_key'],
            algorithm='RS256'
        )
        res = _req.post(
            'https://oauth2.googleapis.com/token',
            data={
                'grant_type': 'urn:ietf:params:oauth:grant-type:jwt-bearer',
                'assertion': assertion,
            },
            timeout=10
        )
        token = res.json()['access_token']
        _fcm_token_cache['token'] = token
        _fcm_token_cache['exp'] = now + 3600
        return token
    except Exception:
        return None

def _send_fcm_to_team(team_id, title, body, url='/'):
    """チームの全ネイティブアプリ端末にFCM経由でプッシュ通知を送る"""
    if not FIREBASE_SERVICE_ACCOUNT_JSON:
        return
    access_token = _fcm_access_token()
    if not access_token:
        return
    try:
        project_id = json.loads(FIREBASE_SERVICE_ACCOUNT_JSON)['project_id']
    except Exception:
        return
    conn = get_db()
    tokens = conn.execute('SELECT * FROM native_push_tokens WHERE team_id=?', (team_id,)).fetchall()
    import requests as _req
    for t in tokens:
        try:
            res = _req.post(
                f'https://fcm.googleapis.com/v1/projects/{project_id}/messages:send',
                headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
                data=json.dumps({'message': {
                    'token': t['token'],
                    'notification': {'title': title, 'body': body},
                    'data': {'url': url},
                }}),
                timeout=10
            )
            # 無効トークン（アンインストール等）は削除する
            if res.status_code in (400, 404):
                conn.execute('DELETE FROM native_push_tokens WHERE id=?', (t['id'],))
                conn.commit()
        except Exception:
            pass
    conn.close()

def send_push_to_team(team_id, title, body, url='/'):
    """チームの全購読者に通知を送る（Web Push＋ネイティブFCMの二系統）"""
    _send_fcm_to_team(team_id, title, body, url)
    if not VAPID_PUBLIC_KEY or not VAPID_PRIVATE_KEY:
        return
    conn = get_db()
    subs = conn.execute('SELECT * FROM push_subscriptions WHERE team_id=?', (team_id,)).fetchall()
    conn.close()
    payload = json.dumps({'title': title, 'body': body, 'url': url})
    try:
        from pywebpush import webpush, WebPushException
        for s in subs:
            try:
                webpush(
                    subscription_info={
                        'endpoint': s['endpoint'],
                        'keys': {'p256dh': s['p256dh'], 'auth': s['auth']},
                    },
                    data=payload,
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims={'sub': VAPID_EMAIL},
                    ttl=86400,
                )
            except WebPushException:
                pass
            except Exception:
                pass
    except ImportError:
        # pywebpush未導入環境のフォールバック（ペイロード暗号化不可のため本文なしで送る）
        import requests as _req
        for s in subs:
            try:
                headers = _vapid_headers(s['endpoint'])
                if not headers:
                    continue
                _req.post(s['endpoint'], headers=headers, timeout=10)
            except Exception:
                pass

# ── メール送信 ────────────────────────────────────────────────────────────

def send_inquiry_email(team_name, name, email, subject, message):
    """お問い合わせをResend経由でGmailに通知する"""
    if not RESEND_API_KEY:
        print('[RESEND] RESEND_API_KEY が未設定')
        return
    try:
        import requests as _req
        res = _req.post(
            'https://api.resend.com/emails',
            headers={
                'Authorization': f'Bearer {RESEND_API_KEY}',
                'Content-Type': 'application/json',
                'User-Agent': 'RakApp/1.0',
            },
            json={
                'from': 'Rak <send@rakapp.jp>',
                'to': [NOTIFY_EMAIL],
                'subject': f'【Rakお問い合わせ】{subject or "（表題なし）"} - {team_name}',
                'text': f'Rakにお問い合わせが届きました。\n\n■ チーム名：{team_name}\n■ お名前：{name}\n■ メールアドレス：{email}\n■ 表題：{subject or "（未選択）"}\n■ メッセージ：\n{message}\n\n---\n返信先：{email}',
            },
            timeout=10
        )
        print(f'[RESEND] ステータス: {res.status_code} {res.text}')
    except Exception as e:
        print(f'[RESEND ERROR] {type(e).__name__}: {e}')


# ── DB ────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS teams (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            sport TEXT DEFAULT '',
            team_code TEXT UNIQUE NOT NULL,
            admin_password TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS events (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            event_date TEXT NOT NULL,
            event_time TEXT DEFAULT '',
            location TEXT DEFAULT '',
            note TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS rsvps (
            id TEXT PRIMARY KEY,
            event_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(event_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS notices (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            body TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS reads (
            notice_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            read_at TEXT NOT NULL,
            PRIMARY KEY(notice_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS members (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            name TEXT NOT NULL,
            number TEXT DEFAULT '',
            position TEXT DEFAULT '',
            joined_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS fees (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            due_date TEXT DEFAULT '',
            note TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS fee_payments (
            id TEXT PRIMARY KEY,
            fee_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            paid INTEGER DEFAULT 0,
            paid_at TEXT DEFAULT '',
            UNIQUE(fee_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS surveys (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS survey_options (
            id TEXT PRIMARY KEY,
            survey_id TEXT NOT NULL,
            label TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS survey_answers (
            id TEXT PRIMARY KEY,
            survey_id TEXT NOT NULL,
            option_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            answered_at TEXT NOT NULL,
            UNIQUE(survey_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS ai_templates (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            body TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS order_forms (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            deadline TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS order_form_fields (
            id TEXT PRIMARY KEY,
            form_id TEXT NOT NULL,
            label TEXT NOT NULL,
            field_type TEXT DEFAULT 'text',
            options TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS order_responses (
            id TEXT PRIMARY KEY,
            form_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL,
            UNIQUE(form_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS order_response_values (
            id TEXT PRIMARY KEY,
            response_id TEXT NOT NULL,
            field_id TEXT NOT NULL,
            value TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS order_form_photos (
            id TEXT PRIMARY KEY,
            form_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            mime_type TEXT DEFAULT 'image/jpeg',
            uploaded_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS app_feedback (
            id TEXT PRIMARY KEY,
            name TEXT DEFAULT '',
            message TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS admin_memos (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS memo_files (
            id TEXT PRIMARY KEY,
            memo_id TEXT NOT NULL,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS uniforms (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS uniform_assignments (
            id TEXT PRIMARY KEY,
            uniform_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            size TEXT DEFAULT '',
            number TEXT DEFAULT '',
            received INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            UNIQUE(uniform_id, member_name)
        );
        CREATE TABLE IF NOT EXISTS ledger (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            type TEXT NOT NULL,
            title TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            category TEXT DEFAULT '',
            entry_date TEXT NOT NULL,
            memo TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            token TEXT UNIQUE NOT NULL,
            expires_at TEXT NOT NULL,
            used INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS push_subscriptions (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            endpoint TEXT NOT NULL,
            p256dh TEXT NOT NULL,
            auth TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(team_id, member_name, endpoint)
        );
        CREATE TABLE IF NOT EXISTS lp_events (
            id TEXT PRIMARY KEY,
            event TEXT NOT NULL,
            src TEXT DEFAULT '',
            ref TEXT DEFAULT '',
            is_mobile INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS native_push_tokens (
            id TEXT PRIMARY KEY,
            team_id TEXT NOT NULL,
            member_name TEXT NOT NULL,
            platform TEXT NOT NULL,
            token TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(team_id, member_name, token)
        );
    ''')
    conn.commit()
    # migration: end_date column
    try:
        conn.execute('ALTER TABLE events ADD COLUMN end_date TEXT DEFAULT ""')
        conn.commit()
    except Exception:
        pass
    # migration: end_time column
    try:
        conn.execute('ALTER TABLE events ADD COLUMN end_time TEXT DEFAULT ""')
        conn.commit()
    except Exception:
        pass
    # migration: plan / stripe columns
    for col_sql in [
        'ALTER TABLE teams ADD COLUMN plan TEXT DEFAULT "free"',
        'ALTER TABLE teams ADD COLUMN stripe_customer_id TEXT DEFAULT ""',
        'ALTER TABLE teams ADD COLUMN stripe_subscription_id TEXT DEFAULT ""',
        'ALTER TABLE app_feedback ADD COLUMN team_name TEXT DEFAULT ""',
        'ALTER TABLE app_feedback ADD COLUMN email TEXT DEFAULT ""',
        'ALTER TABLE app_feedback ADD COLUMN subject TEXT DEFAULT ""',
        'ALTER TABLE teams ADD COLUMN admin_memo TEXT DEFAULT ""',
        'ALTER TABLE teams ADD COLUMN admin_email TEXT DEFAULT ""',
        'ALTER TABLE teams ADD COLUMN trial_end TEXT DEFAULT ""',
        'ALTER TABLE events ADD COLUMN event_color TEXT DEFAULT ""',
        'ALTER TABLE teams ADD COLUMN viewer_token TEXT DEFAULT ""',
        'ALTER TABLE uniform_assignments ADD COLUMN quantity INTEGER DEFAULT 1',
        'ALTER TABLE events ADD COLUMN rsvp_mode TEXT DEFAULT "both"',
        'ALTER TABLE fee_payments ADD COLUMN reported INTEGER DEFAULT 0',
        'ALTER TABLE fee_payments ADD COLUMN reported_at TEXT DEFAULT ""',
    ]:
        try:
            conn.execute(col_sql)
            conn.commit()
        except Exception:
            pass
    # 既存チームに viewer_token が未設定のものへ自動付与
    import secrets as _secrets
    rows = conn.execute("SELECT id FROM teams WHERE viewer_token='' OR viewer_token IS NULL").fetchall()
    for row in rows:
        token = _secrets.token_urlsafe(16)
        conn.execute("UPDATE teams SET viewer_token=? WHERE id=?", (token, row['id']))
    if rows:
        conn.commit()
    conn.close()

init_db()

# ── Helpers ───────────────────────────────────────────────────────

def new_id():
    return str(uuid.uuid4())[:8]

def now_str():
    return datetime.now(JST).strftime('%Y-%m-%d %H:%M')

def is_pro(team):
    if not team:
        return False
    if team['plan'] in ('pro', 'league'):
        return True
    # トライアル期間中かチェック
    trial_end = team['trial_end'] if team['trial_end'] else ''
    if trial_end:
        try:
            end_dt = datetime.strptime(trial_end, '%Y-%m-%d').replace(tzinfo=JST)
            if datetime.now(JST) <= end_dt:
                return True
        except Exception:
            pass
    return False

def get_trial_days_left(team):
    """トライアル残り日数。トライアル中でなければNone。"""
    if not team or not team['trial_end']:
        return None
    if team['plan'] in ('pro', 'league'):
        return None  # 正規課金中はトライアル扱いしない
    try:
        end_dt = datetime.strptime(team['trial_end'], '%Y-%m-%d').replace(tzinfo=JST)
        now = datetime.now(JST)
        if now > end_dt:
            return None
        return max(0, (end_dt.date() - now.date()).days)
    except Exception:
        return None

def count_team_members(team_id, conn=None):
    own = conn is None
    if own:
        conn = get_db()
    n = conn.execute('SELECT COUNT(*) FROM members WHERE team_id=?', (team_id,)).fetchone()[0]
    if own:
        conn.close()
    return n

def can_add_team_member(team, conn=None):
    return True

def member_count_label(team, count):
    return f'{count}名'

def pro_gate(code, team, active='home'):  # noqa
    body = f'''
<div class="container" style="max-width:480px;padding-top:40px">
  <div class="card" style="text-align:center;padding:40px 24px">
    <div style="margin-bottom:16px">{_ICO_LOCK}</div>
    <h1 style="font-size:22px;margin-bottom:8px">Proプランの機能です</h1>
    <p style="color:#666;font-size:14px;margin-bottom:24px">この機能を使うにはRak Proへのアップグレードが必要です。</p>
    <div style="background:#f5f7fb;border-radius:12px;padding:20px;margin-bottom:24px;text-align:left">
      <div style="font-weight:700;margin-bottom:12px;color:#d97706">Proプランでできること</div>
      <div style="font-size:13px;color:#444;line-height:2.2">
        {_CHK} 集金・支払い管理<br>
        {_CHK} 注文フォーム<br>
        {_CHK} アンケート<br>
        {_CHK} AI文章生成<br>
        {_CHK} AIスケジュール自動生成<br>
        {_CHK} Excel出力
      </div>
    </div>
    <div style="font-size:28px;font-weight:900;color:#d97706;margin-bottom:4px">¥980<span style="font-size:14px;font-weight:500;color:#888">/月</span></div>
    <div style="font-size:12px;color:#888;margin-bottom:24px">年払い ¥9,800（2ヶ月分お得）</div>
    <a href="/t/{code}/upgrade" class="btn btn-blue btn-block" style="margin-top:0">Proにアップグレード</a>
    <div style="margin-top:12px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
  </div>
</div>'''
    return page('Proプランへアップグレード', body, code, active=active)

def csv_response(csv_str, filename):
    import urllib.parse
    encoded_name = urllib.parse.quote(filename.encode('utf-8'))
    return Response(
        csv_str.encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_name}"}
    )

def excel_response(rows, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import urllib.parse
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row in enumerate(rows):
        ws.append(list(row))
        if i == 0:
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='2563EB')
                cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded_name = urllib.parse.quote(filename.encode('utf-8'))
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_name}"}
    )

def get_team(code):
    conn = get_db()
    t = conn.execute('SELECT * FROM teams WHERE team_code=?', (code.upper(),)).fetchone()
    conn.close()
    return t

def is_admin(code):
    return session.get(f'admin_{code}') is True

def get_member(code):
    return session.get(f'member_{code}', '')

def fmt_date(s):
    try:
        d = datetime.strptime(s, '%Y-%m-%d')
        wd = ['月','火','水','木','金','土','日'][d.weekday()]
        return f"{d.month}/{d.day}（{wd}）"
    except:
        return s

def fmt_date_range(start, end):
    if not end or end == start:
        return fmt_date(start)
    return f'{fmt_date(start)} 〜 {fmt_date(end)}'

def fmt_datetime(s):
    try:
        d = datetime.strptime(s, '%Y-%m-%d %H:%M')
        return d.strftime('%-m/%-d %H:%M')
    except:
        return s

# ── Base CSS & layout ─────────────────────────────────────────────

FONT = '<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;600;700;800;900&family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">'

# Rak brand mark: amber background, white R + checkmark (v2)
NAV_MARK = (
    '<svg viewBox="0 0 130 120" width="22" height="20" fill="none" xmlns="http://www.w3.org/2000/svg">'
    '<rect width="130" height="120" rx="28" fill="#d97706"/>'
    '<path d="M 32 94 L 32 26 L 60 26 C 74 26 80 36 80 46 C 80 56 74 64 60 64 L 32 64" stroke="white" stroke-width="11" stroke-linejoin="miter" fill="none"/>'
    '<path d="M 54 64 L 72 94 L 112 28" stroke="white" stroke-width="11" stroke-linejoin="miter" fill="none"/>'
    '</svg>'
)
FAVICON_LINK = (
    '<link rel="icon" type="image/svg+xml" href="data:image/svg+xml,'
    "%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E"
    "%3Crect width='100' height='100' fill='%23d97706'/%3E"
    "%3Cpath d='M 18 82 L 18 18 L 46 18 C 60 18 66 28 66 38 C 66 48 60 56 46 56 L 18 56' stroke='white' stroke-width='10' stroke-linejoin='miter' fill='none'/%3E"
    "%3Cpath d='M 40 56 L 58 82 L 92 16' stroke='white' stroke-width='10' stroke-linejoin='miter' fill='none'/%3E"
    "%3C/svg%3E"
    '"><meta name="theme-color" content="#d97706">'
    '<link rel="manifest" href="/manifest.json">'
    '<meta name="apple-mobile-web-app-capable" content="yes">'
    '<meta name="apple-mobile-web-app-status-bar-style" content="default">'
    '<meta name="apple-mobile-web-app-title" content="Rak">'
    '<link rel="apple-touch-icon" href="/apple-touch-icon.png">'
)

PWA_SW = '''<script>
function rakIsNativeApp(){
  return !!(window.Capacitor&&window.Capacitor.isNativePlatform&&window.Capacitor.isNativePlatform());
}
if(!rakIsNativeApp()&&"serviceWorker"in navigator){
  navigator.serviceWorker.register("/sw.js");
}
async function rakSubscribePushNative(code){
  try{
    const Push=window.Capacitor.Plugins.PushNotifications;
    const platform=window.Capacitor.getPlatform();
    Push.addListener("registration",async(t)=>{
      await fetch("/t/"+code+"/push/subscribe-native",{
        method:"POST",headers:{"Content-Type":"application/json"},
        body:JSON.stringify({token:t.value,platform:platform})
      });
      localStorage.setItem("rak_push_"+code,"1");
    });
    Push.addListener("pushNotificationActionPerformed",(n)=>{
      const url=n.notification&&n.notification.data&&n.notification.data.url;
      if(url)location.href=url;
    });
    const perm=await Push.requestPermissions();
    if(perm.receive==="granted")await Push.register();
  }catch(e){}
}
async function rakSubscribePush(code){
  if(rakIsNativeApp()){return rakSubscribePushNative(code);}
  try{
    const reg=await navigator.serviceWorker.ready;
    const res=await fetch("/push/vapid-public-key");
    const {publicKey}=await res.json();
    if(!publicKey)return;
    const sub=await reg.pushManager.subscribe({
      userVisibleOnly:true,
      applicationServerKey:publicKey
    });
    await fetch("/t/"+code+"/push/subscribe",{
      method:"POST",headers:{"Content-Type":"application/json"},
      body:JSON.stringify(sub.toJSON())
    });
    localStorage.setItem("rak_push_"+code,"1");
  }catch(e){}
}
function rakPushBannerOn(code){
  const hide=()=>{const b=document.getElementById("rak-push-banner");if(b)b.style.display="none";};
  if(rakIsNativeApp()){rakSubscribePush(code);hide();return;}
  Notification.requestPermission().then(p=>{if(p==="granted"){rakSubscribePush(code);hide();}});
}
async function rakRequestPush(code){
  if(localStorage.getItem("rak_push_"+code)==="1")return;
  if(rakIsNativeApp()){
    const banner=document.getElementById("rak-push-banner");
    if(banner)banner.style.display="flex";
    return;
  }
  if(!("Notification"in window)||!("serviceWorker"in navigator))return;
  if(Notification.permission==="granted"){rakSubscribePush(code);return;}
  if(Notification.permission==="denied")return;
  const banner=document.getElementById("rak-push-banner");
  if(banner)banner.style.display="flex";
}
</script>'''

# ── オリジナル SVG アイコン (Apple絵文字に依存しない) ────────────────────────
# 空状態イラスト 64×64 (アンバー円 + アイコン)
_SVG_EMPTY_BELL = (
    '<svg width="64" height="64" viewBox="0 0 64 64" fill="none">'
    '<circle cx="32" cy="32" r="30" fill="#fef3c7" stroke="#d97706" stroke-width="2"/>'
    '<path d="M32 13C26 13 21 18 21 24L21 34L16 39H48L43 34V24C43 18 38 13 32 13Z"'
    ' stroke="#d97706" stroke-width="2.5" stroke-linejoin="round" fill="none"/>'
    '<path d="M26 39Q26 44.5 32 44.5Q38 44.5 38 39" stroke="#d97706" stroke-width="2.5" fill="none"/>'
    '</svg>'
)
_SVG_EMPTY_COIN = (
    '<svg width="64" height="64" viewBox="0 0 64 64" fill="none">'
    '<circle cx="32" cy="32" r="30" fill="#fef3c7" stroke="#d97706" stroke-width="2"/>'
    '<circle cx="32" cy="32" r="15" stroke="#d97706" stroke-width="2.5" fill="none"/>'
    '<path d="M26 23L32 30L38 23" stroke="#d97706" stroke-width="2.5" fill="none"'
    ' stroke-linejoin="round" stroke-linecap="round"/>'
    '<line x1="26" y1="28" x2="38" y2="28" stroke="#d97706" stroke-width="2"/>'
    '<line x1="26" y1="33" x2="38" y2="33" stroke="#d97706" stroke-width="2"/>'
    '<line x1="32" y1="30" x2="32" y2="41" stroke="#d97706" stroke-width="2.5" stroke-linecap="round"/>'
    '</svg>'
)
_SVG_EMPTY_FORM = (
    '<svg width="64" height="64" viewBox="0 0 64 64" fill="none">'
    '<circle cx="32" cy="32" r="30" fill="#fef3c7" stroke="#d97706" stroke-width="2"/>'
    '<rect x="18" y="14" width="28" height="36" rx="3" stroke="#d97706" stroke-width="2.5" fill="none"/>'
    '<line x1="23" y1="24" x2="41" y2="24" stroke="#d97706" stroke-width="2"/>'
    '<line x1="23" y1="31" x2="41" y2="31" stroke="#d97706" stroke-width="2"/>'
    '<line x1="23" y1="38" x2="33" y2="38" stroke="#d97706" stroke-width="2"/>'
    '</svg>'
)
_SVG_EMPTY_CHART = (
    '<svg width="64" height="64" viewBox="0 0 64 64" fill="none">'
    '<circle cx="32" cy="32" r="30" fill="#fef3c7" stroke="#d97706" stroke-width="2"/>'
    '<rect x="15" y="38" width="9" height="12" rx="2" fill="#d97706" opacity="0.4"/>'
    '<rect x="27.5" y="28" width="9" height="22" rx="2" fill="#d97706" opacity="0.7"/>'
    '<rect x="40" y="18" width="9" height="32" rx="2" fill="#d97706"/>'
    '</svg>'
)
# セクション / タイル用アイコン 20×20 Lucide スタイル (currentColor)
_ICO_PEOPLE = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<circle cx="8" cy="6" r="3"/>'
    '<path d="M2 18c0-3.3 2.7-6 6-6s6 2.7 6 6"/>'
    '<circle cx="15" cy="7" r="2.5"/>'
    '<path d="M18 18c0-2.7-1.5-5-3.5-6"/>'
    '</svg>'
)
_ICO_CALENDAR = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<rect x="3" y="4" width="14" height="13" rx="2"/>'
    '<path d="M7 2v4M13 2v4M3 8h14"/>'
    '</svg>'
)
_ICO_CLIPBOARD = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<rect x="5" y="2" width="10" height="16" rx="2"/>'
    '<path d="M8 7h4M8 10.5h4M8 14h2.5"/>'
    '</svg>'
)
_ICO_CHART_SM = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<path d="M10 3l1.8 5.5 5.2 1.5-5.2 1.5L10 17l-1.8-5.5L3 10l6-1.5z"/>'
    '</svg>'
)
_ICO_BELL_SM = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" style="vertical-align:middle">'
    '<path d="M5 9c0-2.8 2.2-5 5-5s5 2.2 5 5v3l1.5 2.5h-13L5 12V9z"/>'
    '<path d="M8.5 17.5a1.5 1.5 0 003 0"/>'
    '</svg>'
)
_ICO_MONEY_SM = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" style="vertical-align:middle">'
    '<circle cx="10" cy="10" r="7"/>'
    '<path d="M10 6v8M7.5 8c.5-1.5 5-1.5 5 1.5 0 2.5-5 1.5-5 4 0 2 4.5 1.5 5 0"/>'
    '</svg>'
)
# ユーザー表示 (ナビ)
_ICO_USER_SM = (
    '<svg width="14" height="14" viewBox="0 0 14 14" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" style="vertical-align:middle;color:#6b7280">'
    '<circle cx="7" cy="5" r="3"/>'
    '<path d="M1 14c0-3.3 2.7-6 6-6s6 2.7 6 6"/>'
    '</svg>'
)
# アンバーチェックマーク (機能リスト・ステータス)
_PRO_BADGE = '<span style="font-size:9px;font-weight:700;background:#d97706;color:#fff;border-radius:3px;padding:1px 5px;margin-left:4px;letter-spacing:.02em">PRO</span>'

_CHK = (
    '<svg width="16" height="16" viewBox="0 0 16 16" fill="none"'
    ' style="vertical-align:middle;margin-right:4px">'
    '<circle cx="8" cy="8" r="7" fill="#d97706"/>'
    '<path d="M4.5 8L6.8 10.5L11.5 5.5" stroke="white" stroke-width="2"'
    ' stroke-linecap="round" stroke-linejoin="round" fill="none"/>'
    '</svg>'
)
# ウェルカム画面: Rak ブランドマーク (48px)
_ICO_WELCOME = (
    '<svg viewBox="0 0 130 120" width="48" height="44" fill="none">'
    '<rect width="130" height="120" rx="28" fill="#d97706"/>'
    '<path d="M 32 94 L 32 26 L 60 26 C 74 26 80 36 80 46 C 80 56 74 64 60 64 L 32 64"'
    ' stroke="white" stroke-width="11" stroke-linejoin="miter" fill="none"/>'
    '<path d="M 54 64 L 72 94 L 112 28" stroke="white" stroke-width="11" stroke-linejoin="miter" fill="none"/>'
    '</svg>'
)
# ロック (Proゲート)
_ICO_LOCK = (
    '<svg width="52" height="52" viewBox="0 0 52 52" fill="none">'
    '<rect x="10" y="24" width="32" height="22" rx="5" fill="#fef3c7" stroke="#d97706" stroke-width="2.5"/>'
    '<path d="M18 24V17C18 12 22 8 26 8C30 8 34 12 34 17V24"'
    ' stroke="#d97706" stroke-width="2.5" fill="none" stroke-linecap="round"/>'
    '<circle cx="26" cy="35" r="3.5" fill="#d97706"/>'
    '<line x1="26" y1="38.5" x2="26" y2="42" stroke="#d97706" stroke-width="2.5" stroke-linecap="round"/>'
    '</svg>'
)
# お祝い (成功・アップグレード完了) 56px
_ICO_CELEBRATE = (
    '<svg width="56" height="56" viewBox="0 0 56 56" fill="none">'
    '<circle cx="28" cy="28" r="26" fill="#fef3c7" stroke="#d97706" stroke-width="2"/>'
    '<path d="M28 10L31.5 21H43L33.5 27.5L37 38.5L28 32L19 38.5L22.5 27.5L13 21H24.5L28 10Z"'
    ' fill="#d97706"/>'
    '</svg>'
)
# 小さいお祝い (インライン)
_ICO_CELEBRATE_SM = (
    '<svg width="18" height="18" viewBox="0 0 18 18" fill="none" style="vertical-align:middle">'
    '<path d="M9 2L11 8H17L12 11.5L14 17.5L9 14L4 17.5L6 11.5L1 8H7L9 2Z" fill="#d97706"/>'
    '</svg>'
)

# アドミンタイル用 メモ・メール・プランアイコン（Lucide スタイル）
_ICO_MEMO = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<path d="M13.5 3.5l3 3L7 16H4v-3L13.5 3.5z"/>'
    '<path d="M11.5 5.5l3 3"/>'
    '</svg>'
)
_ICO_MAIL = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<rect x="3" y="6" width="14" height="10" rx="2"/>'
    '<path d="M3 8l7 5 7-5"/>'
    '</svg>'
)
_ICO_CROWN = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<path d="M3 14l2.5-7 4.5 3.5L14 7l2.5 7H3z"/>'
    '<path d="M3 14h14"/>'
    '</svg>'
)
_ICO_HELP = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<circle cx="10" cy="10" r="7"/>'
    '<path d="M8 8c0-1.1.9-2 2-2s2 .9 2 2c0 1-1 1.5-2 2.5"/>'
    '<circle cx="10" cy="14.5" r=".5" fill="currentColor"/>'
    '</svg>'
)
_ICO_UNIFORM = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<path d="M7 3L3 7.5l3 1.5V17h8V9l3-1.5L14 3H7z"/>'
    '<path d="M7 3c.5 1.5 1.5 2.5 3 2.5s2.5-1 3-2.5"/>'
    '</svg>'
)
_ICO_LEDGER = (
    '<svg width="20" height="20" viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle">'
    '<rect x="3" y="2" width="14" height="16" rx="2"/>'
    '<path d="M7 7h6M7 10.5h6M7 14h4"/>'
    '<path d="M3 6h14" stroke-width="1"/>'
    '</svg>'
)

ICONS = {
    'home':     '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9.5L10 3l7 6.5"/><path d="M5 8.5V17h4v-4h2v4h4V8.5"/></svg>',
    'schedule': '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="4" width="14" height="13" rx="2"/><path d="M7 2v4M13 2v4M3 8h14"/></svg>',
    'notices':  '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><path d="M5 9c0-2.8 2.2-5 5-5s5 2.2 5 5v3l1.5 2.5h-13L5 12V9z"/><path d="M8.5 17.5a1.5 1.5 0 003 0"/></svg>',
    'members':  '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><circle cx="8" cy="6" r="3"/><path d="M2 18c0-3.3 2.7-6 6-6s6 2.7 6 6"/><circle cx="15" cy="7" r="2.5"/><path d="M18 18c0-2.7-1.5-5-3.5-6"/></svg>',
    'fees':     '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><circle cx="10" cy="10" r="7"/><path d="M10 6v8M7.5 8c.5-1.5 5-1.5 5 1.5 0 2.5-5 1.5-5 4 0 2 4.5 1.5 5 0"/></svg>',
    'orders':   '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><rect x="5" y="2" width="10" height="16" rx="2"/><path d="M8 7h4M8 10.5h4M8 14h2.5"/></svg>',
    'uniforms': '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M7 3L3 7.5l3 1.5V17h8V9l3-1.5L14 3H7z"/><path d="M7 3c.5 1.5 1.5 2.5 3 2.5s2.5-1 3-2.5"/></svg>',
    'admin':    '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><circle cx="10" cy="10" r="2.5"/><path d="M10 2v2.5M10 15.5V18M2 10h2.5M15.5 10H18M4.9 4.9l1.8 1.8M13.3 13.3l1.8 1.8M4.9 15.1l1.8-1.8M13.3 6.7l1.8-1.8"/></svg>',
    'ai':       '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M10 2.5l1.8 5 5.2 2-5.2 2-1.8 5-1.8-5-5.2-2 5.2-2z"/></svg>',
    'ask':      '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M17 12a2 2 0 01-2 2H6l-3 3V5a2 2 0 012-2h10a2 2 0 012 2v7z"/></svg>',
    'ledger':   '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="2" width="14" height="16" rx="2"/><path d="M7 7h6M7 10.5h6M7 14h4"/></svg>',
    'memo':     '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M13.5 3.5l3 3L7 16H4v-3L13.5 3.5z"/><path d="M11.5 5.5l3 3"/></svg>',
    'contact':  '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="6" width="14" height="10" rx="2"/><path d="M3 8l7 5 7-5"/></svg>',
    'help':     '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><circle cx="10" cy="10" r="7"/><path d="M8 8c0-1.1.9-2 2-2s2 .9 2 2c0 1-1 1.5-2 2.5"/><circle cx="10" cy="14.5" r=".5" fill="currentColor"/></svg>',
    'plan':     '<svg viewBox="0 0 20 20" fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"><path d="M3 14l2.5-7 4.5 3.5L14 7l2.5 7H3z"/><path d="M3 14h14"/></svg>',
}

CSS = '''
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --rak-black:#0a0a0a;
  --rak-ink:#111827;
  --rak-graphite:#4b5563;
  --rak-mute:#6b7280;
  --rak-line:#e5e7eb;
  --rak-line-soft:#f3f4f6;
  --rak-bg:#ffffff;
  --rak-bg-soft:#f9fafb;
  --rak-amber:#d97706;
  --rak-amber-deep:#b45309;
  --rak-amber-tint:#f9fafb;
  --rak-success:#16a34a;
  --rak-success-tint:#f0fdf4;
  --rak-danger:#dc2626;
  --rak-danger-tint:#fef2f2;
  --font-jp:"Noto Sans JP","Hiragino Sans","Hiragino Kaku Gothic ProN",system-ui,sans-serif;
  --font-num:"Inter","Noto Sans JP",system-ui,sans-serif;
}
html,body{font-family:var(--font-jp);background:var(--rak-bg-soft);color:var(--rak-ink);font-size:15px;line-height:1.6;min-height:100vh;-webkit-font-smoothing:antialiased}
a{color:inherit;text-decoration:none}
button{font-family:inherit}

/* Nav */
.nav{background:#fff;border-bottom:1px solid var(--rak-line);padding:0 16px;height:52px;display:flex;align-items:center;gap:10px;position:sticky;top:0;z-index:50}
.nav-logo{font-weight:600;font-size:17px;color:var(--rak-black);display:flex;align-items:center;gap:8px;letter-spacing:-0.01em}
.nav-icon{width:26px;height:26px;border-radius:6px;display:flex;align-items:center;justify-content:center;overflow:hidden}
.nav-team{font-size:13px;color:var(--rak-mute);font-weight:500;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:160px}
.nav-links-desktop{display:flex;gap:2px;margin-left:auto;align-items:center}
.nav-links-desktop a{font-size:13px;color:var(--rak-mute);padding:5px 10px;border-radius:6px;font-weight:500;display:inline-flex;align-items:center;gap:5px}
.nav-links-desktop a:hover{background:var(--rak-bg-soft);color:var(--rak-ink);text-decoration:none}
.nav-links-desktop a.active{color:var(--rak-ink);font-weight:700;background:var(--rak-bg-soft)}
/* Dropdown */
.nav-dropdown{position:relative;display:inline-flex}
.nav-dd-btn{font-size:13px;color:var(--rak-mute);padding:5px 9px;border-radius:6px;font-weight:500;display:inline-flex;align-items:center;gap:3px;cursor:pointer;background:none;border:none;font-family:inherit}
.nav-dd-btn:hover,.nav-dd-btn.active{background:var(--rak-bg-soft);color:var(--rak-ink)}
.nav-dd-btn.active{font-weight:700}
.nav-dd-menu{display:none;position:absolute;top:calc(100% + 6px);right:0;background:#fff;border:1px solid var(--rak-line);border-radius:10px;box-shadow:0 4px 20px rgba(0,0,0,.1);z-index:200;min-width:160px;padding:4px}
.nav-dd-menu.open{display:block}
.nav-dd-menu a{display:flex;align-items:center;gap:8px;padding:9px 14px;font-size:13px;color:var(--rak-mute);border-radius:7px;white-space:nowrap}
.nav-dd-menu a:hover{background:var(--rak-bg-soft);color:var(--rak-ink);text-decoration:none}
.nav-dd-menu a.active{color:var(--rak-ink);font-weight:700;background:var(--rak-bg-soft)}
.nav-dd-sep{height:1px;background:var(--rak-line);margin:4px 8px}

/* Bottom nav */
.bottom-nav{display:none;position:fixed;bottom:0;left:0;right:0;background:#fff;border-top:1px solid var(--rak-line);z-index:100;padding-bottom:env(safe-area-inset-bottom,0)}
.bottom-nav a{position:relative;flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:6px 2px;font-size:9px;color:var(--rak-mute);text-decoration:none;gap:3px;min-height:52px;font-weight:500}
.nav-badge{position:absolute;top:5px;left:calc(50% + 4px);background:#ef4444;color:#fff;border-radius:10px;font-size:9px;font-weight:600;padding:1px 5px;min-width:16px;text-align:center;line-height:14px;pointer-events:none}
.bottom-nav a.active{color:var(--rak-amber)}
.nav-b-icon{display:flex;align-items:center;justify-content:center;position:relative}
.nav-b-icon svg{width:22px;height:22px}
.nav-d-icon{display:inline-flex;vertical-align:-3px;margin-right:4px}
.nav-d-icon svg{width:15px;height:15px}

/* Layout */
.container{max-width:680px;margin:0 auto;padding:24px 16px}

/* Cards */
.card{background:#fff;border-radius:10px;padding:20px;border:1px solid var(--rak-line);margin-bottom:12px}
.card-sm{background:#fff;border-radius:8px;padding:12px 16px;border:1px solid var(--rak-line);margin-bottom:6px}

/* Typography */
h1{font-size:20px;font-weight:600;margin-bottom:4px;letter-spacing:-0.01em;color:var(--rak-black)}
h2{font-size:16px;font-weight:600;margin-bottom:12px}
h3{font-size:14px;font-weight:600}
label{display:block;font-size:12px;font-weight:500;color:var(--rak-mute);margin-bottom:5px;margin-top:14px;letter-spacing:0.01em}
label:first-of-type{margin-top:0}

/* Forms */
input[type=text],input[type=email],input[type=password],input[type=date],input[type=time],textarea,select{width:100%;border:1px solid var(--rak-line);border-radius:8px;padding:10px 12px;font-size:16px;outline:none;font-family:inherit;background:#fff;color:var(--rak-ink);box-sizing:border-box}
input[type=date],input[type=time]{-webkit-appearance:none;appearance:none}
input:focus,textarea:focus,select:focus{border-color:var(--rak-amber);box-shadow:0 0 0 3px rgba(217,119,6,.08)}
textarea{resize:vertical;min-height:80px}

/* Buttons */
.btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:10px 18px;border-radius:8px;font-size:14px;font-weight:500;cursor:pointer;border:none;font-family:inherit;transition:background .1s;text-decoration:none;text-align:center}
.btn:active{transform:scale(0.98)}
.btn-blue{background:var(--rak-black);color:#fff}
.btn-blue:hover{background:#1f2937;text-decoration:none;color:#fff}
.btn-outline{background:#fff;color:var(--rak-ink);border:1px solid var(--rak-line)}
.btn-outline:hover{background:var(--rak-bg-soft);text-decoration:none;color:var(--rak-ink)}
.btn-gray{background:var(--rak-bg-soft);color:var(--rak-graphite)}
.btn-gray:hover{background:var(--rak-line);text-decoration:none;color:var(--rak-graphite)}
.btn-amber{background:var(--rak-amber);color:#fff}
.btn-amber:hover{background:var(--rak-amber-deep);color:#fff;text-decoration:none}
.btn-block{display:block;width:100%;margin-top:12px}
.btn-sm{padding:6px 12px;font-size:13px;border-radius:6px}

/* Badges */
.badge{display:inline-flex;align-items:center;gap:4px;padding:2px 7px;border-radius:4px;font-size:11px;font-weight:500}
.badge-green{background:var(--rak-success-tint);color:var(--rak-success)}
.badge-red{background:var(--rak-danger-tint);color:var(--rak-danger)}
.badge-gray{background:var(--rak-bg-soft);color:var(--rak-mute)}
.badge-blue{background:var(--rak-bg-soft);color:var(--rak-amber)}

/* Alerts */
.msg-ok{background:var(--rak-success-tint);color:var(--rak-success);padding:12px 16px;border-radius:8px;margin-bottom:14px;font-weight:500;border:1px solid #bbf7d0}
.msg-err{background:var(--rak-danger-tint);color:var(--rak-danger);padding:12px 16px;border-radius:8px;margin-bottom:14px;font-weight:500}

/* Section labels */
.section-label{font-size:11px;font-weight:600;letter-spacing:0.08em;color:var(--rak-mute);text-transform:uppercase;display:inline-block;margin-bottom:10px}

/* Misc */
.empty{text-align:center;padding:32px 20px;color:var(--rak-mute);font-size:14px}
.row{display:flex;align-items:center;gap:10px}
.divider{border:none;border-top:1px solid var(--rak-line);margin:14px 0}

/* Dashboard special */
.team-code-card{background:var(--rak-black);color:#fff;border-radius:10px;padding:20px;margin-bottom:14px;position:relative;overflow:hidden;border:none}
.team-code-card::before{content:"";position:absolute;top:-20px;right:-20px;width:100px;height:100px;border-radius:50%;background:var(--rak-amber);opacity:.12}
.mini-stats{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:14px}
.mini-stat{background:var(--rak-bg-soft);border-radius:8px;padding:12px}
.mini-stat .v{font-family:var(--font-num);font-size:20px;font-weight:600;letter-spacing:-0.02em}
.mini-stat .v.amber{color:var(--rak-amber)}
.mini-stat .l{font-size:10px;color:var(--rak-mute);font-weight:500;margin-top:2px}

/* Event list */
.event-list{background:#fff;border:1px solid var(--rak-line);border-radius:10px;overflow:hidden;margin-bottom:14px}
.event-row{padding:14px 16px;border-bottom:1px solid var(--rak-line-soft);display:flex;gap:14px;align-items:flex-start}
.event-row:last-child{border-bottom:none}
.date-block{min-width:48px;text-align:center;background:var(--rak-bg-soft);border-radius:8px;padding:7px 5px}
.date-block.hl{background:var(--rak-black);color:#fff}
.date-block .month{font-size:9px;font-weight:500;opacity:.7}
.date-block .day{font-family:var(--font-num);font-size:20px;font-weight:600;line-height:1.1;letter-spacing:-0.02em}
.date-block .wd{font-size:9px;font-weight:500;opacity:.7}
.att-bar{display:flex;gap:6px;font-size:11px;font-weight:500}
.att-chip{flex:1;background:var(--rak-bg-soft);border-radius:6px;padding:5px 4px;text-align:center}
.att-chip .v{font-family:var(--font-num);font-weight:600;font-size:13px}
.att-chip .l{color:var(--rak-mute);font-size:9px;margin-top:1px}
.att-chip.green .v{color:var(--rak-success)}
.att-chip.red .v{color:var(--rak-danger)}
.att-chip.amber .v{color:var(--rak-amber)}

/* Notice rows */
.notice-list{background:#fff;border:1px solid var(--rak-line);border-radius:10px;overflow:hidden;margin-bottom:14px}
.notice-row{padding:12px 16px;border-bottom:1px solid var(--rak-line-soft);display:flex;align-items:center;gap:12px}
.notice-row:last-child{border-bottom:none}
.notice-row .read-bar{width:48px;height:2px;background:var(--rak-line-soft);border-radius:2px;overflow:hidden;margin-top:3px}
.notice-row .read-bar>i{display:block;height:100%;background:var(--rak-amber)}
.notice-row .read-bar>i.complete{background:var(--rak-success)}

/* Responsive */
@media(max-width:639px){
  .nav-links-desktop{display:none}
  .bottom-nav{display:flex}
  .container{padding-bottom:72px}
}
'''

def page(title, body, code=None, active=None):
    team = get_team(code) if code else None
    team_name = team['name'] if team else ''
    admin = is_admin(code) if code else False
    member = get_member(code) if code else ''

    # notification counts for member
    notifs = {'schedule': 0, 'notices': 0, 'fees': 0, 'orders': 0}
    if code and member and team:
        _c = get_db()
        today_s = datetime.now(JST).strftime('%Y-%m-%d')
        notifs['notices'] = _c.execute(
            'SELECT COUNT(*) FROM notices WHERE team_id=? AND id NOT IN (SELECT notice_id FROM reads WHERE member_name=?)',
            (team['id'], member)
        ).fetchone()[0]
        notifs['fees'] = _c.execute(
            '''SELECT COUNT(*) FROM fees f WHERE f.team_id=?
               AND NOT EXISTS (SELECT 1 FROM fee_payments WHERE fee_id=f.id AND member_name=? AND paid=1)''',
            (team['id'], member)
        ).fetchone()[0]
        notifs['orders'] = _c.execute(
            '''SELECT COUNT(*) FROM order_forms WHERE team_id=?
               AND id NOT IN (SELECT form_id FROM order_responses WHERE member_name=?)''',
            (team['id'], member)
        ).fetchone()[0]
        notifs['schedule'] = _c.execute(
            '''SELECT COUNT(*) FROM events WHERE team_id=? AND event_date>=?
               AND id NOT IN (SELECT event_id FROM rsvps WHERE member_name=?)''',
            (team['id'], today_s, member)
        ).fetchone()[0]
        _c.close()

    desktop_nav = ''
    bottom_nav = ''
    if code:
        home_dest = f'/t/{code}/admin/dash' if (admin and not member) else f'/t/{code}/home'

        # メイン5タブ（デスクトップ＋ボトムナビ共通）
        main_tabs = [
            ('home',     'home',     'ホーム',   home_dest),
            ('schedule', 'schedule', '予定',     f'/t/{code}/schedule'),
            ('notices',  'notices',  'お知らせ', f'/t/{code}/notices'),
            ('members',  'members',  'メンバー', f'/t/{code}/members'),
            ('fees',     'fees',     '集金',     f'/t/{code}/fees'),
        ]
        for key, icon_key, label, url in main_tabs:
            cls = 'active' if active == key else ''
            ico = ICONS[icon_key]
            cnt = notifs.get(key, 0)
            badge = f'<span class="nav-badge">{cnt}</span>' if cnt > 0 else ''
            desktop_nav += f'<a href="{url}" class="{cls}"><span class="nav-d-icon">{ico}</span>{label}</a>'
            bottom_nav += f'<a href="{url}" class="{cls}"><span class="nav-b-icon">{ico}</span><span>{label}</span>{badge}</a>'

        # 「その他▾」ドロップダウン
        overflow_common = [
            ('orders',   'orders',   '注文フォーム', f'/t/{code}/orders'),
            ('uniforms', 'uniforms', 'ユニフォーム', f'/t/{code}/uniforms'),
        ]
        overflow_admin = [
            ('ai',      'ai',      'AI文章',    f'/t/{code}/admin/ai'),
            ('ledger',  'ledger',  '会計',      f'/t/{code}/admin/ledger'),
            ('memo',    'memo',    'メモ',      f'/t/{code}/admin/memos'),
            ('contact', 'contact', '問い合わせ', f'/feedback?from={code}'),
            ('help',    'help',    '使い方',    f'/t/{code}/help'),
            ('plan',    'plan',    'プラン',    f'/t/{code}/upgrade'),
        ]
        overflow_all = overflow_common + (overflow_admin if admin else [])
        overflow_active = active in {k for k, _, _, _ in overflow_all}

        dd_items = ''
        for i, (key, icon_key, label, url) in enumerate(overflow_all):
            if admin and key == 'ai':
                dd_items += '<div class="nav-dd-sep"></div>'
            item_cls = 'active' if active == key else ''
            cnt = notifs.get(key, 0)
            badge_html = f'<span style="background:#ef4444;color:#fff;border-radius:999px;font-size:9px;padding:1px 5px;margin-left:auto">{cnt}</span>' if cnt else ''
            dd_items += f'<a href="{url}" class="{item_cls}"><span class="nav-d-icon">{ICONS[icon_key]}</span>{label}{badge_html}</a>'

        dd_btn_cls = 'active' if overflow_active else ''
        desktop_nav += (
            f'<div class="nav-dropdown">'
            f'<button class="nav-dd-btn {dd_btn_cls}" onclick="(function(b){{var m=b.nextElementSibling;m.classList.toggle(\'open\')}})(this)">その他&nbsp;▾</button>'
            f'<div class="nav-dd-menu">{dd_items}</div>'
            f'</div>'
        )

        if member and not admin:
            desktop_nav += f'<span style="font-size:12px;color:#888;padding:6px 8px">{_ICO_USER_SM} {member}</span>'

        if admin:
            bottom_nav = ''
        else:
            bottom_nav = f'<nav class="bottom-nav">{bottom_nav}</nav>'

    return render_template_string(f'''<!DOCTYPE html>
<html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}
{FONT}<title>{title} | Rak</title>
<style>{CSS}</style></head><body>
<nav class="nav">
  <a class="nav-logo" href="{"/t/"+code if code else "/"}">
    <div class="nav-icon">{NAV_MARK}</div>Rak
  </a>
  {f'<span class="nav-team">{team_name}</span>' if team_name else ''}
  <div class="nav-links-desktop">{desktop_nav}</div>
  {f'<a href="/t/{code}/help" style="margin-left:auto;width:30px;height:30px;border-radius:50%;background:#f1f4f9;display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:700;color:#888;text-decoration:none;flex-shrink:0">?</a>' if code and member and not admin else ''}
</nav>
<script>document.addEventListener('click',function(e){{document.querySelectorAll('.nav-dd-menu.open').forEach(function(m){{if(!m.parentElement.contains(e.target))m.classList.remove('open')}})}})</script>
{'<div style="background:#fffbeb;border-bottom:1px solid #f59e0b;padding:7px 16px;text-align:center;font-size:12px;font-weight:500;color:#92400e">Pro無料トライアル中 — 残り<strong style="color:#d97706">' + str(get_trial_days_left(team)) + '日</strong>　<a href="/t/' + code + '/upgrade" style="color:#d97706;font-weight:700;margin-left:6px">今すぐ継続 →</a></div>' if (code and admin and team and get_trial_days_left(team) is not None) else ''}
{body}
{bottom_nav}
{PWA_SW}
</body></html>''')


# ── PWA ───────────────────────────────────────────────────────────

_PWA_ICON_SVG = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">
<rect width="100" height="100" fill="#d97706"/>
<path d="M 18 82 L 18 18 L 46 18 C 60 18 66 28 66 38 C 66 48 60 56 46 56 L 18 56" stroke="white" stroke-width="10" stroke-linejoin="miter" fill="none"/>
<path d="M 40 56 L 58 82 L 92 16" stroke="white" stroke-width="10" stroke-linejoin="miter" fill="none"/>
</svg>'''

@app.route('/push/vapid-public-key')
def push_vapid_key():
    return jsonify({'publicKey': VAPID_PUBLIC_KEY})

@app.route('/t/<code>/push/subscribe', methods=['POST'])
def push_subscribe(code):
    team = get_team(code)
    if not team:
        return jsonify(error='not found'), 404
    member = get_member(code)
    if not member:
        return jsonify(error='not member'), 403
    data = request.get_json() or {}
    endpoint = data.get('endpoint', '')
    p256dh   = data.get('keys', {}).get('p256dh', '')
    auth     = data.get('keys', {}).get('auth', '')
    if not endpoint or not p256dh or not auth:
        return jsonify(error='invalid'), 400
    conn = get_db()
    conn.execute('''
        INSERT INTO push_subscriptions (id,team_id,member_name,endpoint,p256dh,auth,created_at)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(team_id,member_name,endpoint) DO UPDATE SET p256dh=excluded.p256dh, auth=excluded.auth
    ''', (new_id(), team['id'], member, endpoint, p256dh, auth, now_str()))
    conn.commit()
    conn.close()
    return jsonify(ok=True)

@app.route('/t/<code>/push/subscribe-native', methods=['POST'])
def push_subscribe_native(code):
    team = get_team(code)
    if not team:
        return jsonify(error='not found'), 404
    member = get_member(code)
    if not member:
        return jsonify(error='not member'), 403
    data = request.get_json() or {}
    token    = data.get('token', '')
    platform = data.get('platform', '')
    if not token or platform not in ('ios', 'android'):
        return jsonify(error='invalid'), 400
    conn = get_db()
    conn.execute('''
        INSERT INTO native_push_tokens (id,team_id,member_name,platform,token,created_at)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(team_id,member_name,token) DO UPDATE SET platform=excluded.platform
    ''', (new_id(), team['id'], member, platform, token, now_str()))
    conn.commit()
    conn.close()
    return jsonify(ok=True)

@app.route('/t/<code>/push/unsubscribe-native', methods=['POST'])
def push_unsubscribe_native(code):
    team = get_team(code)
    if not team:
        return jsonify(error='not found'), 404
    data = request.get_json() or {}
    token = data.get('token', '')
    conn = get_db()
    conn.execute('DELETE FROM native_push_tokens WHERE team_id=? AND token=?', (team['id'], token))
    conn.commit()
    conn.close()
    return jsonify(ok=True)

@app.route('/t/<code>/push/unsubscribe', methods=['POST'])
def push_unsubscribe(code):
    team = get_team(code)
    if not team:
        return jsonify(error='not found'), 404
    data = request.get_json() or {}
    endpoint = data.get('endpoint', '')
    conn = get_db()
    conn.execute('DELETE FROM push_subscriptions WHERE team_id=? AND endpoint=?', (team['id'], endpoint))
    conn.commit()
    conn.close()
    return jsonify(ok=True)

@app.route('/manifest.json')
def pwa_manifest():
    return jsonify({
        "name": "Rak",
        "short_name": "Rak",
        "description": "チーム運営の「めんどくさい」を、ぜんぶラクに。",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#d97706",
        "icons": [
            {"src": "/apple-touch-icon.png", "sizes": "1254x1254", "type": "image/png", "purpose": "any maskable"}
        ]
    })

@app.route('/icon.svg')
def pwa_icon():
    return Response(_PWA_ICON_SVG, mimetype='image/svg+xml')

@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
def apple_touch_icon():
    return send_file('static/apple-touch-icon.png', mimetype='image/png')

@app.route('/sw.js')
def service_worker():
    js = """const CACHE='rak-v2';
self.addEventListener('install',e=>{self.skipWaiting();});
self.addEventListener('activate',e=>{
  e.waitUntil(caches.keys().then(keys=>Promise.all(keys.filter(k=>k!==CACHE).map(k=>caches.delete(k)))));
  self.clients.claim();
});
self.addEventListener('fetch',e=>{
  if(e.request.method!=='GET')return;
  e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));
});
self.addEventListener('push',e=>{
  let d={title:'Rak',body:'新しいお知らせがあります',url:'/'};
  try{d=Object.assign(d,e.data.json());}catch(err){}
  e.waitUntil(self.registration.showNotification(d.title,{
    body:d.body,
    icon:'/icon.svg',
    badge:'/icon.svg',
    data:{url:d.url}
  }));
});
self.addEventListener('notificationclick',e=>{
  e.notification.close();
  const url=e.notification.data&&e.notification.data.url||'/';
  e.waitUntil(clients.matchAll({type:'window'}).then(cs=>{
    for(const c of cs){if(c.url===url&&'focus'in c)return c.focus();}
    if(clients.openWindow)return clients.openWindow(url);
  }));
});"""
    return Response(js, mimetype='application/javascript')


@app.route('/legal/terms')
def legal_terms():
    return terms_html()


@app.route('/legal/privacy')
def legal_privacy():
    return privacy_html()


@app.route('/legal/tokushoho')
def legal_tokushoho():
    return tokushoho_html()


# ── SEO ───────────────────────────────────────────────────────────

@app.route('/robots.txt')
def robots_txt():
    return app.response_class(
        'User-agent: *\nAllow: /\nDisallow: /t/\nSitemap: https://www.rakapp.jp/sitemap.xml\n',
        mimetype='text/plain'
    )

@app.route('/sitemap.xml')
def sitemap_xml():
    xml = '''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url><loc>https://www.rakapp.jp/</loc><changefreq>weekly</changefreq><priority>1.0</priority></url>
  <url><loc>https://www.rakapp.jp/create</loc><changefreq>monthly</changefreq><priority>0.8</priority></url>
</urlset>'''
    return app.response_class(xml, mimetype='application/xml')

# ── Home / Create ─────────────────────────────────────────────────

@app.route('/')
def home():
    log_lp_event('lp_view', src=request.args.get('utm_source', ''))
    join_error = request.args.get('error', '')
    join_code  = request.args.get('code', '')
    GOOGLE_VERIFY_TAG = f'<meta name="google-site-verification" content="{GOOGLE_SITE_VERIFICATION}">' if GOOGLE_SITE_VERIFICATION else ''
    LP_LOGO = '<svg width="30" height="27" viewBox="0 0 110 100" fill="none"><path d="M 22 84 L 22 16 L 50 16 C 64 16 70 26 70 36 C 70 46 64 54 50 54 L 22 54" stroke="#d97706" stroke-width="11" stroke-linejoin="miter" fill="none"/><path d="M 44 54 L 62 84 L 102 18" stroke="#d97706" stroke-width="11" stroke-linejoin="miter" fill="none"/></svg>'
    LP_LOGO_W = '<svg width="24" height="22" viewBox="0 0 110 100" fill="none"><path d="M 22 84 L 22 16 L 50 16 C 64 16 70 26 70 36 C 70 46 64 54 50 54 L 22 54" stroke="#d97706" stroke-width="11" stroke-linejoin="miter" fill="none"/><path d="M 44 54 L 62 84 L 102 18" stroke="#d97706" stroke-width="11" stroke-linejoin="miter" fill="none"/></svg>'
    _sv = 'viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"'
    IC_CAL = f'<svg {_sv}><rect x="3" y="4.5" width="18" height="16" rx="2.5"/><path d="M3 9h18M8 2.5v4M16 2.5v4"/></svg>'
    IC_MEGA = f'<svg {_sv}><path d="M3 10.5v3a1 1 0 0 0 1 1h2l5 4V5.5l-5 4H4a1 1 0 0 0-1 1Z"/><path d="M15.5 9a4.5 4.5 0 0 1 0 6"/><path d="M18.5 6.5a8 8 0 0 1 0 11"/></svg>'
    IC_USERS = f'<svg {_sv}><circle cx="9" cy="8" r="3.3"/><path d="M3.5 19a5.5 5.5 0 0 1 11 0"/><path d="M15.5 5.2a3 3 0 0 1 0 5.6"/><path d="M17.8 19a5.3 5.3 0 0 0-2.6-4.5"/></svg>'
    IC_LINK = f'<svg {_sv}><path d="M9.5 14.5a3.5 3.5 0 0 0 5 0l3-3a3.5 3.5 0 0 0-5-5l-1.2 1.2"/><path d="M14.5 9.5a3.5 3.5 0 0 0-5 0l-3 3a3.5 3.5 0 0 0 5 5l1.2-1.2"/></svg>'
    IC_YEN = f'<svg {_sv}><circle cx="12" cy="12" r="8.5"/><path d="M9 8l3 4 3-4M12 12v5M9.5 13.5h5M9.5 15.7h5"/></svg>'
    IC_CLIP = f'<svg {_sv}><path d="M9 4.5H7a2 2 0 0 0-2 2V20a2 2 0 0 0 2 2h10a2 2 0 0 0 2-2V6.5a2 2 0 0 0-2-2h-2"/><rect x="9" y="2.5" width="6" height="4" rx="1.5"/><path d="M8.5 12h7M8.5 15.5h5"/></svg>'
    IC_CHART = f'<svg {_sv}><path d="M4 4v16h16"/><rect x="7.5" y="12" width="2.8" height="5" rx="1"/><rect x="12" y="8.5" width="2.8" height="8.5" rx="1"/><rect x="16.5" y="14" width="2.8" height="3" rx="1"/></svg>'
    IC_AI = f'<svg {_sv}><path d="M12 4l1.7 4.6L18 10.3l-4.3 1.7L12 16.5l-1.7-4.5L6 10.3l4.3-1.7L12 4Z"/><path d="M18.5 14.5l.7 1.9 1.9.7-1.9.7-.7 1.9-.7-1.9-1.9-.7 1.9-.7.7-1.9Z"/></svg>'
    IC_HOME = f'<svg {_sv}><path d="M4 11l8-7 8 7"/><path d="M6 9.5V20h12V9.5"/></svg>'
    IC_BELL = f'<svg {_sv}><path d="M6 9a6 6 0 0 1 12 0c0 5 2 6 2 6H4s2-1 2-6Z"/><path d="M10 19a2 2 0 0 0 4 0"/></svg>'
    IC_BALL = f'<svg {_sv}><circle cx="12" cy="12" r="8.5"/><path d="M12 7.2l3.2 2.4-1.2 3.8h-4L8.8 9.6 12 7.2Z"/></svg>'
    return render_template_string(f'''<!DOCTYPE html>
<html lang="ja"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}
<title>Rak — チーム運営をラクにするアプリ | スケジュール・出欠・集金をひとつに</title>
<meta name="description" content="Rakはスポーツチーム・部活・サークルの運営をまるごと効率化するチーム管理アプリです。スケジュール共有・出欠確認・集金管理・AI文章作成をひとつにまとめて、事務作業をゼロに近づけます。無料から始められます。">
<meta name="keywords" content="チーム管理アプリ,スポーツチーム,スケジュール管理,出欠確認,集金管理,部活管理,少年野球,ソフトボール,サッカー,バスケ">
<link rel="canonical" href="https://www.rakapp.jp/">
{GOOGLE_VERIFY_TAG}
<meta property="og:type" content="website">
<meta property="og:url" content="https://www.rakapp.jp/">
<meta property="og:title" content="Rak — チーム運営をラクにするアプリ">
<meta property="og:description" content="スケジュール・出欠・集金・AIをひとつに。スポーツチーム・部活・サークルの事務作業をまるごと効率化。無料から始められます。">
<meta property="og:image" content="https://www.rakapp.jp/static/ogp.png">
<meta property="og:site_name" content="Rak">
<meta property="og:locale" content="ja_JP">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="Rak — チーム運営をラクにするアプリ">
<meta name="twitter:description" content="スケジュール・出欠・集金・AIをひとつに。無料から始められます。">
<meta name="twitter:image" content="https://www.rakapp.jp/static/ogp.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;600;700;800;900&family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
:root{{
  --rak-black:#111111;--rak-graphite:#525252;--rak-mute:#8a8a8a;
  --rak-line:#e7e7e7;--rak-line-soft:#efefef;
  --rak-bg-soft:#fffaf2;
  --rak-amber:#f59e0b;--rak-amber-deep:#d97706;--rak-amber-tint:#fef3c7;
  --rak-sky:#0ea5e9;--rak-sky-tint:#e0f2fe;
  --font-jp:"Noto Sans JP","Hiragino Sans",system-ui,sans-serif;
  --font-num:"Inter","Noto Sans JP",system-ui,sans-serif;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{font-family:var(--font-jp);color:var(--rak-black);background:#fff;line-height:1.7;-webkit-font-smoothing:antialiased;overflow-x:hidden}}
h1,h2,h3,.sec-title,.sec-sub,.fcard-title,.hero-badge{{word-break:keep-all}}
a{{color:inherit;text-decoration:none}}
button{{font-family:inherit;cursor:pointer}}

/* Nav */
.lp-nav{{background:rgba(255,255,255,.85);backdrop-filter:saturate(180%) blur(12px);border-bottom:1px solid var(--rak-line-soft);padding:0 24px;height:58px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50}}
.lp-logo{{display:flex;align-items:center;gap:10px;font-weight:900;font-size:20px;letter-spacing:-0.02em}}
.lp-nav-links{{display:flex;gap:8px;align-items:center}}
.lp-nav-links a{{font-size:14px;color:var(--rak-graphite);padding:7px 14px;border-radius:8px;font-weight:600}}
.lp-nav-links a:hover{{background:var(--rak-amber-tint);color:var(--rak-amber-deep)}}
.btn-nav{{background:var(--rak-amber)!important;color:#fff!important;border-radius:999px;padding:9px 20px!important;font-weight:800!important;box-shadow:0 4px 14px rgba(245,158,11,.35)}}
.btn-nav:hover{{background:var(--rak-amber-deep)!important}}

/* Hero — bright, app-forward, two columns */
.hero{{position:relative;background:
   radial-gradient(900px 500px at 88% -10%,rgba(245,158,11,.18),transparent 60%),
   radial-gradient(700px 500px at -5% 20%,rgba(14,165,233,.12),transparent 55%),
   linear-gradient(170deg,#fffdf8 0%,#fff5e6 100%);
   padding:64px 24px 72px;overflow:hidden}}
.hero-inner{{max-width:1040px;margin:0 auto;display:grid;grid-template-columns:1.05fr .95fr;gap:40px;align-items:center}}
.hero-l{{text-align:left}}
.hero-badge{{display:inline-flex;align-items:center;gap:7px;background:#fff;color:var(--rak-amber-deep);font-size:12.5px;font-weight:800;padding:7px 15px;border-radius:999px;margin-bottom:22px;letter-spacing:.02em;box-shadow:0 3px 12px rgba(245,158,11,.16)}}
.hero-badge .dot{{width:7px;height:7px;border-radius:50%;background:var(--rak-amber);box-shadow:0 0 0 4px rgba(245,158,11,.18)}}
.hero h1{{font-size:clamp(32px,5.2vw,54px);font-weight:900;line-height:1.16;margin-bottom:20px;color:#0f172a;letter-spacing:-0.03em}}
.highlight{{position:relative;display:inline-block}}
.highlight::before{{content:"";position:absolute;left:-2px;right:-2px;bottom:3px;height:14px;background:var(--rak-amber-tint);z-index:0;border-radius:3px}}
.highlight>span{{position:relative;z-index:1}}
.hero p.lead{{font-size:16.5px;line-height:1.8;color:var(--rak-graphite);max-width:460px;margin:0 0 30px}}
.hero-btns{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:14px}}
.btn-primary{{background:var(--rak-amber);color:#fff;padding:15px 30px;border-radius:999px;font-size:15.5px;font-weight:800;display:inline-flex;align-items:center;gap:8px;transition:transform .12s,box-shadow .12s;box-shadow:0 8px 22px rgba(245,158,11,.4)}}
.btn-primary:hover{{background:var(--rak-amber-deep);color:#fff;transform:translateY(-2px);box-shadow:0 12px 26px rgba(245,158,11,.5)}}
.btn-ghost{{color:#0f172a;background:#fff;padding:14px 24px;border-radius:999px;font-size:15px;font-weight:700;display:inline-flex;align-items:center;border:1.5px solid #f1e3c6;box-shadow:0 2px 8px rgba(0,0,0,.04)}}
.btn-ghost:hover{{background:var(--rak-amber-tint)}}
.hero-note{{font-size:12.5px;color:var(--rak-mute);margin-bottom:26px;display:flex;align-items:center;gap:6px}}
.hero-note b{{color:var(--rak-amber-deep);font-weight:800}}

/* Code input */
.code-wrap{{max-width:380px}}
.code-wrap .lbl{{font-size:12px;color:var(--rak-mute);margin-bottom:9px;font-weight:600}}
.code-bar{{display:flex;gap:8px;background:#fff;border:1.5px solid var(--rak-line);border-radius:999px;padding:6px 6px 6px 18px;box-shadow:0 4px 16px rgba(0,0,0,.06)}}
.code-bar input{{flex:1;border:none;outline:none;font-size:15px;font-weight:700;font-family:var(--font-num);letter-spacing:.1em;text-transform:uppercase;background:transparent;min-width:0}}
.code-bar button{{background:#0f172a;color:#fff;border:none;border-radius:999px;padding:10px 20px;font-size:14px;font-weight:800}}
.code-bar button:hover{{background:#333}}

/* Phone mockup */
.hero-r{{display:flex;justify-content:center;perspective:1400px}}
.phone{{position:relative;width:284px;height:580px;background:#0f172a;border-radius:44px;padding:11px;box-shadow:0 30px 70px rgba(15,23,42,.32),0 0 0 2px rgba(255,255,255,.4) inset;transform:rotateY(-9deg) rotateX(3deg);transition:transform .4s}}
.hero-r:hover .phone{{transform:rotateY(0) rotateX(0)}}
.phone::before{{content:"";position:absolute;top:16px;left:50%;transform:translateX(-50%);width:104px;height:24px;background:#0f172a;border-radius:0 0 16px 16px;z-index:5}}
.scr{{width:100%;height:100%;background:#f6f7f9;border-radius:34px;overflow:hidden;display:flex;flex-direction:column;position:relative}}
.scr-status{{display:flex;justify-content:space-between;align-items:center;padding:14px 22px 6px;font-size:12px;font-weight:700;font-family:var(--font-num);color:#0f172a}}
.scr-head{{background:linear-gradient(135deg,#f59e0b,#d97706);color:#fff;padding:12px 18px 16px;display:flex;align-items:center;justify-content:space-between}}
.scr-team{{display:flex;align-items:center;gap:9px}}
.scr-team .em{{width:34px;height:34px;border-radius:11px;background:rgba(255,255,255,.22);display:flex;align-items:center;justify-content:center;color:#fff}}
.scr-team .em svg{{width:18px;height:18px}}
.scr-team .nm{{font-size:15px;font-weight:800;line-height:1.2}}
.scr-team .sub{{font-size:10.5px;opacity:.85;font-weight:600}}
.scr-bell{{width:32px;height:32px;border-radius:50%;background:rgba(255,255,255,.2);display:flex;align-items:center;justify-content:center;color:#fff;position:relative}}
.scr-bell svg{{width:16px;height:16px}}
.scr-bell::after{{content:"";position:absolute;top:7px;right:8px;width:7px;height:7px;background:#ef4444;border-radius:50%;border:1.5px solid #e5890b}}
.scr-body{{flex:1;padding:14px 14px 0;overflow:hidden}}
.scr-sec{{font-size:11px;font-weight:800;color:#94a3b8;letter-spacing:.06em;margin:4px 2px 8px}}
.mc{{background:#fff;border-radius:15px;padding:13px 14px;box-shadow:0 2px 10px rgba(15,23,42,.05);margin-bottom:11px}}
.mc-ev{{display:flex;gap:11px;align-items:center}}
.mc-date{{width:46px;height:46px;border-radius:12px;background:var(--rak-amber-tint);color:var(--rak-amber-deep);display:flex;flex-direction:column;align-items:center;justify-content:center;flex-shrink:0}}
.mc-date .d{{font-size:17px;font-weight:900;font-family:var(--font-num);line-height:1}}
.mc-date .m{{font-size:9px;font-weight:700}}
.mc-ev-t{{font-size:13.5px;font-weight:800;color:#0f172a;margin-bottom:3px}}
.mc-ev-s{{font-size:11px;color:#64748b;font-weight:600}}
.mc-tag{{display:inline-block;background:#dcfce7;color:#16a34a;font-size:9.5px;font-weight:800;padding:2px 7px;border-radius:6px;margin-top:5px}}
.mc-msg-t{{font-size:12.5px;font-weight:800;color:#0f172a;margin-bottom:4px}}
.mc-msg-b{{font-size:11px;color:#64748b;line-height:1.6;font-weight:500}}
.mc-read{{display:flex;align-items:center;gap:5px;margin-top:9px;font-size:10.5px;font-weight:800;color:var(--rak-amber-deep)}}
.mc-read .bar{{flex:1;height:5px;background:#f1f5f9;border-radius:3px;overflow:hidden}}
.mc-read .bar i{{display:block;height:100%;width:93%;background:var(--rak-amber);border-radius:3px}}
.mc-pay{{display:flex;align-items:center;justify-content:space-between}}
.mc-pay .l{{font-size:12.5px;font-weight:800;color:#0f172a}}
.mc-pay .amt{{font-size:15px;font-weight:900;font-family:var(--font-num);color:#16a34a}}
.scr-nav{{display:flex;justify-content:space-around;align-items:center;background:#fff;border-top:1px solid #eef1f4;padding:9px 6px 12px}}
.scr-nav div{{display:flex;flex-direction:column;align-items:center;gap:3px;font-size:9px;font-weight:700;color:#cbd5e1}}
.scr-nav div.on{{color:var(--rak-amber-deep)}}
.scr-nav .ic{{display:inline-flex}}
.scr-nav .ic svg{{width:19px;height:19px}}

/* Logo strip / catch */
.catch{{background:#fff;padding:30px 24px;border-bottom:1px solid var(--rak-line-soft)}}
.catch-in{{max-width:880px;margin:0 auto;display:flex;flex-wrap:wrap;justify-content:center;gap:14px}}
.chip{{display:inline-flex;align-items:center;gap:7px;background:var(--rak-bg-soft);border:1px solid #f3e7cf;color:#7c5e2a;font-size:13px;font-weight:700;padding:9px 16px;border-radius:999px;white-space:nowrap}}
.chip .d{{width:7px;height:7px;border-radius:50%;background:var(--rak-amber)}}

/* Features */
.features{{padding:76px 24px;background:#fff}}
.sec-label{{font-size:11px;font-weight:800;letter-spacing:.16em;color:var(--rak-amber);margin-bottom:12px;display:block}}
.sec-title{{font-size:clamp(24px,4.2vw,32px);font-weight:900;letter-spacing:-0.025em;line-height:1.22;margin-bottom:8px;color:#0f172a}}
.sec-sub{{font-size:14.5px;color:#64748b;margin-bottom:40px}}
.feat-grid{{max-width:900px;margin:0 auto;display:grid;grid-template-columns:repeat(3,1fr);gap:14px}}
.fcard{{border-radius:16px;padding:22px 14px 18px;display:flex;flex-direction:column;align-items:center;text-align:center;gap:10px;background:#fff;border:1px solid var(--rak-line);box-shadow:0 2px 10px rgba(0,0,0,.03);transition:transform .14s,box-shadow .14s}}
.fcard:hover{{transform:translateY(-3px);box-shadow:0 10px 24px rgba(0,0,0,.08)}}
.fcard-pro{{background:#fffdf8;border:1px solid rgba(245,158,11,.3);border-top:3px solid var(--rak-amber);box-shadow:0 4px 16px rgba(245,158,11,.08)}}
.fcard-pro:hover{{box-shadow:0 10px 28px rgba(245,158,11,.18);transform:translateY(-3px)}}
.fcard-pro .fcard-title{{color:#0f172a}}
.fcard-pro .fcard-desc{{color:#525252}}
.fcard-ic{{display:inline-flex;color:var(--rak-amber);background:var(--rak-amber-tint);border-radius:14px;padding:12px}}
.fcard-ic svg{{width:28px;height:28px}}
.fcard-title{{font-size:14px;font-weight:800;line-height:1.4;color:#0f172a}}
.fcard-desc{{font-size:11.5px;color:#64748b;line-height:1.6;min-height:2.5em}}

/* Pricing */
.pricing{{padding:76px 24px;background:linear-gradient(180deg,#fffaf2,#fff5e6)}}
.plan-grid{{max-width:780px;margin:0 auto;display:grid;gap:16px}}
.plan-card{{background:#fff;border:1px solid var(--rak-line);border-radius:18px;padding:26px;box-shadow:0 4px 18px rgba(0,0,0,.04)}}
.plan-card.dark{{background:linear-gradient(160deg,#1c1c1c,#000);color:#fff;border:none;position:relative;box-shadow:0 16px 40px rgba(15,23,42,.18)}}
.plan-name{{font-size:13px;font-weight:800;color:var(--rak-mute);letter-spacing:.08em;text-transform:uppercase;margin-bottom:10px}}
.plan-card.dark .plan-name{{color:#aaa}}
.plan-price{{display:flex;align-items:baseline;gap:4px;margin-bottom:16px}}
.plan-price .num{{font-family:var(--font-num);font-size:40px;font-weight:900;letter-spacing:-0.03em}}
.plan-price .per{{color:var(--rak-mute);font-size:13px}}
.plan-card.dark .plan-price .per{{color:#aaa}}
.plan-items{{font-size:13px;color:var(--rak-graphite);line-height:1.95;margin-bottom:20px}}
.plan-card.dark .plan-items{{color:var(--rak-amber)}}
.plan-card.dark .plan-items .acc{{color:var(--rak-amber)}}
.plan-rec{{position:absolute;top:-11px;right:18px;background:var(--rak-amber);color:#fff;font-size:10px;font-weight:800;padding:5px 12px;border-radius:999px;letter-spacing:.05em;box-shadow:0 4px 12px rgba(245,158,11,.4)}}
.plan-btn-w{{display:block;text-align:center;padding:14px;border-radius:999px;font-weight:800;font-size:14px;background:#fff;color:var(--rak-amber-deep);border:1.5px solid var(--rak-amber)}}
.plan-btn-w:hover{{background:var(--rak-amber-tint)}}
.plan-btn-b{{display:block;text-align:center;padding:14px;border-radius:999px;font-weight:800;font-size:14px;background:var(--rak-amber);color:#fff;border:none;box-shadow:0 8px 20px rgba(245,158,11,.4)}}
.plan-btn-b:hover{{background:var(--rak-amber-deep)}}
@media(min-width:560px){{.plan-grid{{grid-template-columns:1fr 1fr}}}}

/* Footer CTA */
.cta-sec{{background:
   radial-gradient(600px 300px at 80% 0%,rgba(245,158,11,.25),transparent 60%),
   #0f172a;padding:80px 24px;text-align:center;color:#fff}}
.cta-sec h2{{font-size:clamp(26px,4.5vw,38px);font-weight:900;margin-bottom:14px;letter-spacing:-0.025em}}
.cta-sec p{{font-size:15px;color:#94a3b8;margin-bottom:30px}}
.btn-amber-solid{{background:var(--rak-amber);color:#fff;padding:16px 40px;border-radius:999px;font-size:16px;font-weight:800;display:inline-block;box-shadow:0 10px 30px rgba(245,158,11,.45)}}
.btn-amber-solid:hover{{background:var(--rak-amber-deep);color:#fff;transform:translateY(-2px)}}
footer{{background:#0f172a;border-top:1px solid #1e293b;color:#475569;padding:24px;text-align:center;font-size:12px}}
.footer-links{{display:flex;gap:20px;justify-content:center;margin-bottom:10px;flex-wrap:wrap}}
footer a{{color:#475569}}
footer a:hover{{color:#94a3b8}}
.footer-logo{{display:flex;align-items:center;gap:8px;justify-content:center;margin-bottom:12px;font-weight:700;font-size:16px;color:#475569}}

@media(max-width:840px){{
  .hero-inner{{grid-template-columns:1fr;gap:36px;text-align:center;max-width:440px}}
  .hero-l{{text-align:center}}
  .hero p.lead{{margin-left:auto;margin-right:auto}}
  .hero-btns,.hero-note{{justify-content:center}}
  .code-wrap{{margin:0 auto}}
  .phone{{transform:none}}
  .hero-r{{order:-1}}
}}
@media(max-width:600px){{
  .hero{{padding:48px 20px 56px}}
  .features,.pricing{{padding:56px 20px}}
  .cta-sec{{padding:60px 20px}}
  .lp-nav-links .hide-sp{{display:none}}
  .feat-grid{{grid-template-columns:repeat(2,1fr)}}
}}
</style>
</head><body>

<nav class="lp-nav">
  <a class="lp-logo" href="/">
    {LP_LOGO}Rak
  </a>
  <div class="lp-nav-links">
    <a href="#features" class="hide-sp">機能</a>
    <a href="#pricing" class="hide-sp">料金</a>
    <a href="/login">ログイン</a>
  </div>
</nav>

<section class="hero">
  <div class="hero-inner">
    <div class="hero-l">
      <div class="hero-badge"><span class="dot"></span>部活・チーム・サークルの運営担当へ</div>
      <h1><span style="white-space:nowrap">メンバーは<span style="color:var(--rak-amber)">今のまま</span>。</span><br><span style="white-space:nowrap">あなたの<span class="highlight"><span>事務</span></span>だけ、</span><span style="white-space:nowrap">ラクに。</span></h1>
      <p class="lead">出欠・集金・予定づくり——運営担当の「めんどくさい」をRakが引き受けます。メンバーはアプリも登録も不要。<br>連絡アプリに届くリンクを開いて、タップで答えるだけです。</p>
      <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:10px">
        <div style="flex:1;min-width:180px;background:#fff;border:2px solid var(--rak-amber);border-radius:16px;padding:16px;text-align:center">
          <div style="font-size:11px;font-weight:800;color:var(--rak-amber-deep);margin-bottom:4px">PRO</div>
          <div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:2px">14日間無料トライアル</div>
          <div style="font-size:11px;color:#888;margin-bottom:4px">その後 ¥980/月・いつでも解約可</div>
          <div style="display:inline-block;font-size:11px;font-weight:700;color:#16a34a;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:2px 8px;margin-bottom:10px">クレカ不要</div>
          <a href="/create?intent=pro&src=fcard" class="btn-primary" style="display:block;font-size:13px;padding:10px 0">試してみる →</a>
        </div>
        <div style="flex:1;min-width:180px;background:#fff;border:2px solid #e2e8f0;border-radius:16px;padding:16px;text-align:center">
          <div style="font-size:11px;font-weight:800;color:#64748b;margin-bottom:4px">FREE</div>
          <div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:2px">無料プラン</div>
          <div style="font-size:11px;color:#888;margin-bottom:4px">基本機能のみ</div>
          <div style="display:inline-block;font-size:11px;font-weight:700;color:#16a34a;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:2px 8px;margin-bottom:10px">クレカ不要</div>
          <a href="/create?src=fcard_free" class="btn-ghost" style="display:block;font-size:13px;padding:10px 0">無料で始める</a>
        </div>
      </div>
      <p class="hero-note"><b>メールとパスワードだけ</b>・30秒で開始</p>
    </div>
    <div class="hero-r">
      <div class="phone"><div class="scr">
        <div class="scr-status"><span>9:41</span><span>● ● ● ▮</span></div>
        <div style="background:linear-gradient(135deg,#d97706,#f59e0b);color:#fff;text-align:center;padding:16px 12px 14px">
          <div style="font-size:11px;opacity:.9">FC サンライズ</div>
          <div style="font-size:17px;font-weight:800;margin-top:2px">出欠の回答</div>
          <div style="display:inline-block;font-size:10px;background:rgba(255,255,255,.22);border-radius:20px;padding:3px 10px;margin-top:6px">登録不要・タップで回答</div>
        </div>
        <div style="padding:12px;background:#f7f8fa;flex:1">
          <div style="font-size:10px;color:#9ca3af;text-align:center;margin-bottom:10px">田中 さんとして回答中</div>
          <div style="background:#fff;border-radius:12px;padding:12px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,.06)">
            <div style="font-size:13px;font-weight:700;color:#111">練習試合 vs FCさくら</div>
            <div style="font-size:11px;color:#6b7280;margin-top:2px">6月20日（土）9:00　市民グラウンド</div>
            <div style="display:flex;gap:8px;margin-top:10px">
              <div style="flex:1;text-align:center;padding:9px;border-radius:8px;border:2px solid #16a34a;background:#f0fdf4;color:#16a34a;font-size:12px;font-weight:700">✓ 出席</div>
              <div style="flex:1;text-align:center;padding:9px;border-radius:8px;border:2px solid #e5e7eb;color:#9ca3af;font-size:12px;font-weight:700">欠席</div>
            </div>
          </div>
          <div style="background:#fff;border-radius:12px;padding:12px;box-shadow:0 1px 4px rgba(0,0,0,.06)">
            <div style="font-size:13px;font-weight:700;color:#111">通常練習</div>
            <div style="font-size:11px;color:#6b7280;margin-top:2px">6月23日（火）19:00　市民体育館</div>
            <div style="display:flex;gap:8px;margin-top:10px">
              <div style="flex:1;text-align:center;padding:9px;border-radius:8px;border:2px solid #e5e7eb;color:#9ca3af;font-size:12px;font-weight:700">出席</div>
              <div style="flex:1;text-align:center;padding:9px;border-radius:8px;border:2px solid #dc2626;background:#fef2f2;color:#dc2626;font-size:12px;font-weight:700">✓ 欠席</div>
            </div>
          </div>
        </div>
      </div></div>
    </div>
  </div>
</section>

<section class="catch">
  <div class="catch-in">
    <span class="chip"><span class="d"></span>メンバーは登録不要</span>
    <span class="chip"><span class="d"></span>いつもの連絡アプリで完結</span>
    <span class="chip"><span class="d"></span>出欠は自動で集計</span>
    <span class="chip"><span class="d"></span>事務作業がゼロに近づく</span>
  </div>
</section>

<section style="padding:64px 24px;background:#fffdf7">
  <div style="max-width:760px;margin:0 auto">
    <div style="text-align:center;margin-bottom:36px">
      <span class="sec-label">HOW IT WORKS</span>
      <div class="sec-title" style="margin-top:6px">使い方は、3ステップ。</div>
      <div class="sec-sub">アプリを入れるのは、あなただけ。</div>
    </div>
    <div style="display:grid;gap:16px">
      <div style="display:flex;gap:16px;align-items:flex-start;background:#fff;border-radius:14px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.05)">
        <div style="flex:0 0 auto;width:36px;height:36px;border-radius:50%;background:#d97706;color:#fff;font-weight:800;display:flex;align-items:center;justify-content:center">1</div>
        <div><div style="font-weight:700;font-size:16px">Rakで作る</div><div style="font-size:14px;color:#6b7280;margin-top:3px">出欠・予定・お知らせを、アプリでサッと作成。AIが文章や予定づくりも手伝います。</div></div>
      </div>
      <div style="display:flex;gap:16px;align-items:flex-start;background:#fff;border-radius:14px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.05)">
        <div style="flex:0 0 auto;width:36px;height:36px;border-radius:50%;background:#d97706;color:#fff;font-weight:800;display:flex;align-items:center;justify-content:center">2</div>
        <div><div style="font-weight:700;font-size:16px">リンクを連絡アプリに貼る</div><div style="font-size:14px;color:#6b7280;margin-top:3px">コピーして、いつもの連絡アプリのグループへ。新しい連絡ツールに乗り換えさせる必要はありません。</div></div>
      </div>
      <div style="display:flex;gap:16px;align-items:flex-start;background:#fff;border-radius:14px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.05)">
        <div style="flex:0 0 auto;width:36px;height:36px;border-radius:50%;background:#d97706;color:#fff;font-weight:800;display:flex;align-items:center;justify-content:center">3</div>
        <div><div style="font-weight:700;font-size:16px">メンバーはタップで回答</div><div style="font-size:14px;color:#6b7280;margin-top:3px">リンクを開いて出席/欠席を押すだけ。登録もアプリも不要。回答はRakで自動集計されます。</div></div>
      </div>
    </div>
  </div>
</section>

<section class="features" id="features">
  <div style="max-width:900px;margin:0 auto">
    <span class="sec-label">FEATURES</span>
    <div class="sec-title">あなたの事務作業、<br>ぜんぶここで。</div>
    <div class="sec-sub">作るのはアプリで。配るのはいつもの連絡アプリで。集計はRakが自動で。</div>
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:14px">
      <span style="font-size:13px;font-weight:800;letter-spacing:.06em;color:#16a34a;white-space:nowrap">FREE　無料で使える</span>
      <div style="flex:1;height:1px;background:#bbf7d0"></div>
    </div>
    <div class="feat-grid" style="margin-bottom:32px">
      <div class="fcard">
        <span class="fcard-ic">{IC_CAL}</span>
        <div class="fcard-title">スケジュール管理</div>
        <div class="fcard-desc">練習・試合の日程をチームで共有</div>
      </div>
      <div class="fcard">
        <span class="fcard-ic">{IC_MEGA}</span>
        <div class="fcard-title">お知らせ作成</div>
        <div class="fcard-desc">連絡文を作成。コピーして連絡アプリに貼るだけ</div>
      </div>
      <div class="fcard">
        <span class="fcard-ic">{IC_USERS}</span>
        <div class="fcard-title">メンバー管理</div>
        <div class="fcard-desc">メンバー情報を一元管理</div>
      </div>
      <div class="fcard">
        <span class="fcard-ic">{IC_LINK}</span>
        <div class="fcard-title">共有リンク（登録不要）</div>
        <div class="fcard-desc">出欠・予定のリンクを連絡アプリに貼るだけ。メンバーはタップで回答</div>
      </div>
      <div class="fcard">
        <span class="fcard-ic">{IC_CLIP}</span>
        <div class="fcard-title">メモ・資料保管</div>
        <div class="fcard-desc">運営メモに写真やファイルを添付して保管。大会要項・名簿などをチームの保管庫に</div>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:14px">
      <span style="font-size:13px;font-weight:800;letter-spacing:.06em;color:var(--rak-amber-deep);white-space:nowrap">PRO　アップグレードで解放</span>
      <div style="flex:1;height:1px;background:#fde68a"></div>
    </div>
    <div class="feat-grid">
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_YEN}</span>
        <div class="fcard-title">集金・費用管理</div>
        <div class="fcard-desc">会費や遠征費の集金を自動化</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_CLIP}</span>
        <div class="fcard-title">注文フォーム</div>
        <div class="fcard-desc">弁当・グッズの注文をまとめて管理</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_CHART}</span>
        <div class="fcard-title">会計・収支記録</div>
        <div class="fcard-desc">収支を記録して会計報告を楽に</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_AI}</span>
        <div class="fcard-title">AI文章生成</div>
        <div class="fcard-desc">練習メモ→学校提出の<strong>活動報告書</strong>を自動生成。お知らせ・保護者連絡も対応</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_CAL}</span>
        <div class="fcard-title">AIスケジュール自動生成</div>
        <div class="fcard-desc">制約を入力するとAIが仮予定を自動作成</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_USERS}</span>
        <div class="fcard-title">ユニフォーム管理</div>
        <div class="fcard-desc">サイズ・枚数・背番号を一括管理。受取チェックまで完結</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_CLIP}</span>
        <div class="fcard-title">アンケート</div>
        <div class="fcard-desc">選択・テキスト形式のアンケートを作成・集計</div>
      </div>
      <div class="fcard fcard-pro">
        <span class="fcard-ic">{IC_CHART}</span>
        <div class="fcard-title">Excelエクスポート</div>
        <div class="fcard-desc">スケジュール・集金データをExcelに出力</div>
      </div>
    </div>
  </div>
</section>

<section style="padding:72px 24px;background:#fff">
  <div style="max-width:780px;margin:0 auto">
    <span class="sec-label">WHY RAK</span>
    <div class="sec-title">バラバラのツールを<br>卒業する理由。</div>
    <div class="sec-sub" style="margin-bottom:40px">「とりあえず連絡アプリ」では情報が流れる。「とりあえず表計算」では手間が増える。</div>

    <div style="overflow-x:auto;-webkit-overflow-scrolling:touch">
      <table style="width:100%;border-collapse:collapse;min-width:480px">
        <thead>
          <tr>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#9ca3af;font-weight:600;border-bottom:2px solid #f3f4f6;width:40%"></th>
            <th style="padding:12px 16px;text-align:center;font-size:13px;color:#6b7280;font-weight:600;border-bottom:2px solid #f3f4f6">従来の方法</th>
            <th style="padding:12px 16px;text-align:center;font-size:14px;font-weight:800;color:#fff;background:#d97706;border-bottom:2px solid #d97706;border-radius:10px 10px 0 0">Rak</th>
          </tr>
        </thead>
        <tbody>
          <tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">スケジュール共有</td>
            <td style="padding:14px 16px;text-align:center;font-size:13px;color:#9ca3af">投稿が流れて埋もれる</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓</td>
          </tr>
          <tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">出欠確認</td>
            <td style="padding:14px 16px;text-align:center;font-size:13px;color:#9ca3af">リアクション集計・手入力</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓</td>
          </tr>
          <tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">集金・費用管理</td>
            <td style="padding:14px 16px;text-align:center;font-size:13px;color:#9ca3af">表計算で手入力・追いかけ</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓</td>
          </tr>
          <tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">活動報告書・お知らせ作成</td>
            <td style="padding:14px 16px;text-align:center;font-size:13px;color:#9ca3af">毎回一から手書き</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓ <span style="font-size:10px;font-weight:700">AI</span></td>
          </tr>
          <tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">スケジュール自動作成</td>
            <td style="padding:14px 16px;text-align:center;font-size:20px;color:#e5e7eb">—</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓ <span style="font-size:10px;font-weight:700">AI</span></td>
          </tr>
          <tr>
            <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">すべてひとつで完結</td>
            <td style="padding:14px 16px;text-align:center;font-size:20px;color:#e5e7eb">—</td>
            <td style="padding:14px 16px;text-align:center;background:#fffbeb;border-radius:0 0 10px 10px;font-size:18px;color:#d97706">✓</td>
          </tr>
        </tbody>
      </table>
    </div>

    <!-- 他社A vs Rak -->
    <div style="margin-top:48px">
      <div style="font-size:12px;font-weight:700;letter-spacing:.08em;color:#9ca3af;text-transform:uppercase;margin-bottom:8px">VS 他社A</div>
      <div style="font-size:22px;font-weight:800;color:#111;margin-bottom:6px">スポーツアプリ専門と何が違うのか？</div>
      <div style="font-size:14px;color:#6b7280;margin-bottom:24px">機能は同じでも、料金は<strong style="color:#111">3分の1</strong>。しかも他社Aにない機能がある。</div>
      <div style="overflow-x:auto;-webkit-overflow-scrolling:touch">
        <table style="width:100%;border-collapse:collapse;min-width:460px">
          <thead>
            <tr>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#9ca3af;font-weight:600;border-bottom:2px solid #f3f4f6;width:44%"></th>
              <th style="padding:12px 16px;text-align:center;font-size:13px;color:#6b7280;font-weight:600;border-bottom:2px solid #f3f4f6">他社A</th>
              <th style="padding:12px 16px;text-align:center;font-size:14px;font-weight:800;color:#fff;background:#d97706;border-bottom:2px solid #d97706;border-radius:10px 10px 0 0">Rak</th>
            </tr>
          </thead>
          <tbody>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">スケジュール共有</td>
              <td style="padding:14px 16px;text-align:center;font-size:13px;color:#6b7280">あり</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:13px;color:#d97706;font-weight:600">専用タブで流れない</td>
            </tr>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">出欠確認</td>
              <td style="padding:14px 16px;text-align:center;font-size:13px;color:#6b7280">あり</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:13px;color:#d97706;font-weight:600">自動リマインド送信</td>
            </tr>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:600">集金・費用管理</td>
              <td style="padding:14px 16px;text-align:center;font-size:20px;color:#e5e7eb">—</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓</td>
            </tr>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:600">ユニフォーム・備品管理</td>
              <td style="padding:14px 16px;text-align:center;font-size:20px;color:#e5e7eb">—</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓</td>
            </tr>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:600">AI文章・スケジュール自動生成</td>
              <td style="padding:14px 16px;text-align:center;font-size:20px;color:#e5e7eb">—</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:18px;color:#d97706">✓ <span style="font-size:10px;font-weight:700">AI</span></td>
            </tr>
            <tr style="border-bottom:1px solid #f3f4f6">
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:500">アプリ内広告</td>
              <td style="padding:14px 16px;text-align:center;font-size:13px;color:#ef4444;font-weight:600">あり</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;font-size:13px;color:#22c55e;font-weight:600">なし</td>
            </tr>
            <tr>
              <td style="padding:14px 16px;font-size:14px;color:#374151;font-weight:600">Proプラン料金</td>
              <td style="padding:14px 16px;text-align:center;font-size:13px;color:#ef4444;font-weight:700">高価格</td>
              <td style="padding:14px 16px;text-align:center;background:#fffbeb;border-radius:0 0 10px 10px;font-size:13px;color:#d97706;font-weight:800">¥980/月〜<br><span style="font-size:10px;font-weight:500;color:#92400e">1チームあたり</span></td>
            </tr>
          </tbody>
        </table>
      </div>
      <div style="margin-top:16px;padding:14px 18px;background:#fffbeb;border-radius:10px;border:1px solid #fde68a;font-size:13px;color:#92400e;line-height:1.7">
        <strong>集金・AI・ユニフォーム管理をまとめて</strong>、料金は他社Aの3分の1。チームの事務作業をひとつのアプリで完結できるのは<span style="white-space:nowrap">Rakだけです。</span>
      </div>
    </div>

    <div style="margin-top:32px;background:#f8f9fb;border-radius:14px;padding:24px 28px;display:flex;gap:20px;align-items:flex-start;flex-wrap:wrap">
      <div style="flex:1;min-width:200px">
        <div style="font-size:13px;font-weight:800;color:#111;margin-bottom:6px">Rakだけができること</div>
        <div style="font-size:13px;color:#6b7280;line-height:1.9">
          AIが予定を自動生成<br>
          AIが連絡文を自動作成<br>
          集金・注文・会計をアプリ内で完結<br>
          カラーコードで活動を視覚管理
        </div>
      </div>
      <div style="flex:1;min-width:200px">
        <div style="font-size:13px;font-weight:800;color:#111;margin-bottom:6px">こんな人に選ばれています</div>
        <div style="font-size:13px;color:#6b7280;line-height:1.9">
          集金のたびに連絡を送るのが辛い<br>
          表計算ソフトの管理が追いつかない<br>
          コーチ業より事務作業が多い<br>
          メンバーへの連絡が面倒
        </div>
      </div>
    </div>
  </div>
</section>

<section class="pricing" id="pricing">
  <div style="max-width:780px;margin:0 auto">
    <span class="sec-label">PRICING</span>
    <div class="sec-title">シンプルな料金。</div>
    <div class="sec-sub" style="margin-bottom:32px">まず無料で始めて、必要になったらアップグレード。</div>
    <div class="plan-grid">
      <div class="plan-card">
        <div class="plan-name">Free</div>
        <div class="plan-price"><span class="num">¥0</span><span class="per">/月</span></div>
        <div class="plan-items">
          <div>✓ スケジュール管理</div>
          <div>✓ お知らせ作成</div>
          <div>✓ メンバー管理（人数制限なし）</div>
          <div>✓ 出欠・予定の共有リンク（登録不要）</div>
        </div>
        <a href="/create?src=plan_free" class="plan-btn-w">無料で始める</a>
      </div>
      <div class="plan-card dark">
        <div class="plan-rec">おすすめ</div>
        <div class="plan-name">Pro</div>
        <div class="plan-price"><span class="num">¥980</span><span class="per">/月</span></div>
        <div style="font-size:11px;color:#aaa;margin-top:-12px;margin-bottom:4px">1チームあたり・税込</div>
        <div style="display:inline-block;font-size:11px;font-weight:700;color:#fff;background:#d97706;border-radius:6px;padding:2px 8px;margin-bottom:6px">リリース記念価格 — ご利用中のチームはずっとこの価格</div>
        <div style="display:inline-block;font-size:11px;font-weight:700;color:#fbbf24;background:rgba(251,191,36,.12);border:1px solid rgba(251,191,36,.3);border-radius:6px;padding:2px 8px;margin-bottom:10px">年額 ¥9,800 で2ヶ月分お得</div>
        <div class="plan-items">
          <div class="acc">＋ 集金・費用管理</div>
          <div class="acc">＋ 注文フォーム</div>
          <div class="acc">＋ ユニフォーム管理</div>
          <div class="acc">＋ 会計・収支記録</div>
          <div class="acc">＋ AI文章生成（活動報告書・お知らせ・保護者連絡）</div>
          <div class="acc">＋ AIスケジュール自動生成</div>
          <div class="acc">＋ Excelエクスポート</div>
        </div>
        <a href="/create?intent=pro&src=plan_pro" class="plan-btn-b">Proを試す（14日無料・クレカ不要）</a>
      </div>
    </div>
  </div>
</section>

<section style="padding:48px 20px;max-width:680px;margin:0 auto">
  <span class="sec-label">FAQ</span>
  <div class="sec-title" style="font-size:22px;margin-bottom:24px">よくある質問</div>
  <div style="display:flex;flex-direction:column;gap:2px">
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">メンバーもアプリを入れる必要がありますか？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">いいえ。アプリを入れて使うのは運営担当のあなただけです。メンバーは、連絡アプリなどで届いた専用リンクを開いて出席/欠席などをタップするだけ。アカウント登録もアプリのインストールも一切不要です。</div>
    </details>
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden;margin-top:8px">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">クレジットカードは必要ですか？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">無料プランもProトライアル（14日間）も、クレジットカードは不要です。Proにアップグレードする際に初めてカード情報が必要になります。</div>
    </details>
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden;margin-top:8px">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">解約はいつでもできますか？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">はい。管理者ページからいつでも解約できます。解約後は次の請求日以降の課金が止まり、期間終了まで引き続きご利用いただけます。</div>
    </details>
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden;margin-top:8px">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">無料プランとProプランの違いは？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">無料プランはスケジュール・連絡・メンバー管理などの基本機能が使えます。Proプランでは集金・費用管理、注文フォーム、AI文章生成・スケジュール自動生成、ユニフォーム管理、Excelエクスポートなどが追加されます。</div>
    </details>
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden;margin-top:8px">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">年度替わりのデータはどうなりますか？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">同じアカウントをそのまま使えます。新年度のメンバー名簿に入れ替えるだけで続けて使え、過去のデータも保持されます。</div>
    </details>
    <details style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:0;overflow:hidden;margin-top:8px">
      <summary style="padding:16px 20px;font-size:14px;font-weight:700;cursor:pointer;list-style:none;display:flex;justify-content:space-between;align-items:center">子どもの個人情報は安全ですか？<span style="color:#9ca3af;font-size:18px">＋</span></summary>
      <div style="padding:0 20px 16px;font-size:13px;color:#6b7280;line-height:1.8">メンバーはアカウント登録不要で、共有された専用リンクからのみアクセスします。リンクを知っている人だけが閲覧・回答でき、第三者への情報提供は行いません。詳細はプライバシーポリシーをご確認ください。</div>
    </details>
  </div>
</section>

<section class="cta-sec">
  <h2>今日から、<br>あなたの事務を<span style="color:var(--rak-amber)">ラク</span>に。</h2>
  <p>メールとパスワードだけ・30秒で開始。無料からはじめられます。メンバーには何もインストールさせません。</p>
  <a href="/create?src=bottom" class="btn-amber-solid">無料ではじめる →</a>
</section>

<footer>
  <div class="footer-logo">{LP_LOGO_W}Rak</div>
  <div class="footer-links">
    <a href="/legal/terms">利用規約</a>
    <a href="/legal/privacy">プライバシー</a>
    <a href="/legal/tokushoho">特定商取引法</a>
    <a href="/feedback">お問い合わせ</a>
    <a href="/create?src=footer">チームを作る</a>
  </div>
  <p>© 2026 Rak</p>
</footer>
{PWA_SW}
</body></html>''')

@app.route('/join', methods=['POST'])
def join():
    code = request.form.get('code', '').strip().upper()
    team = get_team(code)
    if not team:
        return redirect('/?error=notfound&code=' + code)
    return redirect(url_for('team_portal', code=code))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ''
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        if not email or not password:
            error = 'メールアドレスとパスワードを入力してください'
        else:
            conn = get_db()
            team = conn.execute(
                'SELECT team_code FROM teams WHERE LOWER(admin_email)=? AND admin_password=? ORDER BY created_at DESC LIMIT 1',
                (email, password)).fetchone()
            conn.close()
            if team:
                code = team['team_code']
                session.permanent = True
                session[f'admin_{code}'] = True
                return redirect(url_for('admin_dash', code=code))
            error = 'メールアドレスかパスワードが違います'
    body = f'''
<div class="container" style="max-width:420px">
  <div class="card">
    <h1>ログイン</h1>
    <p style="color:#666;font-size:13px;margin-bottom:16px">登録したメールアドレスとパスワードでログインします。</p>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>メールアドレス</label>
      <input type="email" name="email" placeholder="例：admin@example.com" required>
      <label>パスワード</label>
      <div style="position:relative">
        <input type="password" name="password" id="lg-pw" required style="padding-right:44px">
        <button type="button" onclick="var i=document.getElementById('lg-pw');i.type=i.type==='password'?'text':'password';this.textContent=i.type==='password'?'表示':'隠す'" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;color:#888;font-size:12px;cursor:pointer;padding:4px">表示</button>
      </div>
      <button class="btn btn-blue btn-block" type="submit" style="margin-top:14px">ログイン</button>
    </form>
    <div style="text-align:center;margin-top:12px"><a href="/forgot-password" style="font-size:12px;color:#888">パスワードを忘れた方</a></div>
  </div>
  <div style="text-align:center;font-size:13px;color:#888">アカウントがない方は <a href="/create" style="color:#d97706;font-weight:600">新規作成</a></div>
</div>'''
    return page('ログイン', body)


@app.route('/create', methods=['GET', 'POST'])
def create_team():
    error = ''
    intent = request.args.get('intent', '') or request.form.get('intent', '')
    if request.method == 'GET':
        log_lp_event('create_view', src=request.args.get('src', '') or ('pro' if intent == 'pro' else ''))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        password = request.form.get('password', '').strip()
        email = request.form.get('email', '').strip()
        if not password or not email:
            error = 'メールアドレスとパスワードを入力してください'
        elif len(password) < 6:
            error = 'パスワードは6文字以上で設定してください'
        else:
            team_id = new_id()
            code = new_id().upper()[:6]
            trial_end_val = ''
            if intent == 'pro':
                trial_end_val = (datetime.now(JST) + timedelta(days=14)).strftime('%Y-%m-%d')
            import secrets as _secrets
            viewer_token = _secrets.token_urlsafe(16)
            conn = get_db()
            conn.execute(
                'INSERT INTO teams (id,name,sport,team_code,admin_password,created_at,admin_email,trial_end,viewer_token) VALUES (?,?,?,?,?,?,?,?,?)',
                (team_id, name, '', code, password, now_str(), email, trial_end_val, viewer_token)
            )
            conn.commit()
            conn.close()
            log_lp_event('team_created', src=('pro' if intent == 'pro' else ''))
            session.permanent = True
            session[f'admin_{code}'] = True
            if intent == 'pro':
                return redirect(url_for('admin_dash', code=code, created='1'))
            return redirect(url_for('admin_dash', code=code, created='1'))

    pro_badge = '''<div style="background:linear-gradient(135deg,#d97706,#f59e0b);color:#fff;border-radius:12px;padding:14px 16px;margin-bottom:16px">
      <div style="font-size:13px;font-weight:800;margin-bottom:4px">✦ Pro 14日間無料トライアル付きで作成</div>
      <div style="font-size:11px;opacity:.9">クレカ不要・いつでも解約可・14日後に自動課金なし</div>
    </div>''' if intent == 'pro' else ''
    submit_label = 'チームを作成してProトライアルへ →' if intent == 'pro' else 'チームを作成してコードを発行 →'
    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    {pro_badge}
    <h1>はじめる</h1>
    <p style="color:#666;font-size:13px;margin-bottom:16px">メールとパスワードだけ。30秒で完了・クレジットカード不要。</p>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <input type="hidden" name="intent" value="{intent}">
      <label>メールアドレス *</label>
      <input type="email" name="email" placeholder="例：admin@example.com" required>
      <div style="font-size:12px;color:#888;margin-top:4px;margin-bottom:4px">ログインとパスワード再設定に使います。メンバーには不要・公開されません。</div>
      <label>パスワード *</label>
      <div style="position:relative">
        <input type="password" name="password" id="pw-input" placeholder="6文字以上" required style="padding-right:44px">
        <button type="button" onclick="var i=document.getElementById('pw-input');i.type=i.type==='password'?'text':'password';this.textContent=i.type==='password'?'表示':'隠す'" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;color:#888;font-size:12px;cursor:pointer;padding:4px">表示</button>
      </div>
      <div style="font-size:12px;color:#888;margin-top:6px">6文字以上ならOK。チーム名は作成後に設定できます。</div>
      <button class="btn btn-blue btn-block" type="submit">{submit_label}</button>
    </form>
  </div>
  <div style="text-align:center;font-size:13px;color:#888">すでにアカウントをお持ちの方は <a href="/login" style="color:#d97706;font-weight:600">ログイン</a></div>
</div>'''
    return page('はじめる', body)


# ── Password reset ────────────────────────────────────────────────

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    msg = ''
    error = ''
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        conn = get_db()
        team = conn.execute('SELECT * FROM teams WHERE LOWER(admin_email)=?', (email,)).fetchone()
        if team:
            import secrets
            from datetime import datetime as _dt, timedelta as _td
            token = secrets.token_urlsafe(32)
            expires = (_dt.now(JST) + _td(hours=1)).strftime('%Y-%m-%d %H:%M')
            conn.execute('INSERT INTO password_reset_tokens (id,team_id,token,expires_at) VALUES (?,?,?,?)',
                         (new_id(), team['id'], token, expires))
            conn.commit()
            reset_url = f"{base_url()}reset-password?token={token}"
            if RESEND_API_KEY:
                try:
                    import requests as _req
                    _req.post('https://api.resend.com/emails',
                        headers={'Authorization': f'Bearer {RESEND_API_KEY}', 'Content-Type': 'application/json'},
                        json={'from': 'Rak <send@rakapp.jp>', 'to': [email],
                              'subject': '【Rak】パスワードリセットのご案内',
                              'text': f'チーム「{team["name"]}」の管理者パスワードをリセットします。\n\n以下のURLから1時間以内に新しいパスワードを設定してください。\n\n{reset_url}\n\n※このメールに心当たりがない場合は無視してください。'},
                        timeout=10)
                except Exception:
                    pass
        conn.close()
        msg = 'メールアドレスが登録されている場合、リセット用のメールを送信しました'

    body = f'''
<div class="container" style="max-width:440px;padding-top:48px">
  <div class="card">
    <h1 style="margin-bottom:8px">パスワードをお忘れの方</h1>
    <p style="font-size:13px;color:#666;margin-bottom:20px">登録したメールアドレスを入力してください。パスワード再設定のリンクをお送りします。</p>
    {f'<div class="msg-ok">{msg}</div>' if msg else ''}
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>メールアドレス</label>
      <input type="email" name="email" placeholder="登録済みのメールアドレス" required autofocus>
      <button class="btn btn-blue btn-block" type="submit">リセットメールを送る</button>
    </form>
  </div>
  <div style="text-align:center;margin-top:12px"><a href="/" style="font-size:13px;color:#888">← トップに戻る</a></div>
</div>'''
    return page('パスワードリセット', body)


@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    token = request.args.get('token', '') or request.form.get('token', '')
    error = ''
    conn = get_db()
    now_s = datetime.now(JST).strftime('%Y-%m-%d %H:%M')
    rec = conn.execute(
        'SELECT * FROM password_reset_tokens WHERE token=? AND used=0 AND expires_at>=?',
        (token, now_s)
    ).fetchone()
    conn.close()

    if not rec:
        body = '''<div class="container" style="max-width:440px;padding-top:48px">
  <div class="card" style="text-align:center;padding:40px">
    <h1 style="font-size:20px;margin-bottom:12px">リンクが無効です</h1>
    <p style="color:#666;font-size:13px;margin-bottom:24px">リンクの有効期限が切れているか、すでに使用済みです。</p>
    <a href="/forgot-password" class="btn btn-blue">再度リセット申請する</a>
  </div></div>'''
        return page('リンク無効', body)

    if request.method == 'POST':
        new_pw = request.form.get('password', '').strip()
        if len(new_pw) < 6:
            error = 'パスワードは6文字以上にしてください'
        elif not any(c.isalpha() for c in new_pw):
            error = '英字を1文字以上含めてください'
        elif not any(c.isdigit() for c in new_pw):
            error = '数字を1文字以上含めてください'
        else:
            conn = get_db()
            conn.execute('UPDATE teams SET admin_password=? WHERE id=?', (new_pw, rec['team_id']))
            conn.execute('UPDATE password_reset_tokens SET used=1 WHERE id=?', (rec['id'],))
            conn.commit()
            conn.close()
            body = '''<div class="container" style="max-width:440px;padding-top:48px">
  <div class="card" style="text-align:center;padding:40px">
    <h1 style="font-size:20px;margin-bottom:12px">パスワードを変更しました</h1>
    <p style="color:#666;font-size:13px;margin-bottom:24px">新しいパスワードでログインしてください。</p>
    <a href="/" class="btn btn-blue">トップへ戻る</a>
  </div></div>'''
            return page('変更完了', body)

    body = f'''
<div class="container" style="max-width:440px;padding-top:48px">
  <div class="card">
    <h1 style="margin-bottom:8px">新しいパスワードを設定</h1>
    <p style="font-size:13px;color:#666;margin-bottom:20px">英字・数字を含む6文字以上で設定してください。</p>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <input type="hidden" name="token" value="{token}">
      <label>新しいパスワード</label>
      <div style="position:relative">
        <input type="password" name="password" id="reset-pw" placeholder="例：soccer2026" required style="padding-right:44px">
        <button type="button" onclick="var i=document.getElementById('reset-pw');i.type=i.type==='password'?'text':'password';this.textContent=i.type==='password'?'表示':'隠す'" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;color:#888;font-size:12px;cursor:pointer;padding:4px">表示</button>
      </div>
      <button class="btn btn-blue btn-block" type="submit">パスワードを変更する</button>
    </form>
  </div>
</div>'''
    return page('パスワード再設定', body)


# ── Member portal ─────────────────────────────────────────────────

@app.route('/t/<code>')
def team_portal(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    if not member and not is_admin(code):
        # 既存メンバー名簿を取得
        _conn = get_db()
        roster = [r['name'] for r in _conn.execute('SELECT name FROM members WHERE team_id=? ORDER BY name', (team['id'],)).fetchall()]
        _conn.close()

        if roster:
            # 名簿あり：一覧から選ぶUI（二重登録防止）
            name_buttons = ''.join(
                f'<form method="POST" action="/t/{code}/join" style="margin:0"><input type="hidden" name="name" value="{n}"><button type="submit" class="btn btn-outline" style="width:100%;text-align:left;padding:11px 14px;font-size:14px">{n}</button></form>'
                for n in roster
            )
            body = f'''
<div class="container" style="max-width:480px;padding-top:60px">
  <div class="card">
    <div style="text-align:center;margin-bottom:20px">
      <div style="margin-bottom:8px">{_ICO_WELCOME}</div>
      <h1 style="margin-bottom:4px">{team["name"]}</h1>
      <p style="color:#666;font-size:13px">あなたの名前を選んでください</p>
    </div>
    <div id="roster-list" style="display:flex;flex-direction:column;gap:6px;margin-bottom:16px">
      {name_buttons}
    </div>
    <div style="text-align:center">
      <button onclick="document.getElementById('roster-list').style.display='none';document.getElementById('name-form').style.display='block';this.style.display='none'" style="background:none;border:none;color:#aaa;font-size:12px;cursor:pointer;padding:6px">自分の名前がない →</button>
    </div>
    <!-- 名前がない場合の手入力フォーム -->
    <div id="name-form" style="display:none;margin-top:12px">
      <div style="font-size:12px;color:#d97706;background:#fffbeb;border-radius:6px;padding:8px 12px;margin-bottom:12px">管理者に名前の登録を依頼するか、以下から入力してください</div>
      <form method="POST" action="/t/{code}/join" onsubmit="rakSave(this)">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
          <div>
            <label style="font-size:11px;color:#888;margin-bottom:4px;display:block">苗字</label>
            <input type="text" name="last_name" placeholder="田中" required style="text-align:center">
          </div>
          <div>
            <label style="font-size:11px;color:#888;margin-bottom:4px;display:block">名前</label>
            <input type="text" name="first_name" placeholder="花子" required style="text-align:center">
          </div>
        </div>
        <button class="btn btn-blue btn-block" type="submit">入る →</button>
      </form>
    </div>
  </div>
</div>'''
        else:
            # 名簿なし：従来の手入力フォーム
            body = f'''
<div class="container" style="max-width:480px;padding-top:60px">
  <div class="card" style="text-align:center">
    <div style="margin-bottom:12px">{_ICO_WELCOME}</div>
    <h1 style="margin-bottom:6px">{team["name"]}</h1>
    <p style="color:#666;font-size:13px;margin-bottom:20px">氏名を入力してください</p>
    <!-- localStorage に記憶がある場合のワンタップ入室 -->
    <div id="quick-join" style="display:none;margin-bottom:12px">
      <form method="POST" action="/t/{code}/join" id="quick-form">
        <input type="hidden" name="last_name" id="q-last">
        <input type="hidden" name="first_name" id="q-first">
        <button type="submit" class="btn btn-blue btn-block" style="font-size:16px;padding:14px">
          <span id="q-name">入る</span>さんとして入る →
        </button>
      </form>
      <button onclick="document.getElementById('quick-join').style.display='none';document.getElementById('name-form').style.display='block'" style="background:none;border:none;color:#aaa;font-size:12px;margin-top:10px;cursor:pointer;width:100%">別の名前で入る</button>
    </div>
    <div id="name-form">
      <form method="POST" action="/t/{code}/join" onsubmit="rakSave(this)">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:0">
          <div>
            <label style="font-size:11px;color:#888;text-align:left;margin-bottom:4px;display:block">苗字</label>
            <input type="text" name="last_name" placeholder="田中" required style="text-align:center">
          </div>
          <div>
            <label style="font-size:11px;color:#888;text-align:left;margin-bottom:4px;display:block">名前</label>
            <input type="text" name="first_name" placeholder="花子" required style="text-align:center">
          </div>
        </div>
        <button class="btn btn-blue btn-block" type="submit">入る →</button>
      </form>
    </div>
    <div style="margin-top:16px"><a href="/t/{code}/help" style="font-size:12px;color:#aaa">使い方を見る →</a></div>
    <script>
    (function(){{
      var k='rak_m_{code}';
      try{{
        var d=JSON.parse(localStorage.getItem(k)||'null');
        if(d&&d.name){{
          document.getElementById('q-name').textContent=d.name;
          document.getElementById('q-last').value=d.last||'';
          document.getElementById('q-first').value=d.first||'';
          document.getElementById('quick-join').style.display='block';
          document.getElementById('name-form').style.display='none';
        }}
      }}catch(e){{}}
    }})();
    function rakSave(f){{
      try{{
        var l=f.last_name.value.trim(),fn=f.first_name.value.trim();
        if(l||fn)localStorage.setItem('rak_m_{code}',JSON.stringify({{last:l,first:fn,name:(l+' '+fn).trim()}}));
      }}catch(e){{}}
    }}
    </script>
  </div>
</div>'''
        return page(team['name'], body, code)

    # logged-in member or admin → member home
    if is_admin(code):
        return redirect(url_for('admin_dash', code=code))
    return redirect(url_for('member_home', code=code))

@app.route('/t/<code>/join', methods=['POST'])
def member_join(code):
    last_name = request.form.get('last_name', '').strip()
    first_name = request.form.get('first_name', '').strip()
    name = f'{last_name} {first_name}'.strip() if (last_name or first_name) else request.form.get('name', '').strip()
    if name:
        session.permanent = True
        session[f'member_{code}'] = name
    return redirect(url_for('member_home', code=code))

@app.route('/t/<code>/home')
def member_home(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))
    if admin and not member:
        return redirect(url_for('admin_dash', code=code))

    now = datetime.now(JST)
    today_s = now.strftime('%Y-%m-%d')
    try:
        vy = int(request.args.get('y', now.year))
        vm = int(request.args.get('m', now.month))
        vm = max(1, min(12, vm))
    except Exception:
        vy, vm = now.year, now.month
    py, pm = (vy - 1, 12) if vm == 1 else (vy, vm - 1)
    ny, nm = (vy + 1, 1) if vm == 12 else (vy, vm + 1)
    month_start = f'{vy}-{vm:02d}-01'
    month_end = f'{ny}-{nm:02d}-01'

    conn = get_db()

    # カレンダー用データ
    cal_events = conn.execute(
        'SELECT event_date, end_date FROM events WHERE team_id=? AND event_date>=? AND event_date<? ORDER BY event_date',
        (team['id'], month_start, month_end)
    ).fetchall()
    cal_fees = conn.execute(
        "SELECT due_date FROM fees WHERE team_id=? AND due_date>=? AND due_date<? AND due_date!=''",
        (team['id'], month_start, month_end)
    ).fetchall()
    cal_orders = conn.execute(
        "SELECT deadline FROM order_forms WHERE team_id=? AND deadline>=? AND deadline<? AND deadline!=''",
        (team['id'], month_start, month_end)
    ).fetchall()

    ev_dates = set()
    for ev in cal_events:
        cur = datetime.strptime(ev['event_date'], '%Y-%m-%d')
        end_d = datetime.strptime(ev['end_date'], '%Y-%m-%d') if ev['end_date'] else cur
        while cur <= end_d:
            ev_dates.add(cur.strftime('%Y-%m-%d'))
            cur += timedelta(days=1)
    fee_dates_cal = set(f['due_date'] for f in cal_fees)
    order_dates_cal = set(o['deadline'] for o in cal_orders)
    calendar_html = build_calendar(vy, vm, ev_dates, fee_dates_cal, order_dates_cal)

    # バッジカウント
    unread = conn.execute(
        'SELECT COUNT(*) FROM notices WHERE team_id=? AND id NOT IN (SELECT notice_id FROM reads WHERE member_name=?)',
        (team['id'], member)
    ).fetchone()[0]
    unpaid = conn.execute(
        '''SELECT COUNT(*) FROM fees f WHERE f.team_id=?
           AND NOT EXISTS (SELECT 1 FROM fee_payments WHERE fee_id=f.id AND member_name=? AND paid=1)''',
        (team['id'], member)
    ).fetchone()[0]

    # 直近の予定リスト（今日以降4件）
    upcoming = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? ORDER BY event_date LIMIT 4',
        (team['id'], today_s)
    ).fetchall()

    # 自分のRSVP状態を一括取得
    upcoming_ids = [ev['id'] for ev in upcoming]
    rsvp_map = {}
    if upcoming_ids:
        placeholders = ','.join('?' * len(upcoming_ids))
        rows = conn.execute(
            f'SELECT event_id, status FROM rsvps WHERE member_name=? AND event_id IN ({placeholders})',
            [member] + upcoming_ids
        ).fetchall()
        rsvp_map = {r['event_id']: r['status'] for r in rows}

    unanswered_rsvp = sum(1 for ev in upcoming if ev['id'] not in rsvp_map)
    conn.close()

    event_items = ''
    for ev in upcoming:
        my_rsvp = rsvp_map.get(ev['id'], '')

        rsvp_area = f'''
      <form method="POST" action="/t/{code}/rsvp/{ev['id']}" style="display:flex;gap:6px;margin-top:8px">
        <button name="status" value="attending" type="submit"
          style="flex:1;padding:6px;border-radius:7px;font-size:12px;font-weight:600;border:1.5px solid {'#16a34a' if my_rsvp=='attending' else 'var(--rak-line)'};background:{'#f0fdf4' if my_rsvp=='attending' else '#fff'};color:{'#16a34a' if my_rsvp=='attending' else 'var(--rak-mute)'};cursor:pointer">
          {'✓ ' if my_rsvp=='attending' else ''}出席
        </button>
        <button name="status" value="absent" type="submit"
          style="flex:1;padding:6px;border-radius:7px;font-size:12px;font-weight:600;border:1.5px solid {'#dc2626' if my_rsvp=='absent' else 'var(--rak-line)'};background:{'#fef2f2' if my_rsvp=='absent' else '#fff'};color:{'#dc2626' if my_rsvp=='absent' else 'var(--rak-mute)'};cursor:pointer">
          {'✓ ' if my_rsvp=='absent' else ''}欠席
        </button>
      </form>'''

        unanswered_badge = '' if my_rsvp else '<span style="font-size:10px;background:#fef3c7;color:#d97706;font-weight:700;padding:1px 7px;border-radius:999px;margin-left:6px">未回答</span>'
        event_items += f'''
    <div style="padding:10px 0;border-bottom:1px solid var(--rak-line-soft)">
      <div style="display:flex;gap:12px;align-items:flex-start">
        <div style="min-width:40px;text-align:center;background:var(--rak-bg-soft);border-radius:8px;padding:5px 4px;flex-shrink:0">
          <div style="font-size:9px;color:var(--rak-mute)">{fmt_date(ev["event_date"])[:5]}</div>
          <div style="font-size:17px;font-weight:600;font-family:var(--font-num);line-height:1.2">{int(ev["event_date"].split("-")[2])}</div>
        </div>
        <div style="flex:1;min-width:0">
          <div style="font-weight:500;font-size:14px">{ev["title"]}{unanswered_badge}</div>
          {f'<div style="font-size:12px;color:var(--rak-mute)">{ev["event_time"]}{("　" + ev["location"]) if ev["location"] else ""}</div>' if ev["event_time"] or ev["location"] else ''}
          {rsvp_area}
        </div>
      </div>
    </div>'''
    if not event_items:
        event_items = '<div style="padding:14px 0;text-align:center;color:var(--rak-mute);font-size:13px">予定はありません</div>'

    first_name = member.split()[-1] if ' ' in member else member

    body = f'''
<div class="container" style="max-width:480px">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px">
    <div>
      <div style="font-size:12px;color:var(--rak-mute)">{team["name"]}</div>
      <h1 style="margin-top:1px">{first_name}さん</h1>
    </div>
    <div style="display:flex;gap:8px;margin-left:auto">
      <a href="/t/{code}/notices" style="position:relative;display:flex;flex-direction:column;align-items:center;background:#fff;border:1px solid var(--rak-line);border-radius:10px;padding:10px 14px;text-decoration:none;min-width:60px">
        {"" if not unread else f'<span style="position:absolute;top:-6px;right:-6px;background:#dc2626;color:#fff;border-radius:10px;font-size:10px;font-weight:700;padding:1px 5px">{unread}</span>'}
        <div style="font-size:18px;font-weight:600;color:{"var(--rak-amber)" if unread else "var(--rak-black)"}">{unread}</div>
        <div style="font-size:10px;color:var(--rak-mute);margin-top:1px">未読</div>
      </a>
      <a href="/t/{code}/schedule" style="display:flex;flex-direction:column;align-items:center;background:#fff;border:1px solid var(--rak-line);border-radius:10px;padding:10px 14px;text-decoration:none;min-width:60px">
        <div style="font-size:18px;font-weight:600;color:{"var(--rak-amber)" if unanswered_rsvp else "var(--rak-black)"}">{unanswered_rsvp}</div>
        <div style="font-size:10px;color:var(--rak-mute);margin-top:1px">未回答</div>
      </a>
      <a href="/t/{code}/fees" style="display:flex;flex-direction:column;align-items:center;background:#fff;border:1px solid var(--rak-line);border-radius:10px;padding:10px 14px;text-decoration:none;min-width:60px">
        <div style="font-size:18px;font-weight:600;color:{"var(--rak-danger)" if unpaid else "var(--rak-black)"}">{unpaid}</div>
        <div style="font-size:10px;color:var(--rak-mute);margin-top:1px">未払い</div>
      </a>
    </div>
  </div>

  <div class="card" style="margin-bottom:12px">
    <div style="display:flex;align-items:center;margin-bottom:12px">
      <a href="/t/{code}/home?y={py}&m={pm}" style="width:32px;height:32px;display:flex;align-items:center;justify-content:center;border-radius:50%;background:var(--rak-bg-soft);color:var(--rak-mute);font-size:18px;text-decoration:none">‹</a>
      <div style="flex:1;text-align:center;font-weight:600;font-size:15px">{vy}年{vm}月</div>
      <a href="/t/{code}/home?y={ny}&m={nm}" style="width:32px;height:32px;display:flex;align-items:center;justify-content:center;border-radius:50%;background:var(--rak-bg-soft);color:var(--rak-mute);font-size:18px;text-decoration:none">›</a>
    </div>
    {calendar_html}
    <div style="display:flex;gap:14px;margin-top:10px;font-size:11px;color:var(--rak-mute);justify-content:center">
      <span><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:#111;margin-right:3px;vertical-align:middle"></span>予定</span>
      <span><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:#dc2626;margin-right:3px;vertical-align:middle"></span>集金</span>
      <span><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:#16a34a;margin-right:3px;vertical-align:middle"></span>注文</span>
    </div>
  </div>

  <div class="card">
    <div style="font-size:11px;font-weight:600;color:var(--rak-mute);letter-spacing:.06em;text-transform:uppercase;margin-bottom:4px">直近の予定</div>
    {event_items}
    <a href="/t/{code}/schedule" style="display:block;text-align:center;font-size:13px;color:var(--rak-amber);margin-top:10px">すべて見る →</a>
  </div>

  <div id="rak-push-banner" style="display:none;align-items:center;gap:12px;background:#fff;border:1px solid var(--rak-line);border-radius:10px;padding:14px 16px;margin-top:12px">
    <div style="flex:1;font-size:13px;color:var(--rak-ink)">
      <div style="font-weight:600;margin-bottom:2px">お知らせ通知をオンにする</div>
      <div style="font-size:12px;color:var(--rak-mute)">管理者が投稿したときにすぐ届きます</div>
    </div>
    <button onclick="rakPushBannerOn('{code}')"
      class="btn btn-amber btn-sm">通知をオン</button>
    <button onclick="localStorage.setItem('rak_push_{code}','skip');this.closest('#rak-push-banner').style.display='none';"
      style="background:none;border:none;color:var(--rak-mute);font-size:18px;cursor:pointer;padding:0 4px">×</button>
  </div>
</div>
<script>
(function(){{
  try{{
    var parts='{member}'.split(' ');
    localStorage.setItem('rak_m_{code}',JSON.stringify({{last:parts[0]||'',first:parts[1]||'',name:'{member}'}}));
  }}catch(e){{}}
  rakRequestPush('{code}');
}})();
</script>'''
    return page('ホーム', body, code, active='home')


# ── Schedule ──────────────────────────────────────────────────────

EVENT_COLORS = [
    ('#3b82f6', '青'),
    ('#ef4444', '赤'),
    ('#f97316', '橙'),
    ('#22c55e', '緑'),
    ('#a855f7', '紫'),
    ('#6b7280', 'グレー'),
]

def color_picker_html(selected=''):
    items = ''
    for hex_color, label in EVENT_COLORS:
        checked = 'checked' if selected == hex_color else ''
        ring = f'box-shadow:0 0 0 3px #fff,0 0 0 5px {hex_color}' if selected == hex_color else ''
        items += f'''<label style="display:flex;flex-direction:column;align-items:center;gap:4px;cursor:pointer;margin:0">
          <input type="radio" name="event_color" value="{hex_color}" {checked} style="display:none" onchange="document.querySelectorAll('.cpick-dot').forEach(d=>d.style.boxShadow='');this.nextElementSibling.style.boxShadow='0 0 0 3px #fff,0 0 0 5px '+this.value">
          <div class="cpick-dot" style="width:28px;height:28px;border-radius:50%;background:{hex_color};cursor:pointer;{ring};transition:box-shadow .15s;flex-shrink:0"></div>
          <span style="font-size:10px;color:#666;white-space:nowrap;line-height:1">{label}</span>
        </label>'''
    return f'<div style="display:flex;gap:14px;flex-wrap:wrap;align-items:flex-start;margin-bottom:14px">{items}</div>'

def build_calendar(year, month, event_dates, fee_dates=None, order_dates=None, admin_code=None):
    import calendar as _cal
    fee_dates = fee_dates or set()
    order_dates = order_dates or set()
    if isinstance(event_dates, dict):
        ev_color_map = event_dates
    else:
        ev_color_map = {d: ['#6b7280'] for d in event_dates}
    cal = _cal.monthcalendar(year, month)
    wd_labels = ['月','火','水','木','金','土','日']
    wd_colors = ['#374151','#374151','#374151','#374151','#374151','#2563eb','#dc2626']
    header = ''.join(
        f'<div style="text-align:center;font-size:11px;font-weight:500;color:{wd_colors[i]};padding:4px 0">{d}</div>'
        for i, d in enumerate(wd_labels)
    )
    today_str = datetime.now(JST).strftime('%Y-%m-%d')
    rows = ''
    for week in cal:
        for wd_idx, day in enumerate(week):
            if day == 0:
                rows += '<div></div>'
            else:
                date_str = f'{year}-{month:02d}-{day:02d}'
                has_ev  = date_str in ev_color_map
                has_fee = date_str in fee_dates
                has_ord = date_str in order_dates
                has_any = has_ev or has_fee or has_ord
                is_today = date_str == today_str
                bg = 'background:#0a0a0a;color:#fff;' if is_today else ''
                fw = '700' if (is_today or has_any) else '400'
                # 土日の文字色（今日は白で上書き）
                if not is_today:
                    day_color = wd_colors[wd_idx]
                else:
                    day_color = '#fff'
                # クリック動作: 予定があればスクロール、管理者かつ予定なしなら追加ページへ
                if has_ev:
                    cursor = 'pointer'
                    onclick = f'onclick="scrollToDate(\'{date_str}\')"'
                elif admin_code:
                    cursor = 'pointer'
                    onclick = f'onclick="location.href=\'/t/{admin_code}/admin/events/new?date={date_str}\'"'
                else:
                    cursor = 'default'
                    onclick = ''
                dots = ''
                if has_ev:
                    for c in ev_color_map[date_str][:3]:
                        dots += f'<div style="width:5px;height:5px;border-radius:50%;background:{c};display:inline-block;margin:0 1px"></div>'
                if has_fee: dots += '<div style="width:5px;height:5px;border-radius:50%;background:#dc2626;display:inline-block;margin:0 1px"></div>'
                if has_ord: dots += '<div style="width:5px;height:5px;border-radius:50%;background:#16a34a;display:inline-block;margin:0 1px"></div>'
                dot_row = f'<div style="display:flex;justify-content:center;min-height:7px;margin-top:2px">{dots}</div>'
                rows += f'<div style="text-align:center;padding:5px 2px;border-radius:8px;cursor:{cursor};{bg}" {onclick}><div style="font-size:13px;font-weight:{fw};color:{day_color}">{day}</div>{dot_row}</div>'
    return f'<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:2px">{header}{rows}</div>'

# ── 保護者用 閲覧専用ページ ───────────────────────────────────────────────
@app.route('/t/<code>/view/<token>')
def viewer_page(code, token):
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return render_template_string('<!DOCTYPE html><html lang="ja"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Rak</title></head><body style="font-family:sans-serif;text-align:center;padding:60px 20px;color:#888"><p>リンクが無効または期限切れです。</p><a href="/" style="color:#d97706">トップへ戻る</a></body></html>'), 404

    now = datetime.now(JST)
    today = now.strftime('%Y-%m-%d')

    try:
        vy = int(request.args.get('y', now.year))
        vm = int(request.args.get('m', now.month))
        vm = max(1, min(12, vm))
    except Exception:
        vy, vm = now.year, now.month

    month_start = f'{vy}-{vm:02d}-01'
    ny, nm = (vy + 1, 1) if vm == 12 else (vy, vm + 1)
    month_end = f'{ny}-{nm:02d}-01'
    py, pm = (vy - 1, 12) if vm == 1 else (vy, vm - 1)

    conn = get_db()
    all_events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? AND event_date<? ORDER BY event_date,event_time',
        (team['id'], month_start, month_end)
    ).fetchall()
    upcoming = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? ORDER BY event_date,event_time LIMIT 10',
        (team['id'], today)
    ).fetchall()
    notices = conn.execute(
        'SELECT * FROM notices WHERE team_id=? ORDER BY created_at DESC LIMIT 5',
        (team['id'],)
    ).fetchall()
    conn.close()

    ev_color_map = {}
    for ev in all_events:
        c = ev['event_color'] or '#6b7280'
        cur = datetime.strptime(ev['event_date'], '%Y-%m-%d')
        end_d = datetime.strptime(ev['end_date'], '%Y-%m-%d') if ev['end_date'] else cur
        while cur <= end_d:
            ds = cur.strftime('%Y-%m-%d')
            ev_color_map.setdefault(ds, [])
            if c not in ev_color_map[ds]:
                ev_color_map[ds].append(c)
            cur += timedelta(days=1)
    calendar_html = build_calendar(vy, vm, ev_color_map)

    wd_jp = ['月','火','水','木','金','土','日']
    def dl(d):
        try:
            from datetime import date as _d
            dt = _d.fromisoformat(d)
            return f'{dt.month}/{dt.day}({wd_jp[dt.weekday()]})'
        except Exception:
            return d

    event_rows = ''
    for ev in upcoming:
        border = ev['event_color'] or '#e5e7eb'
        loc = f'<span style="font-size:11px;color:#9ca3af;margin-left:6px">📍{ev["location"]}</span>' if ev['location'] else ''
        time_s = f'<span style="font-size:11px;color:#9ca3af;margin-left:4px">{ev["event_time"]}</span>' if ev['event_time'] else ''
        event_rows += f'''
        <div style="border-left:4px solid {border};padding:10px 12px;background:#fff;border-radius:0 8px 8px 0;margin-bottom:8px;box-shadow:0 1px 3px rgba(0,0,0,.06)">
          <div style="font-size:11px;color:#6b7280;margin-bottom:2px">{dl(ev["event_date"])}{time_s}</div>
          <div style="font-size:14px;font-weight:600;color:#111">{ev["title"]}{loc}</div>
          {f'<div style="font-size:12px;color:#6b7280;margin-top:4px">{ev["note"]}</div>' if ev["note"] else ''}
        </div>'''

    notice_rows = ''
    for n in notices:
        notice_rows += f'''
        <div style="padding:12px 14px;background:#fff;border-radius:8px;margin-bottom:8px;box-shadow:0 1px 3px rgba(0,0,0,.06)">
          <div style="font-weight:600;color:#111;font-size:14px">{n["title"]}</div>
          <div style="font-size:11px;color:#9ca3af;margin-top:2px">{fmt_datetime(n["created_at"])}</div>
          <div style="font-size:13px;color:#374151;margin-top:6px;white-space:pre-wrap;line-height:1.6">{n["body"][:200]}{"..." if len(n["body"])>200 else ""}</div>
        </div>'''

    prev_link = f'/t/{code}/view/{token}?y={py}&m={pm:02d}'
    next_link = f'/t/{code}/view/{token}?y={ny}&m={nm:02d}'

    html = f'''<!DOCTYPE html>
<html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}
{FONT}
<title>{team["name"] or "チーム"} | Rak 閲覧</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Hiragino Sans","Yu Gothic",sans-serif;background:#f5f5f7;color:#1a1a1a;min-height:100vh}}
.hdr{{background:#0a0a0a;color:#fff;padding:14px 20px;display:flex;align-items:center;gap:10px}}
.hdr-logo{{font-weight:900;font-size:18px;letter-spacing:-.02em}}
.hdr-name{{font-size:14px;opacity:.7}}
.badge{{background:#d97706;color:#fff;font-size:10px;font-weight:700;padding:2px 7px;border-radius:999px;margin-left:8px}}
.wrap{{max-width:520px;margin:0 auto;padding:16px 14px 40px}}
.sec-hd{{font-size:12px;font-weight:700;color:#9ca3af;letter-spacing:.06em;text-transform:uppercase;margin:20px 0 10px}}
.cal-nav{{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px}}
.cal-nav a{{color:#d97706;font-size:22px;text-decoration:none;padding:4px 10px}}
.cal-nav span{{font-weight:700;font-size:15px}}
.footer{{text-align:center;padding:20px;font-size:11px;color:#bbb}}
</style>
</head><body>
<div class="hdr">
  <div class="hdr-logo">Rak</div>
  <div style="flex:1">
    <div class="hdr-name">{team["name"] or "チーム"}</div>
  </div>
  <span class="badge">閲覧専用</span>
</div>
<div class="wrap">
  <div class="sec-hd">カレンダー</div>
  <div class="cal-nav">
    <a href="{prev_link}">‹</a>
    <span>{vy}年{vm}月</span>
    <a href="{next_link}">›</a>
  </div>
  {calendar_html}

  <div class="sec-hd" style="margin-top:24px">直近の予定</div>
  {event_rows if upcoming else '<div style="color:#9ca3af;font-size:13px;padding:12px 0">直近の予定はありません</div>'}

  <div class="sec-hd" style="margin-top:24px">お知らせ</div>
  {notice_rows if notices else '<div style="color:#9ca3af;font-size:13px;padding:12px 0">お知らせはありません</div>'}
</div>
<div class="footer">Powered by Rak — 閲覧専用ページ</div>
</body></html>'''
    return html

@app.route('/t/<code>/schedule')
def schedule(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    now = datetime.now(JST)
    today = now.strftime('%Y-%m-%d')

    try:
        vy = int(request.args.get('y', now.year))
        vm = int(request.args.get('m', now.month))
        vm = max(1, min(12, vm))
    except Exception:
        vy, vm = now.year, now.month

    month_start = f'{vy}-{vm:02d}-01'
    ny, nm = (vy + 1, 1) if vm == 12 else (vy, vm + 1)
    month_end = f'{ny}-{nm:02d}-01'
    py, pm = (vy - 1, 12) if vm == 1 else (vy, vm - 1)

    conn = get_db()
    all_events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? AND event_date<? ORDER BY event_date,event_time',
        (team['id'], month_start, month_end)
    ).fetchall()

    fees_in_month = conn.execute(
        "SELECT * FROM fees WHERE team_id=? AND due_date>=? AND due_date<? AND due_date!='' ORDER BY due_date",
        (team['id'], month_start, month_end)
    ).fetchall()
    orders_in_month = conn.execute(
        "SELECT deadline FROM order_forms WHERE team_id=? AND deadline>=? AND deadline<? AND deadline!=''",
        (team['id'], month_start, month_end)
    ).fetchall()

    ev_color_map = {}
    for ev in all_events:
        c = ev['event_color'] or '#6b7280'
        cur = datetime.strptime(ev['event_date'], '%Y-%m-%d')
        end_d = datetime.strptime(ev['end_date'], '%Y-%m-%d') if ev['end_date'] else cur
        while cur <= end_d:
            ds = cur.strftime('%Y-%m-%d')
            ev_color_map.setdefault(ds, [])
            if c not in ev_color_map[ds]:
                ev_color_map[ds].append(c)
            cur += timedelta(days=1)
    fee_dates = set(f['due_date'] for f in fees_in_month)
    order_dates = set(o['deadline'] for o in orders_in_month)
    calendar_html = build_calendar(vy, vm, ev_color_map, fee_dates, order_dates, admin_code=code if admin else None)

    event_cards = ''
    for ev in all_events:
        rsvps = conn.execute('SELECT * FROM rsvps WHERE event_id=?', (ev['id'],)).fetchall()
        attending = sum(1 for r in rsvps if r['status'] == 'attending')
        absent = sum(1 for r in rsvps if r['status'] == 'absent')
        _mode = ev['rsvp_mode'] or 'both'
        if _mode == 'none':
            rsvp_badges = ''
        elif _mode == 'absent_only':
            rsvp_badges = f'<span class="badge badge-red">欠席 {absent}</span>'
        else:
            rsvp_badges = f'<span class="badge badge-green">出席 {attending}</span><span class="badge badge-red">欠席 {absent}</span>'
        my_rsvp = ''
        if member:
            r = conn.execute('SELECT status FROM rsvps WHERE event_id=? AND member_name=?',
                             (ev['id'], member)).fetchone()
            my_rsvp = r['status'] if r else ''
        rsvp_btns = ''
        if member:
            rsvp_btns = f'''
            <form method="POST" action="/t/{code}/rsvp/{ev['id']}" style="display:flex;gap:8px;margin-top:12px">
              <button name="status" value="attending" class="btn btn-sm {'btn-blue' if my_rsvp=='attending' else 'btn-outline'}" type="submit">出席</button>
              <button name="status" value="absent" class="btn btn-sm btn-gray" type="submit" style="{'background:#fee2e2;color:#dc2626' if my_rsvp=='absent' else ''}">欠席</button>
            </form>'''
        admin_btns = ''
        if admin:
            admin_btns = f'''
            <div style="display:flex;gap:6px;margin-top:10px">
              <a href="/t/{code}/admin/events/{ev['id']}/edit" class="btn btn-sm btn-outline">編集</a>
              <form method="POST" action="/t/{code}/admin/events/{ev['id']}/delete" onsubmit="return confirm('削除しますか？')" style="margin:0">
                <button class="btn btn-sm btn-gray" type="submit" style="color:#dc2626">削除</button>
              </form>
            </div>'''
        ev_color_border = ev['event_color'] or '#e5e7eb'
        event_cards += f'''
        <div class="card-sm" id="ev-{ev['event_date']}" style="border-left:4px solid {ev_color_border}">
          <div class="row" style="flex-wrap:wrap;gap:6px">
            <div style="flex:1;min-width:0">
              <div style="font-weight:700;font-size:16px">{ev['title']}</div>
              <div style="font-size:13px;color:#666;margin-top:2px">{fmt_date_range(ev['event_date'], ev['end_date'])}{'　' + ev['event_time'] + ('〜' + (ev['end_time'] if ev['end_time'] else '') if ev['event_time'] else '') if ev['event_time'] else ''}{('　' + ev['location']) if ev['location'] else ''}</div>
            </div>
            <div style="display:flex;gap:6px;align-items:center">{rsvp_badges}</div>
          </div>
          {f'<div style="font-size:13px;color:#666;margin-top:8px;background:#f8faff;padding:8px 12px;border-radius:8px">{ev["note"]}</div>' if ev['note'] else ''}
          {rsvp_btns}
          {admin_btns}
        </div>'''

    fee_cards = ''
    for f in fees_in_month:
        paid_row = conn.execute('SELECT paid FROM fee_payments WHERE fee_id=? AND member_name=?',
                                (f['id'], member)).fetchone() if member else None
        paid = paid_row['paid'] if paid_row else None
        status_badge = ''
        if member:
            status_badge = '<span class="badge badge-green">支払済</span>' if paid else '<span class="badge badge-red">未払い</span>'
        fee_cards += f'''
        <div class="card-sm" id="ev-{f['due_date']}" style="border-left:3px solid #f59e0b;background:#fffbeb">
          <div class="row" style="justify-content:space-between;align-items:center">
            <div>
              <div style="font-weight:700">集金期限：{f['title']}</div>
              <div style="font-size:13px;color:#666;margin-top:2px">{fmt_date(f['due_date'])}　¥{f['amount']:,}</div>
            </div>
            {status_badge}
          </div>
        </div>'''

    conn.close()

    is_this_month = (vy == now.year and vm == now.month)
    today_btn = '' if is_this_month else f'<a href="/t/{code}/schedule" style="font-size:12px;color:#d97706;padding:3px 10px;border:1.5px solid #d97706;border-radius:8px;text-decoration:none">今月</a>'
    new_btn = f'<a href="/t/{code}/admin/events/new" class="btn btn-blue btn-sm">＋ 追加</a>' if admin else ''
    ai_btn = ''
    excel_btn = ''
    if admin:
        if is_pro(team):
            ai_btn = f'<a href="/t/{code}/admin/ai-schedule" class="btn btn-sm" style="background:#d97706;color:#fff">✦ AI予定作成</a>'
            excel_btn = f'<a href="/t/{code}/admin/schedule/excel?y={vy}&m={vm}" class="btn btn-sm btn-gray">📥 Excel</a>'
        else:
            ai_btn = f'<a href="/t/{code}/upgrade" class="btn btn-sm" style="background:#f5f5f5;color:#d97706;border:1px solid #d97706">✦ AI予定作成</a>'
    combined = (event_cards + fee_cards) or '<div class="empty card">この月の予定はありません</div>'

    answer_link_card = ''
    if admin:
        view_url = f"{base_url()}t/{code}/view/{team['viewer_token']}"
        ans_url = f"{base_url()}t/{code}/answer/{team['viewer_token']}"
        answer_link_card = f'''
  <div class="card" style="margin-bottom:16px;border:1.5px solid #d97706;background:#fffdf7">
    <div style="font-weight:700;font-size:14px;margin-bottom:4px">🔗 メンバーへの共有リンク</div>
    <div style="font-size:12px;color:#666;margin-bottom:12px">ふだんの連絡アプリに貼るだけ。メンバーは登録不要で開けます。</div>
    <div style="font-size:12px;font-weight:700;color:#111;margin-bottom:4px">📅 予定表を見せる</div>
    <div style="font-size:11px;color:#888;margin-bottom:6px">カレンダーと予定をそのまま見てもらえます</div>
    <div style="background:#fff;border:1px solid #eee;border-radius:8px;padding:8px 10px;font-size:11px;color:#374151;word-break:break-all;font-family:monospace" id="view-url">{view_url}</div>
    <button type="button" onclick="navigator.clipboard.writeText(document.getElementById('view-url').textContent).then(function(){{var b=document.getElementById('view-copy');b.textContent='コピーしました ✓';b.style.background='#16a34a';setTimeout(function(){{b.textContent='予定表リンクをコピー';b.style.background='#111'}},2000)}})" id="view-copy" style="width:100%;margin-top:8px;padding:10px;border:none;border-radius:8px;background:#111;color:#fff;font-size:13px;font-weight:700;cursor:pointer">予定表リンクをコピー</button>
    <div style="height:1px;background:#f0e6d0;margin:14px 0"></div>
    <div style="font-size:12px;font-weight:700;color:#111;margin-bottom:4px">📣 出欠を答えてもらう</div>
    <div style="font-size:11px;color:#888;margin-bottom:6px">開いて出席/欠席を押すだけ。自動で集計されます</div>
    <div style="background:#fff;border:1px solid #eee;border-radius:8px;padding:8px 10px;font-size:11px;color:#374151;word-break:break-all;font-family:monospace" id="ans-url">{ans_url}</div>
    <button type="button" onclick="navigator.clipboard.writeText(document.getElementById('ans-url').textContent).then(function(){{var b=document.getElementById('ans-copy');b.textContent='コピーしました ✓';b.style.background='#16a34a';setTimeout(function(){{b.textContent='出欠リンクをコピー';b.style.background='#d97706'}},2000)}})" id="ans-copy" style="width:100%;margin-top:8px;padding:10px;border:none;border-radius:8px;background:#d97706;color:#fff;font-size:13px;font-weight:700;cursor:pointer">出欠リンクをコピー</button>
  </div>'''

    body = f'''
<div class="container">
  <div class="row" style="margin-bottom:16px">
    <div><span class="section-label">スケジュール</span></div>
    <div style="display:flex;gap:8px;margin-left:auto">{excel_btn}{ai_btn}{new_btn}</div>
  </div>
  {answer_link_card}
  <div class="card" style="margin-bottom:16px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
      <a href="/t/{code}/schedule?y={py}&m={pm}" style="width:36px;height:36px;display:flex;align-items:center;justify-content:center;border-radius:8px;background:#f1f4f9;color:#333;font-size:18px;text-decoration:none;flex-shrink:0">‹</a>
      <div style="display:flex;align-items:center;gap:8px">
        <span style="font-weight:700;font-size:16px">{vy}年{vm}月</span>
        {today_btn}
      </div>
      <a href="/t/{code}/schedule?y={ny}&m={nm}" style="width:36px;height:36px;display:flex;align-items:center;justify-content:center;border-radius:8px;background:#f1f4f9;color:#333;font-size:18px;text-decoration:none;flex-shrink:0">›</a>
    </div>
    {calendar_html}
  </div>
  {combined}
  {'<div style="margin-top:8px"><a href="/t/' + code + '/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>' if admin else ''}
</div>
<script>
function scrollToDate(date) {{
  var el = document.getElementById('ev-' + date);
  if (el) el.scrollIntoView({{behavior:'smooth', block:'center'}});
}}
</script>'''
    return page('スケジュール', body, code, active='schedule')

@app.route('/t/<code>/admin/schedule/excel')
def admin_schedule_excel(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='schedule')

    now = datetime.now(JST)
    try:
        vy = int(request.args.get('y', now.year))
        vm = int(request.args.get('m', now.month))
        vm = max(1, min(12, vm))
    except Exception:
        vy, vm = now.year, now.month

    month_start = f'{vy}-{vm:02d}-01'
    ny, nm = (vy + 1, 1) if vm == 12 else (vy, vm + 1)
    month_end = f'{ny}-{nm:02d}-01'

    conn = get_db()
    events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? AND event_date<? ORDER BY event_date, event_time',
        (team['id'], month_start, month_end)
    ).fetchall()

    WD = ['月', '火', '水', '木', '金', '土', '日']
    rows = [('日付', '曜日', 'タイトル', '開始', '終了', '場所', '備考', '出席数', '欠席数')]
    for ev in events:
        try:
            d = datetime.strptime(ev['event_date'], '%Y-%m-%d')
            date_str = f"{d.month}/{d.day}"
            wd = WD[d.weekday()]
        except Exception:
            date_str = ev['event_date']
            wd = ''
        rsvps = conn.execute('SELECT status FROM rsvps WHERE event_id=?', (ev['id'],)).fetchall()
        attending = sum(1 for r in rsvps if r['status'] == 'attending')
        absent = sum(1 for r in rsvps if r['status'] == 'absent')
        rows.append((
            date_str, wd,
            ev['title'] or '',
            ev['event_time'] or '',
            ev['end_time'] or '',
            ev['location'] or '',
            ev['note'] or '',
            attending, absent
        ))
    conn.close()

    filename = f"スケジュール_{vy}年{vm}月.xlsx"
    return excel_response(rows, filename)


@app.route('/t/<code>/admin/ai-schedule', methods=['GET', 'POST'])
def admin_ai_schedule(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='schedule')

    import json as _json
    sess_key = f'ai_sched_{code}'
    now = datetime.now(JST)
    default_ym = f'{now.year}-{now.month + 1:02d}' if now.month < 12 else f'{now.year + 1}-01'

    error = ''
    success_msg = ''

    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'generate':
            constraints = request.form.get('constraints', '').strip()
            target_ym = request.form.get('target_ym', default_ym).strip()
            if not constraints:
                error = '制約・要望を入力してください'
            elif not ANTHROPIC_API_KEY:
                error = 'ANTHROPIC_API_KEYが設定されていません'
            elif not HAS_ANTHROPIC:
                error = 'anthropicライブラリがインストールされていません'
            else:
                try:
                    ty, tm = int(target_ym[:4]), int(target_ym[5:7])
                    target_label = f'{ty}年{tm}月'
                    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
                    message = client.messages.create(
                        model='claude-haiku-4-5-20251001',
                        max_tokens=1500,
                        messages=[{
                            'role': 'user',
                            'content': f'''あなたはスポーツチームのスケジュール管理AIです。
以下の制約・要望をもとに、{target_label}の練習・試合スケジュールを提案してください。

制約・要望：
{constraints}

以下のJSON配列形式で返してください（6〜12件程度）：
[
  {{"title": "...", "date": "YYYY-MM-DD", "time": "HH:MM", "end_time": "HH:MM", "location": "...", "note": "...", "color": "#xxxxxx"}},
  ...
]

カラー選定ルール（colorフィールドに6桁HEXで設定）：
- 通常練習・レギュラー活動：#3b82f6（青）
- 試合・公式戦：#ef4444（赤）
- 練習試合・オープン戦：#f97316（橙）
- 遠征・合宿・特別イベント：#22c55e（緑）
- 記念行事・特別な活動：#a855f7（紫）
- その他・判断できない場合：#6b7280（グレー）

その他ルール：
- dateは{target_label}内（{ty}-{tm:02d}-01〜{ty}-{tm:02d}-28以降最終日）に収める
- timeとend_timeは不明なら空文字列
- locationは不明なら空文字列
- noteには参加対象（男子・女子・全体など）や特記事項を簡潔に書く
- JSONのみ返す。説明・マークダウン不要。'''
                        }]
                    )
                    text = message.content[0].text.strip()
                    if text.startswith('```'):
                        text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
                    events = _json.loads(text)
                    if not isinstance(events, list):
                        raise ValueError('not a list')
                    session[sess_key] = {
                        'events': events,
                        'constraints': constraints,
                        'target_ym': target_ym
                    }
                    return redirect(url_for('admin_ai_schedule', code=code))
                except Exception as e:
                    error = f'AI生成に失敗しました: {str(e)}'

        elif action == 'register_one':
            idx = request.form.get('idx', '')
            sess = session.get(sess_key, {})
            events = sess.get('events', [])
            try:
                idx = int(idx)
                if 0 <= idx < len(events):
                    ev = events[idx]
                    title = str(ev.get('title', '')).strip()
                    date = str(ev.get('date', '')).strip()
                    time = str(ev.get('time', '')).strip()
                    end_time = str(ev.get('end_time', '')).strip()
                    location = str(ev.get('location', '')).strip()
                    note = str(ev.get('note', '')).strip()
                    if title and date:
                        ev_color = str(ev.get('color', '')).strip()
                        conn = get_db()
                        conn.execute(
                            'INSERT INTO events (id,team_id,title,event_date,event_time,location,note,created_at,end_date,end_time,event_color) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                            (new_id(), team['id'], title, date, time, location, note, now_str(), '', end_time, ev_color)
                        )
                        conn.commit()
                        conn.close()
                        events.pop(idx)
                        sess['events'] = events
                        session[sess_key] = sess
            except Exception:
                pass
            return redirect(url_for('admin_ai_schedule', code=code))

        elif action == 'register_all':
            sess = session.get(sess_key, {})
            for ev in sess.get('events', []):
                title = str(ev.get('title', '')).strip()
                date = str(ev.get('date', '')).strip()
                time = str(ev.get('time', '')).strip()
                end_time = str(ev.get('end_time', '')).strip()
                location = str(ev.get('location', '')).strip()
                note = str(ev.get('note', '')).strip()
                if title and date:
                    ev_color = str(ev.get('color', '')).strip()
                    conn = get_db()
                    conn.execute(
                        'INSERT INTO events (id,team_id,title,event_date,event_time,location,note,created_at,end_date,end_time,event_color) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                        (new_id(), team['id'], title, date, time, location, note, now_str(), '', end_time, ev_color)
                    )
                    conn.commit()
                    conn.close()
            session.pop(sess_key, None)
            return redirect(url_for('schedule', code=code))

        elif action == 'clear':
            session.pop(sess_key, None)
            return redirect(url_for('admin_ai_schedule', code=code))

    sess = session.get(sess_key, {})
    pending_events = sess.get('events', [])
    saved_constraints = sess.get('constraints', '')
    saved_ym = sess.get('target_ym', default_ym)

    event_cards = ''
    for i, ev in enumerate(pending_events):
        title = str(ev.get('title', ''))
        date = str(ev.get('date', ''))
        time = str(ev.get('time', ''))
        end_time = str(ev.get('end_time', ''))
        location = str(ev.get('location', ''))
        note = str(ev.get('note', ''))
        ev_color = str(ev.get('color', '#6b7280')).strip() or '#6b7280'
        time_str = ''
        if time:
            time_str = f'　{time}〜{end_time}' if end_time else f'　{time}'
        detail = ''
        if location:
            detail += f'<div style="font-size:13px;color:#666;margin-top:2px">📍 {location}</div>'
        if note:
            detail += f'<div style="font-size:13px;color:#666;margin-top:2px">{note}</div>'
        event_cards += f'''
        <div class="card-sm" style="border-left:4px solid {ev_color}">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px">
            <div style="flex:1;min-width:0">
              <div style="display:flex;align-items:center;gap:8px">
                <div style="width:10px;height:10px;border-radius:50%;background:{ev_color};flex-shrink:0"></div>
                <div style="font-weight:700;font-size:16px">{title}</div>
              </div>
              <div style="font-size:13px;color:#666;margin-top:2px">{fmt_date(date)}{time_str}</div>
              {detail}
            </div>
            <form method="POST" style="flex-shrink:0">
              <input type="hidden" name="action" value="register_one">
              <input type="hidden" name="idx" value="{i}">
              <button class="btn btn-sm btn-blue" type="submit">登録</button>
            </form>
          </div>
        </div>'''

    gen_section = ''
    if pending_events:
        gen_section = f'''
        <div class="card" style="margin-top:16px;border:2px solid #d97706">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
            <div style="font-weight:700;color:#d97706">✦ AI生成スケジュール（{len(pending_events)}件）</div>
            <form method="POST" style="margin:0">
              <input type="hidden" name="action" value="clear">
              <button class="btn btn-sm btn-gray" type="submit" style="font-size:12px">クリア</button>
            </form>
          </div>
          <div style="font-size:12px;color:#888;margin-bottom:12px">各予定の「登録」ボタンで個別登録、または下の「全て登録」で一括登録できます。</div>
          {event_cards}
          <form method="POST" style="margin-top:16px">
            <input type="hidden" name="action" value="register_all">
            <button class="btn btn-blue btn-block" type="submit">全て登録してスケジュールへ →</button>
          </form>
        </div>'''

    body = f'''
<div class="container" style="max-width:560px">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
    <span class="section-label" style="color:#d97706">✦ AI予定作成</span>
    <span style="font-size:11px;background:#fef3c7;color:#d97706;padding:3px 8px;border-radius:20px">Pro機能</span>
  </div>
  {f'<div class="msg-err">{error}</div>' if error else ''}
  <div class="card">
    <form method="POST">
      <input type="hidden" name="action" value="generate">
      <label>対象月</label>
      <input type="month" name="target_ym" value="{saved_ym}" required style="font-size:16px">
      <label style="margin-top:12px">制約・要望 *</label>
      <textarea name="constraints" rows="6" placeholder="例：
・体育館は土日のみ使用可（第2日曜は使用不可）
・7月15日は外部コーチ不在
・男子チームは7月下旬に練習試合希望
・女子チームは毎週水曜放課後に自主練習あり
・月4回の全体練習を組み込む" style="font-size:13px">{saved_constraints}</textarea>
      <button class="btn btn-block" type="submit" style="background:#d97706;color:#fff;margin-top:8px">✦ AIでスケジュール生成</button>
    </form>
  </div>
  {gen_section}
  <div style="display:flex;justify-content:space-between;margin-top:12px">
    <a href="/t/{code}/schedule" style="font-size:13px;color:#888">← スケジュールに戻る</a>
    <a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>
  </div>
</div>'''
    return page('AI予定作成', body, code, active='schedule')


@app.route('/t/<code>/rsvp/<event_id>', methods=['POST'])
def rsvp(code, event_id):
    member = get_member(code)
    if not member:
        return redirect(url_for('team_portal', code=code))
    status = request.form.get('status', 'attending')
    conn = get_db()
    conn.execute('''
        INSERT INTO rsvps (id,event_id,member_name,status,updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(event_id,member_name) DO UPDATE SET status=excluded.status, updated_at=excluded.updated_at
    ''', (new_id(), event_id, member, status, now_str()))
    conn.commit()
    conn.close()
    return redirect(url_for('schedule', code=code))


# ── 無登録の出欠回答ページ（メンバーはリンクを開いてタップするだけ）────────────

def _answer_invalid_page():
    return render_template_string('<!DOCTYPE html><html lang="ja"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Rak</title></head><body style="font-family:sans-serif;text-align:center;padding:60px 20px;color:#888"><p>リンクが無効か、期限切れです。</p></body></html>'), 404


@app.route('/t/<code>/answer/<token>')
def answer_page(code, token):
    """出欠回答ページ（公開・ログイン不要）。token=viewer_token で認証"""
    import html as _html
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()

    # 名前選択（?name=... が来たらクッキーに保存してリダイレクト）
    pick = request.args.get('name', '').strip()
    if pick:
        r = redirect(f'/t/{code}/answer/{token}')
        r.set_cookie(f'rak_ans_{code}', pick, max_age=60 * 60 * 24 * 180, samesite='Lax')
        return r
    if request.args.get('reset'):
        r = redirect(f'/t/{code}/answer/{token}')
        r.delete_cookie(f'rak_ans_{code}')
        return r

    my_name = request.cookies.get(f'rak_ans_{code}', '')
    today = datetime.now(JST).strftime('%Y-%m-%d')
    conn = get_db()
    members = conn.execute('SELECT name FROM members WHERE team_id=? ORDER BY CAST(number AS INTEGER), name', (team['id'],)).fetchall()
    events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? ORDER BY event_date,event_time LIMIT 20',
        (team['id'], today)).fetchall()
    my_rsvps = {}
    if my_name:
        for r in conn.execute('SELECT event_id,status FROM rsvps WHERE member_name=?', (my_name,)).fetchall():
            my_rsvps[r['event_id']] = r['status']
    conn.close()

    head = f'''<!DOCTYPE html><html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}{FONT}
<title>{_html.escape(team["name"] or "チーム")}｜出欠の回答</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans JP',-apple-system,sans-serif;background:#f7f8fa;color:#111;padding:20px 16px 60px;max-width:520px;margin:0 auto}}
.head{{text-align:center;margin-bottom:20px}}
.head .tm{{font-size:13px;color:#888}}
.head h1{{font-size:20px;margin-top:4px}}
.card{{background:#fff;border-radius:14px;padding:16px;margin-bottom:12px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
.ev-title{{font-weight:700;font-size:16px}}
.ev-meta{{font-size:13px;color:#666;margin-top:2px}}
.btns{{display:flex;gap:10px;margin-top:14px}}
.btns button{{flex:1;padding:13px;border-radius:10px;font-size:15px;font-weight:700;border:2px solid #e5e7eb;background:#fff;color:#9ca3af;cursor:pointer}}
.btns button.on-go{{border-color:#16a34a;background:#f0fdf4;color:#16a34a}}
.btns button.on-no{{border-color:#dc2626;background:#fef2f2;color:#dc2626}}
.namebox{{text-align:center;padding:8px 0}}
.namebox a{{display:block;padding:14px;margin:8px 0;border:1.5px solid #e5e7eb;border-radius:10px;text-decoration:none;color:#111;font-weight:600;font-size:15px}}
.who{{text-align:center;font-size:13px;color:#666;margin-bottom:14px}}
.who b{{color:#111}}
.who a{{color:#d97706;font-size:12px;margin-left:6px}}
.foot{{text-align:center;font-size:11px;color:#bbb;margin-top:24px}}
input[type=text]{{width:100%;padding:12px;border:1.5px solid #e5e7eb;border-radius:10px;font-size:15px;margin-top:8px}}
.subbtn{{width:100%;padding:13px;border:none;border-radius:10px;background:#111;color:#fff;font-size:15px;font-weight:700;margin-top:10px;cursor:pointer}}
</style></head><body>
<div class="head"><div class="tm">{_html.escape(team["name"] or "チーム")}</div><h1>出欠の回答</h1></div>'''

    wd = ['月','火','水','木','金','土','日']
    def dl(d, t):
        try:
            from datetime import date as _d
            x = _d.fromisoformat(d)
            s = f'{x.month}月{x.day}日（{wd[x.weekday()]}）'
            return s + (f'　{t}' if t else '')
        except Exception:
            return d

    if not my_name:
        # 名前選択画面
        opts = ''.join(
            f'<a href="/t/{code}/answer/{token}?name={_html.escape(m["name"])}">{_html.escape(m["name"])}</a>'
            for m in members)
        if not opts:
            opts = '<div style="color:#888;font-size:13px;padding:8px">名簿が未登録です。下に名前を入力してください。</div>'
        body = f'''
<div class="card">
  <div style="font-weight:700;margin-bottom:6px">あなたのお名前を選んでください</div>
  <div style="font-size:12px;color:#888;margin-bottom:8px">登録不要。選ぶだけで回答できます。</div>
  <div class="namebox">{opts}</div>
  <form method="GET" action="/t/{code}/answer/{token}">
    <input type="text" name="name" placeholder="名簿にない方はここに入力" maxlength="20">
    <button class="subbtn" type="submit">この名前で回答する</button>
  </form>
</div>
<div class="foot">Powered by Rak</div></body></html>'''
        return head + body

    # 回答画面
    ev_html = ''
    if not events:
        ev_html = '<div class="card" style="text-align:center;color:#888">回答する予定はまだありません</div>'
    for ev in events:
        mode = ev['rsvp_mode'] or 'both'
        st = my_rsvps.get(ev['id'], '')
        go_cls = 'on-go' if st == 'attending' else ''
        no_cls = 'on-no' if st == 'absent' else ''
        meta = f'{dl(ev["event_date"], ev["event_time"])}{("　📍" + _html.escape(ev["location"])) if ev["location"] else ""}'
        if mode == 'none':
            buttons = '<div style="font-size:12px;color:#9ca3af;margin-top:10px">この予定は出欠の回答は不要です（連絡のみ）</div>'
        elif mode == 'absent_only':
            # 基本は出席。欠席する人だけ押す
            buttons = f'''
  <div style="font-size:12px;color:#666;margin-top:8px">基本は参加でお願いします。休む場合だけ下を押してください。</div>
  <form method="POST" action="/t/{code}/answer/{token}/rsvp/{ev["id"]}" class="btns">
    <button name="status" value="absent" class="{no_cls}" type="submit" style="flex:1">{'✓ 欠席で連絡済み' if st=='absent' else '欠席します'}</button>
    {('<button name="status" value="attending" type="submit" style="flex:0 0 auto;border-color:#16a34a;color:#16a34a">やっぱり出席</button>') if st=='absent' else ''}
  </form>'''
        else:  # both
            buttons = f'''
  <form method="POST" action="/t/{code}/answer/{token}/rsvp/{ev["id"]}" class="btns">
    <button name="status" value="attending" class="{go_cls}" type="submit">{'✓ ' if st=='attending' else ''}出席</button>
    <button name="status" value="absent" class="{no_cls}" type="submit">{'✓ ' if st=='absent' else ''}欠席</button>
  </form>'''
        ev_html += f'''
<div class="card">
  <div class="ev-title">{_html.escape(ev["title"])}</div>
  <div class="ev-meta">{meta}</div>
  {buttons}
</div>'''
    body = f'''
<div class="who"><b>{_html.escape(my_name)}</b> さんとして回答中<a href="/t/{code}/answer/{token}?reset=1">（変更）</a></div>
{ev_html}
<div class="foot">回答すると主催者にすぐ反映されます　·　Powered by Rak</div></body></html>'''
    return head + body


@app.route('/t/<code>/answer/<token>/rsvp/<event_id>', methods=['POST'])
def answer_rsvp(code, token, event_id):
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()
    name = request.cookies.get(f'rak_ans_{code}', '').strip()
    if not name:
        return redirect(f'/t/{code}/answer/{token}')
    status = request.form.get('status', 'attending')
    if status not in ('attending', 'absent'):
        status = 'attending'
    conn = get_db()
    # イベントがこのチームのものか検証
    ev = conn.execute('SELECT id FROM events WHERE id=? AND team_id=?', (event_id, team['id'])).fetchone()
    if ev:
        conn.execute('''
            INSERT INTO rsvps (id,event_id,member_name,status,updated_at)
            VALUES (?,?,?,?,?)
            ON CONFLICT(event_id,member_name) DO UPDATE SET status=excluded.status, updated_at=excluded.updated_at
        ''', (new_id(), event_id, name, status, now_str()))
        conn.commit()
    conn.close()
    return redirect(f'/t/{code}/answer/{token}')


@app.route('/t/<code>/order-answer/<token>/<form_id>')
def order_answer_page(code, token, form_id):
    """注文フォーム回答ページ（公開・ログイン不要）。token=viewer_token で認証"""
    import html as _html
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()

    # 名前選択（出欠リンクと同じクッキーを共有）
    pick = request.args.get('name', '').strip()
    if pick:
        r = redirect(f'/t/{code}/order-answer/{token}/{form_id}')
        r.set_cookie(f'rak_ans_{code}', pick, max_age=60 * 60 * 24 * 180, samesite='Lax')
        return r
    if request.args.get('reset'):
        r = redirect(f'/t/{code}/order-answer/{token}/{form_id}')
        r.delete_cookie(f'rak_ans_{code}')
        return r

    conn = get_db()
    form = conn.execute('SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])).fetchone()
    if not form:
        conn.close()
        return _answer_invalid_page()
    fields = conn.execute('SELECT * FROM order_form_fields WHERE form_id=? ORDER BY sort_order', (form_id,)).fetchall()
    photos = conn.execute('SELECT * FROM order_form_photos WHERE form_id=? ORDER BY uploaded_at', (form_id,)).fetchall()
    members = conn.execute('SELECT name FROM members WHERE team_id=? ORDER BY CAST(number AS INTEGER), name', (team['id'],)).fetchall()

    my_name = request.cookies.get(f'rak_ans_{code}', '')
    my_values = {}
    if my_name:
        mr = conn.execute('SELECT id FROM order_responses WHERE form_id=? AND member_name=?', (form_id, my_name)).fetchone()
        if mr:
            for v in conn.execute('SELECT field_id,value FROM order_response_values WHERE response_id=?', (mr['id'],)).fetchall():
                my_values[v['field_id']] = v['value']
    conn.close()

    head = f'''<!DOCTYPE html><html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}{FONT}
<title>{_html.escape(team["name"] or "チーム")}｜{_html.escape(form["title"])}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans JP',-apple-system,sans-serif;background:#f7f8fa;color:#111;padding:20px 16px 60px;max-width:520px;margin:0 auto}}
.head{{text-align:center;margin-bottom:20px}}
.head .tm{{font-size:13px;color:#888}}
.head h1{{font-size:20px;margin-top:4px}}
.card{{background:#fff;border-radius:14px;padding:16px;margin-bottom:12px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
.namebox a{{display:block;padding:14px;margin:8px 0;border:1.5px solid #e5e7eb;border-radius:10px;text-decoration:none;color:#111;font-weight:600;font-size:15px;text-align:center}}
.who{{text-align:center;font-size:13px;color:#666;margin-bottom:14px}}
.who b{{color:#111}}
.who a{{color:#d97706;font-size:12px;margin-left:6px}}
.foot{{text-align:center;font-size:11px;color:#bbb;margin-top:24px}}
label{{display:block;font-size:13px;color:#555;font-weight:600;margin-top:14px;margin-bottom:4px}}
input[type=text],select{{width:100%;padding:12px;border:1.5px solid #e5e7eb;border-radius:10px;font-size:15px;background:#fff}}
.subbtn{{width:100%;padding:14px;border:none;border-radius:10px;background:#111;color:#fff;font-size:15px;font-weight:700;margin-top:16px;cursor:pointer}}
.dl-meta{{font-size:13px;color:#f59e0b;margin-top:4px}}
.desc{{font-size:13px;color:#666;margin-top:6px;line-height:1.6}}
.fphoto{{width:100%;border-radius:10px;border:1.5px solid #eee;margin-bottom:10px;display:block}}
.okmsg{{background:#f0fdf4;border:1px solid #bbf7d0;color:#16a34a;border-radius:10px;padding:10px 12px;font-size:13px;font-weight:600;margin-bottom:12px}}
</style></head><body>
<div class="head"><div class="tm">{_html.escape(team["name"] or "チーム")}</div><h1>{_html.escape(form["title"])}</h1></div>'''

    if not my_name:
        opts = ''.join(
            f'<a href="/t/{code}/order-answer/{token}/{form_id}?name={_html.escape(m["name"])}">{_html.escape(m["name"])}</a>'
            for m in members)
        if not opts:
            opts = '<div style="color:#888;font-size:13px;padding:8px;text-align:center">名簿が未登録です。下に名前を入力してください。</div>'
        body = f'''
<div class="card">
  <div style="font-weight:700;margin-bottom:6px">あなたのお名前を選んでください</div>
  <div style="font-size:12px;color:#888;margin-bottom:8px">登録不要。選ぶだけで回答できます。</div>
  <div class="namebox">{opts}</div>
  <form method="GET" action="/t/{code}/order-answer/{token}/{form_id}">
    <input type="text" name="name" placeholder="名簿にない方はここに入力" maxlength="20">
    <button class="subbtn" type="submit">この名前で回答する</button>
  </form>
</div>
<div class="foot">Powered by Rak</div></body></html>'''
        return head + body

    # 回答フォーム
    photos_html = ''.join(f'<img src="/uploads/{p["id"]}" class="fphoto">' for p in photos)
    desc_html = f'<div class="desc">{_html.escape(form["description"])}</div>' if form['description'] else ''
    dl_html = f'<div class="dl-meta">期限：{fmt_date(form["deadline"])}</div>' if form['deadline'] else ''
    answered = bool(my_values)
    field_inputs = ''
    for field in fields:
        cur = my_values.get(field['id'], '')
        if field['field_type'] == 'select' and field['options']:
            opts_list = [o.strip() for o in field['options'].split(',') if o.strip()]
            options_html = '<option value="">選択してください</option>' + ''.join(
                f'<option value="{_html.escape(o)}" {"selected" if cur == o else ""}>{_html.escape(o)}</option>'
                for o in opts_list)
            field_inputs += f'<label>{_html.escape(field["label"])}</label><select name="field_{field["id"]}">{options_html}</select>'
        else:
            field_inputs += (
                f'<label>{_html.escape(field["label"])}</label>'
                f'<input type="text" name="field_{field["id"]}" value="{_html.escape(cur)}" placeholder="入力してください">'
            )
    if not fields:
        form_html = '<div style="text-align:center;color:#888;font-size:13px;padding:12px">この注文フォームはまだ準備中です</div>'
    else:
        form_html = (
            f'<form method="POST" action="/t/{code}/order-answer/{token}/{form_id}/submit">'
            f'{field_inputs}<button class="subbtn" type="submit">{"回答を更新する" if answered else "送信する"}</button></form>'
        )
    body = f'''
<div class="who"><b>{_html.escape(my_name)}</b> さんとして回答中<a href="/t/{code}/order-answer/{token}/{form_id}?reset=1">（変更）</a></div>
<div class="card">
  {'<div class="okmsg">' + _CHK + ' 回答済みです。修正して再送信できます。</div>' if answered else ''}
  {desc_html}
  {dl_html}
  {photos_html}
  {form_html}
</div>
<div class="foot">回答すると主催者にすぐ反映されます　·　Powered by Rak</div></body></html>'''
    return head + body


@app.route('/t/<code>/order-answer/<token>/<form_id>/submit', methods=['POST'])
def order_answer_submit(code, token, form_id):
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()
    name = request.cookies.get(f'rak_ans_{code}', '').strip()
    if not name:
        return redirect(f'/t/{code}/order-answer/{token}/{form_id}')
    conn = get_db()
    form = conn.execute('SELECT id FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])).fetchone()
    if not form:
        conn.close()
        return _answer_invalid_page()
    fields = conn.execute('SELECT * FROM order_form_fields WHERE form_id=? ORDER BY sort_order', (form_id,)).fetchall()
    resp = conn.execute('SELECT id FROM order_responses WHERE form_id=? AND member_name=?', (form_id, name)).fetchone()
    if resp:
        resp_id = resp['id']
        conn.execute('UPDATE order_responses SET submitted_at=? WHERE id=?', (now_str(), resp_id))
        conn.execute('DELETE FROM order_response_values WHERE response_id=?', (resp_id,))
    else:
        resp_id = new_id()
        conn.execute('INSERT INTO order_responses VALUES (?,?,?,?)', (resp_id, form_id, name, now_str()))
    for field in fields:
        value = request.form.get(f'field_{field["id"]}', '').strip()
        conn.execute('INSERT INTO order_response_values VALUES (?,?,?,?)', (new_id(), resp_id, field['id'], value))
    conn.commit()
    conn.close()
    return redirect(f'/t/{code}/order-answer/{token}/{form_id}')


@app.route('/t/<code>/pay-answer/<token>')
def pay_answer_page(code, token):
    """集金の支払い自己申告ページ（公開・ログイン不要）。token=viewer_token で認証"""
    import html as _html
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()

    # 名前選択（出欠・注文リンクと同じクッキーを共有）
    pick = request.args.get('name', '').strip()
    if pick:
        r = redirect(f'/t/{code}/pay-answer/{token}')
        r.set_cookie(f'rak_ans_{code}', pick, max_age=60 * 60 * 24 * 180, samesite='Lax')
        return r
    if request.args.get('reset'):
        r = redirect(f'/t/{code}/pay-answer/{token}')
        r.delete_cookie(f'rak_ans_{code}')
        return r

    my_name = request.cookies.get(f'rak_ans_{code}', '')
    conn = get_db()
    members = conn.execute('SELECT name FROM members WHERE team_id=? ORDER BY CAST(number AS INTEGER), name', (team['id'],)).fetchall()
    fees = conn.execute('SELECT * FROM fees WHERE team_id=? ORDER BY due_date,created_at', (team['id'],)).fetchall()
    my_pay = {}
    if my_name:
        for p in conn.execute('SELECT fee_id,paid,reported FROM fee_payments WHERE member_name=?', (my_name,)).fetchall():
            my_pay[p['fee_id']] = p
    conn.close()

    head = f'''<!DOCTYPE html><html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
{FAVICON_LINK}{FONT}
<title>{_html.escape(team["name"] or "チーム")}｜集金の支払い</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans JP',-apple-system,sans-serif;background:#f7f8fa;color:#111;padding:20px 16px 60px;max-width:520px;margin:0 auto}}
.head{{text-align:center;margin-bottom:20px}}
.head .tm{{font-size:13px;color:#888}}
.head h1{{font-size:20px;margin-top:4px}}
.card{{background:#fff;border-radius:14px;padding:16px;margin-bottom:12px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
.fee-title{{font-weight:700;font-size:16px}}
.fee-meta{{font-size:14px;color:#555;margin-top:2px}}
.namebox a{{display:block;padding:14px;margin:8px 0;border:1.5px solid #e5e7eb;border-radius:10px;text-decoration:none;color:#111;font-weight:600;font-size:15px;text-align:center}}
.who{{text-align:center;font-size:13px;color:#666;margin-bottom:14px}}
.who b{{color:#111}}
.who a{{color:#d97706;font-size:12px;margin-left:6px}}
.foot{{text-align:center;font-size:11px;color:#bbb;margin-top:24px}}
.paybtn{{width:100%;padding:13px;border-radius:10px;font-size:15px;font-weight:700;border:2px solid #16a34a;background:#16a34a;color:#fff;cursor:pointer;margin-top:12px}}
.subbtn{{width:100%;padding:13px;border:none;border-radius:10px;background:#111;color:#fff;font-size:15px;font-weight:700;margin-top:10px;cursor:pointer}}
input[type=text]{{width:100%;padding:12px;border:1.5px solid #e5e7eb;border-radius:10px;font-size:15px;margin-top:8px}}
.badge-paid{{display:inline-block;background:#f0fdf4;border:1px solid #bbf7d0;color:#16a34a;border-radius:8px;padding:8px 12px;font-size:14px;font-weight:700;margin-top:12px;text-align:center;width:100%}}
.badge-rep{{display:inline-block;background:#fffbeb;border:1px solid #fde68a;color:#b45309;border-radius:8px;padding:8px 12px;font-size:13px;font-weight:600;margin-top:12px;text-align:center;width:100%}}
</style></head><body>
<div class="head"><div class="tm">{_html.escape(team["name"] or "チーム")}</div><h1>集金の支払い</h1></div>'''

    if not my_name:
        opts = ''.join(
            f'<a href="/t/{code}/pay-answer/{token}?name={_html.escape(m["name"])}">{_html.escape(m["name"])}</a>'
            for m in members)
        if not opts:
            opts = '<div style="color:#888;font-size:13px;padding:8px;text-align:center">名簿が未登録です。下に名前を入力してください。</div>'
        body = f'''
<div class="card">
  <div style="font-weight:700;margin-bottom:6px">あなたのお名前を選んでください</div>
  <div style="font-size:12px;color:#888;margin-bottom:8px">登録不要。選ぶだけで申告できます。</div>
  <div class="namebox">{opts}</div>
  <form method="GET" action="/t/{code}/pay-answer/{token}">
    <input type="text" name="name" placeholder="名簿にない方はここに入力" maxlength="20">
    <button class="subbtn" type="submit">この名前で申告する</button>
  </form>
</div>
<div class="foot">Powered by Rak</div></body></html>'''
        return head + body

    fee_html = ''
    if not fees:
        fee_html = '<div class="card" style="text-align:center;color:#888">支払う集金はまだありません</div>'
    for f in fees:
        p = my_pay.get(f['id'])
        paid = p['paid'] if p else 0
        reported = p['reported'] if p else 0
        meta = f'¥{f["amount"]:,}' + (f'　期限：{fmt_date(f["due_date"])}' if f['due_date'] else '')
        note_html = f'<div style="font-size:12px;color:#888;margin-top:4px">{_html.escape(f["note"])}</div>' if f['note'] else ''
        if paid:
            action = '<div class="badge-paid">' + _CHK + ' 支払い確認済み</div>'
        elif reported:
            action = '''<div class="badge-rep">⏳ 支払い申告ずみ（主催者の確認待ち）</div>
  <form method="POST" action="/t/{code}/pay-answer/{token}/report/{fid}" style="margin-top:6px"><button name="undo" value="1" type="submit" style="width:100%;padding:9px;border-radius:8px;border:1px solid #e5e7eb;background:#fff;color:#888;font-size:12px;cursor:pointer">申告を取り消す</button></form>'''.replace('{code}', code).replace('{token}', token).replace('{fid}', f['id'])
        else:
            action = f'''<form method="POST" action="/t/{code}/pay-answer/{token}/report/{f["id"]}">
    <button class="paybtn" type="submit">支払いました</button>
  </form>
  <div style="font-size:11px;color:#aaa;text-align:center;margin-top:6px">※ 実際のお支払い（現金・振込）の後に押してください</div>'''
        fee_html += f'''
<div class="card">
  <div class="fee-title">{_html.escape(f["title"])}</div>
  <div class="fee-meta">{meta}</div>
  {note_html}
  {action}
</div>'''
    body = f'''
<div class="who"><b>{_html.escape(my_name)}</b> さんとして申告中<a href="/t/{code}/pay-answer/{token}?reset=1">（変更）</a></div>
{fee_html}
<div class="foot">「支払いました」を押すと主催者に届きます　·　Powered by Rak</div></body></html>'''
    return head + body


@app.route('/t/<code>/pay-answer/<token>/report/<fee_id>', methods=['POST'])
def pay_answer_report(code, token, fee_id):
    team = get_team(code)
    if not team or team['viewer_token'] != token:
        return _answer_invalid_page()
    name = request.cookies.get(f'rak_ans_{code}', '').strip()
    if not name:
        return redirect(f'/t/{code}/pay-answer/{token}')
    undo = request.form.get('undo') == '1'
    conn = get_db()
    fee = conn.execute('SELECT id FROM fees WHERE id=? AND team_id=?', (fee_id, team['id'])).fetchone()
    if fee:
        if undo:
            conn.execute('''INSERT INTO fee_payments (id,fee_id,member_name,paid,paid_at,reported,reported_at)
                VALUES (?,?,?,0,'',0,'')
                ON CONFLICT(fee_id,member_name) DO UPDATE SET reported=0, reported_at=''
                WHERE fee_payments.paid=0''',
                (new_id(), fee_id, name))
        else:
            conn.execute('''INSERT INTO fee_payments (id,fee_id,member_name,paid,paid_at,reported,reported_at)
                VALUES (?,?,?,0,'',1,?)
                ON CONFLICT(fee_id,member_name) DO UPDATE SET reported=1, reported_at=excluded.reported_at''',
                (new_id(), fee_id, name, now_str()))
        conn.commit()
    conn.close()
    return redirect(f'/t/{code}/pay-answer/{token}')


# ── Notices ───────────────────────────────────────────────────────

@app.route('/t/<code>/notices')
def notices(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    ns = conn.execute(
        'SELECT * FROM notices WHERE team_id=? ORDER BY created_at DESC',
        (team['id'],)
    ).fetchall()

    cards = ''
    for n in ns:
        read_count = conn.execute('SELECT COUNT(*) FROM reads WHERE notice_id=?', (n['id'],)).fetchone()[0]
        is_read = bool(conn.execute('SELECT 1 FROM reads WHERE notice_id=? AND member_name=?',
                                    (n['id'], member)).fetchone()) if member else True
        cards += f'''
        <a href="/t/{code}/notices/{n['id']}" style="text-decoration:none;display:block">
          <div class="card-sm" style="{'opacity:.7' if is_read else ''}">
            <div class="row">
              <div style="flex:1">
                <div style="font-weight:700;color:#1a1a1a">{'📌 ' if not is_read else ''}{n['title']}</div>
                <div style="font-size:12px;color:#888;margin-top:2px">{fmt_datetime(n['created_at'])}</div>
              </div>
              <div>
                {'<span class="badge badge-blue">NEW</span>' if not is_read else f'<span class="badge badge-gray">既読 {read_count}</span>'}
              </div>
            </div>
          </div>
        </a>'''
    conn.close()

    new_btn = f'<a href="/t/{code}/admin/notices/new" class="btn btn-blue btn-sm">＋ お知らせ作成</a>' if admin else ''
    back_link = f'<a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>' if admin else f'<a href="/t/{code}/schedule" style="font-size:13px;color:#888">← ホームに戻る</a>'
    body = f'''
<div class="container">
  <div class="row" style="margin-bottom:16px">
    <span class="section-label">{_ICO_BELL_SM} お知らせ</span>
    <div style="margin-left:auto">{new_btn}</div>
  </div>
  {cards if ns else (f'<div class="empty card"><div style="margin-bottom:8px">{_SVG_EMPTY_BELL}</div><div style="font-weight:700;margin-bottom:4px">お知らせはまだありません</div><div style="font-size:12px;color:#aaa;margin-bottom:16px">最初のお知らせを送ってみましょう。AIが下書きを作れます。</div><div style="display:flex;gap:8px;justify-content:center"><a href="/t/{code}/admin/notices/new" class="btn btn-blue btn-sm">＋ 作成する</a><a href="/t/{code}/admin/ai" class="btn btn-outline btn-sm">✨ AIで下書き</a></div></div>' if admin else f'<div class="empty card"><div style="margin-bottom:8px">{_SVG_EMPTY_BELL}</div>お知らせはまだありません</div>')}
  {'<div style="margin-top:8px">' + back_link + '</div>' if admin else ''}
</div>'''
    return page('お知らせ', body, code, active='notices')

@app.route('/t/<code>/notices/<notice_id>')
def notice_detail(code, notice_id):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    n = conn.execute('SELECT * FROM notices WHERE id=? AND team_id=?', (notice_id, team['id'])).fetchone()
    if not n:
        conn.close()
        return redirect(url_for('notices', code=code))

    if member:
        conn.execute('''
            INSERT OR IGNORE INTO reads (notice_id,member_name,read_at) VALUES (?,?,?)
        ''', (notice_id, member, now_str()))
        conn.commit()

    readers = conn.execute('SELECT member_name, read_at FROM reads WHERE notice_id=? ORDER BY read_at', (notice_id,)).fetchall()
    conn.close()

    reader_list = ''
    if admin:
        reader_list = ''.join(f'<div style="font-size:13px;padding:6px 0;border-bottom:1px solid #f0f0f0;color:#555">{r["member_name"]} <span style="color:#aaa;font-size:11px">{fmt_datetime(r["read_at"])}</span></div>' for r in readers)
        reader_list = f'<hr class="divider"><div style="font-size:12px;font-weight:700;color:#d97706;margin-bottom:8px">既読 {len(readers)}名</div>{reader_list}'

    body = f'''
<div class="container">
  <div class="card">
    <div style="font-size:12px;color:#888;margin-bottom:8px">{fmt_datetime(n['created_at'])}</div>
    <h1 style="margin-bottom:16px">{n['title']}</h1>
    <div style="white-space:pre-wrap;line-height:1.8;color:#333">{n['body']}</div>
    {reader_list}
  </div>
  <div style="display:flex;justify-content:space-between;align-items:center">
    <a href="/t/{code}/notices" style="font-size:13px;color:#888">← お知らせ一覧</a>
    {'<a href="/t/' + code + '/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>' if admin else '<a href="/t/' + code + '/schedule" style="font-size:13px;color:#888">← ホームに戻る</a>'}
  </div>
</div>'''
    return page(n['title'], body, code, active='notices')


# ── Admin login ───────────────────────────────────────────────────

@app.route('/t/<code>/admin', methods=['GET', 'POST'])
def admin_login(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    if is_admin(code):
        return redirect(url_for('admin_dash', code=code))

    lock_key = f'login_fail_{code}'
    lock_until_key = f'login_lock_{code}'
    import time
    now_ts = time.time()

    # ロック中チェック
    lock_until = session.get(lock_until_key, 0)
    if now_ts < lock_until:
        remain = int(lock_until - now_ts)
        body = f'''<div class="container" style="max-width:400px;padding-top:60px">
  <div class="card"><div class="msg-err">ログインを{remain}秒間ブロックしています。しばらくお待ちください。</div></div></div>'''
        return page('管理者ログイン', body, code)

    error = ''
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if pw == team['admin_password']:
            session.pop(lock_key, None)
            session.pop(lock_until_key, None)
            session.permanent = True
            session[f'admin_{code}'] = True
            return redirect(url_for('admin_dash', code=code))
        fails = session.get(lock_key, 0) + 1
        session[lock_key] = fails
        if fails >= 10:
            session[lock_until_key] = now_ts + 900  # 15分ロック
            session[lock_key] = 0
            error = 'ログイン試行回数が上限を超えました。15分後に再試行してください。'
        else:
            error = f'パスワードが違います（{fails}/10回）'

    body = f'''
<div class="container" style="max-width:400px;padding-top:60px">
  <div class="card">
    <h1 style="margin-bottom:4px">管理者ログイン</h1>
    <p style="color:#666;font-size:13px;margin-bottom:16px">{team['name']}</p>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>パスワード</label>
      <input type="password" name="password" autofocus required>
      <button class="btn btn-blue btn-block" type="submit">ログイン</button>
    </form>
    <div style="text-align:center;margin-top:14px">
      <a href="/forgot-password" style="font-size:12px;color:#888">パスワードを忘れた方</a>
    </div>
  </div>
</div>'''
    return page('管理者ログイン', body, code)

@app.route('/t/<code>/admin/settings', methods=['GET', 'POST'])
def admin_settings(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    msg = ''
    error = ''

    if request.method == 'POST':
        action = request.form.get('action')
        conn = get_db()
        if action == 'delete_team':
            confirm_pw = request.form.get('confirm_password', '')
            if confirm_pw != team['admin_password']:
                error = 'パスワードが違います。チームの削除はできませんでした。'
                conn.close()
            else:
                tid = team['id']
                for tbl in ['rsvps','reads','fee_payments','order_responses','push_subscriptions',
                            'admin_memos','uniform_assignments','ledger','password_reset_tokens']:
                    try: conn.execute(f'DELETE FROM {tbl} WHERE team_id=?', (tid,))
                    except: pass
                for tbl in ['events','notices','members','fees','order_forms','uniforms']:
                    try: conn.execute(f'DELETE FROM {tbl} WHERE team_id=?', (tid,))
                    except: pass
                conn.execute('DELETE FROM teams WHERE id=?', (tid,))
                conn.commit()
                conn.close()
                session.pop(f'admin_{code}', None)
                return redirect('/?deleted=1')
        if action == 'name':
            new_name = request.form.get('name', '').strip()
            if not new_name or len(new_name) < 2:
                error = 'チーム名は2文字以上で入力してください'
            else:
                conn.execute('UPDATE teams SET name=? WHERE id=?', (new_name, team['id']))
                conn.commit()
                msg = 'チーム名を変更しました'
        elif action == 'email':
            new_email = request.form.get('email', '').strip()
            if not new_email:
                error = 'メールアドレスを入力してください'
            else:
                conn.execute('UPDATE teams SET admin_email=? WHERE id=?', (new_email, team['id']))
                conn.commit()
                msg = 'メールアドレスを変更しました'
        elif action == 'password':
            current_pw = request.form.get('current_password', '')
            new_pw = request.form.get('new_password', '').strip()
            if current_pw != team['admin_password']:
                error = '現在のパスワードが違います'
            elif len(new_pw) < 6:
                error = '新しいパスワードは6文字以上にしてください'
            elif not any(c.isalpha() for c in new_pw):
                error = '英字を1文字以上含めてください'
            elif not any(c.isdigit() for c in new_pw):
                error = '数字を1文字以上含めてください'
            else:
                conn.execute('UPDATE teams SET admin_password=? WHERE id=?', (new_pw, team['id']))
                conn.commit()
                msg = 'パスワードを変更しました'
        conn.close()
        team = get_team(code)

    body = f'''
<div class="container" style="max-width:480px">
  {f'<div class="msg-ok">{_CHK} {msg}</div>' if msg else ''}
  {f'<div class="msg-err">{error}</div>' if error else ''}

  <div class="card" style="margin-bottom:12px">
    <h2 style="margin-bottom:16px">チーム名の変更</h2>
    <form method="POST">
      <input type="hidden" name="action" value="name">
      <label>チーム名</label>
      <input type="text" name="name" value="{team['name']}" required>
      <button class="btn btn-blue btn-block" type="submit">変更する</button>
    </form>
  </div>

  <div class="card" style="margin-bottom:12px">
    <h2 style="margin-bottom:4px">メールアドレスの変更</h2>
    <div style="font-size:12px;color:#888;margin-bottom:16px">パスワードリセットに使用するメールアドレス</div>
    <form method="POST">
      <input type="hidden" name="action" value="email">
      <label>メールアドレス</label>
      <input type="email" name="email" value="{team['admin_email'] or ''}" placeholder="admin@example.com" required>
      <button class="btn btn-blue btn-block" type="submit">変更する</button>
    </form>
  </div>

  <div class="card" style="margin-bottom:12px">
    <h2 style="margin-bottom:4px">管理者パスワードの確認・変更</h2>
    <div style="font-size:12px;color:#888;margin-bottom:16px">現在のパスワードを入力して新しいパスワードに変更できます</div>
    <form method="POST">
      <input type="hidden" name="action" value="password">
      <label>現在のパスワード</label>
      <div style="position:relative">
        <input type="password" name="current_password" id="cur-pw" placeholder="現在のパスワード" required style="padding-right:44px">
        <button type="button" onclick="var i=document.getElementById('cur-pw');i.type=i.type==='password'?'text':'password';this.textContent=i.type==='password'?'表示':'隠す'" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;color:#888;font-size:12px;cursor:pointer;padding:4px">表示</button>
      </div>
      <label style="margin-top:14px">新しいパスワード</label>
      <div style="position:relative">
        <input type="password" name="new_password" id="new-pw" placeholder="英字・数字を含む6文字以上" required style="padding-right:44px">
        <button type="button" onclick="var i=document.getElementById('new-pw');i.type=i.type==='password'?'text':'password';this.textContent=i.type==='password'?'表示':'隠す'" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;color:#888;font-size:12px;cursor:pointer;padding:4px">表示</button>
      </div>
      <button class="btn btn-blue btn-block" type="submit">パスワードを変更する</button>
    </form>
  </div>

  <div class="card" style="margin-bottom:12px;background:#fafafa">
    <h2 style="margin-bottom:4px">ログアウト</h2>
    <div style="font-size:12px;color:#888;margin-bottom:16px">このデバイスの管理者セッションを終了します</div>
    <a href="/t/{code}/admin/logout" class="btn btn-outline btn-block">ログアウトする</a>
  </div>

  <div class="card" style="margin-bottom:12px;border-color:#fecaca;background:#fff5f5">
    <h2 style="margin-bottom:4px;color:#991b1b">チームを削除する</h2>
    <div style="font-size:12px;color:#888;margin-bottom:12px">チームと全データ（メンバー・予定・集金・お知らせ）を完全に削除します。この操作は取り消せません。</div>
    <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px;margin-bottom:14px;font-size:13px;color:#991b1b">
      ⚠️ 削除後30日以内にサーバーからデータが消去されます。有料プランは次回更新日前に解約してください。
    </div>
    <form method="POST" onsubmit="return confirm('本当にチームを削除しますか？この操作は取り消せません。')">
      <input type="hidden" name="action" value="delete_team">
      <label style="color:#991b1b">パスワードを入力して確認</label>
      <input type="password" name="confirm_password" placeholder="管理者パスワード" required>
      <button class="btn btn-block" type="submit" style="background:#dc2626;color:#fff;border:none;border-radius:8px;padding:11px;font-size:14px;font-weight:700;cursor:pointer">チームを完全に削除する</button>
    </form>
  </div>

  <div style="margin-top:8px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('設定', body, code, active='home')


@app.route('/t/<code>/admin/logout')
def admin_logout(code):
    session.pop(f'admin_{code}', None)
    return redirect(url_for('team_portal', code=code))


# ── Admin dashboard ───────────────────────────────────────────────

@app.route('/t/<code>/admin/dash')
def admin_dash(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    created = request.args.get('created') == '1'

    conn = get_db()
    today = datetime.now(JST).strftime('%Y-%m-%d')
    events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? ORDER BY event_date LIMIT 7',
        (team['id'], today)
    ).fetchall()
    notices = conn.execute(
        'SELECT * FROM notices WHERE team_id=? ORDER BY created_at DESC LIMIT 3',
        (team['id'],)
    ).fetchall()

    # 集金未払いリスト
    fees = conn.execute('SELECT * FROM fees WHERE team_id=? ORDER BY due_date, created_at', (team['id'],)).fetchall()
    unpaid_summary = []
    for f in fees:
        members_for_fee = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY name', (team['id'],)).fetchall()
        for m in members_for_fee:
            p = conn.execute('SELECT paid FROM fee_payments WHERE fee_id=? AND member_name=?', (f['id'], m['name'])).fetchone()
            if not p or not p['paid']:
                unpaid_summary.append({'fee_title': f['title'], 'member': m['name'], 'amount': f['amount'], 'due_date': f['due_date'], 'fee_id': f['id']})

    # 締切付き集金・注文フォーム
    upcoming_fees = conn.execute(
        "SELECT * FROM fees WHERE team_id=? AND due_date>=? AND due_date!='' ORDER BY due_date LIMIT 5",
        (team['id'], today)
    ).fetchall()
    upcoming_orders = conn.execute(
        "SELECT * FROM order_forms WHERE team_id=? AND deadline>=? AND deadline!='' ORDER BY deadline LIMIT 5",
        (team['id'], today)
    ).fetchall()

    members_all = conn.execute('SELECT name FROM members WHERE team_id=?', (team['id'],)).fetchall()
    member_names = [m['name'] for m in members_all]
    has_any_event = conn.execute('SELECT 1 FROM events WHERE team_id=? LIMIT 1', (team['id'],)).fetchone()
    has_any_notice = conn.execute('SELECT 1 FROM notices WHERE team_id=? LIMIT 1', (team['id'],)).fetchone()

    def get_no_answer(ev_id):
        answered = set(r['member_name'] for r in conn.execute('SELECT member_name FROM rsvps WHERE event_id=?', (ev_id,)).fetchall())
        return [n for n in member_names if n not in answered]

    # 今月のカレンダー用データ
    now_dt = datetime.now(JST)
    cy, cm = now_dt.year, now_dt.month
    cm_start = f'{cy}-{cm:02d}-01'
    cn_y, cn_m = (cy + 1, 1) if cm == 12 else (cy, cm + 1)
    cm_end = f'{cn_y}-{cn_m:02d}-01'
    cal_events = conn.execute(
        'SELECT * FROM events WHERE team_id=? AND event_date>=? AND event_date<?',
        (team['id'], cm_start, cm_end)
    ).fetchall()
    dash_ev_color_map = {}
    for ev in cal_events:
        c = ev['event_color'] or '#6b7280'
        cur = datetime.strptime(ev['event_date'], '%Y-%m-%d')
        end_d = datetime.strptime(ev['end_date'], '%Y-%m-%d') if ev['end_date'] else cur
        while cur <= end_d:
            ds = cur.strftime('%Y-%m-%d')
            dash_ev_color_map.setdefault(ds, [])
            if c not in dash_ev_color_map[ds]:
                dash_ev_color_map[ds].append(c)
            cur += timedelta(days=1)
    dash_calendar_html = build_calendar(cy, cm, dash_ev_color_map, admin_code=code)

    # 統合タイムライン（予定 + 集金締切 + 注文締切）
    wd_jp = ['月','火','水','木','金','土','日']
    def date_label(d):
        try:
            from datetime import date as _date
            dt = _date.fromisoformat(d)
            return f'{dt.month}/{dt.day}({wd_jp[dt.weekday()]})'
        except Exception:
            return d

    timeline = []
    for ev in events:
        no_answer = get_no_answer(ev['id'])
        sub = ''
        if no_answer:
            sub = f'{len(no_answer)}名未回答'
        timeline.append({'date': ev['event_date'], 'type': 'event', 'title': ev['title'],
                         'time': ev['event_time'] or '', 'sub': sub,
                         'url': f'/t/{code}/admin/events/{ev["id"]}', 'label': '詳細'})
    for f in upcoming_fees:
        timeline.append({'date': f['due_date'], 'type': 'fee', 'title': f'{f["title"]}　集金締切',
                         'time': '', 'sub': f'¥{f["amount"]:,}',
                         'url': f'/t/{code}/admin/fees', 'label': '管理'})
    for o in upcoming_orders:
        timeline.append({'date': o['deadline'], 'type': 'order', 'title': f'{o["title"]}　注文締切',
                         'time': '', 'sub': '',
                         'url': f'/t/{code}/orders', 'label': '詳細'})
    timeline.sort(key=lambda x: (x['date'], x['type']))

    type_color = {'event': '#111827', 'fee': '#d97706', 'order': '#6366f1'}
    type_badge = {'event': '', 'fee': '<span style="font-size:10px;background:#fffbeb;color:#d97706;border-radius:4px;padding:1px 5px;margin-left:4px">集金</span>',
                  'order': '<span style="font-size:10px;background:#eef2ff;color:#6366f1;border-radius:4px;padding:1px 5px;margin-left:4px">注文</span>'}

    timeline_rows = ''
    for item in timeline:
        sub_html = ''
        if item['time']:
            sub_html += f'<span style="color:#9ca3af">{item["time"]}</span>'
        if item['sub']:
            sub_html += f'{"　" if sub_html else ""}<span style="color:#dc2626;font-size:11px">{item["sub"]}</span>'
        timeline_rows += f'''
    <div style="display:flex;align-items:center;gap:10px;padding:9px 14px;border-bottom:1px solid #f3f4f6">
      <div style="width:6px;height:6px;border-radius:50%;background:{type_color[item["type"]]};flex-shrink:0"></div>
      <div style="min-width:52px;font-size:11px;color:#6b7280;font-family:var(--font-num);flex-shrink:0">{date_label(item["date"])}</div>
      <div style="flex:1;min-width:0">
        <div style="font-size:13px;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{item["title"]}{type_badge[item["type"]]}</div>
        {f'<div style="font-size:11px;margin-top:1px">{sub_html}</div>' if sub_html else ''}
      </div>
      <a href="{item["url"]}" style="font-size:12px;color:var(--rak-amber);flex-shrink:0">{item["label"]} →</a>
    </div>'''

    if not timeline:
        timeline_rows = '<div style="padding:20px;text-align:center;color:#9ca3af;font-size:13px">直近の予定・締切はありません</div>'

    conn.close()

    notice_rows = ''.join(f'''
    <div class="card-sm row" style="justify-content:space-between">
      <div>
        <div style="font-weight:700">{n['title']}</div>
        <div style="font-size:12px;color:#888">{fmt_datetime(n['created_at'])}</div>
      </div>
      <a href="/t/{code}/notices/{n['id']}" class="btn btn-sm btn-outline">確認</a>
    </div>''' for n in notices) or '<div class="empty">お知らせなし</div>'

    trial_days = get_trial_days_left(team)
    if trial_days is not None:
        plan_card = f'<div class="card" style="background:linear-gradient(135deg,#1c1a00,#3d2e00);color:#fff;border:none;text-align:center"><div style="font-size:12px;opacity:.8;margin-bottom:4px">現在のプラン</div><div style="font-size:20px;font-weight:900;margin-bottom:6px">Rak Pro ✦ トライアル中</div><div style="font-size:13px;color:#fbbf24;margin-bottom:10px">残り{trial_days}日</div><a href="/t/{code}/upgrade" style="font-size:12px;color:#fbbf24;text-decoration:underline">トライアル終了後に続けるには →</a></div>'
    elif is_pro(team):
        plan_card = '<div class="card" style="background:linear-gradient(135deg,#111,#333);color:#fff;border:none;text-align:center"><div style="font-size:12px;opacity:.8;margin-bottom:4px">現在のプラン</div><div style="font-size:20px;font-weight:900;margin-bottom:8px">Rak Pro ✦</div><div style="font-size:12px;opacity:.7">すべての機能をご利用中</div></div>'
    else:
        plan_card = f'<div class="card" style="border:2px solid #d97706;text-align:center;padding:20px"><div style="font-size:12px;color:#888;margin-bottom:4px">現在のプラン</div><div style="font-size:18px;font-weight:700;margin-bottom:12px">Free</div><a href="/t/{code}/upgrade" class="btn btn-blue" style="font-size:14px;padding:10px 24px">Proにアップグレード ¥980/月</a></div>'

    unpaid_badge = f'<span style="background:#dc2626;color:#fff;border-radius:10px;font-size:10px;padding:1px 6px;margin-left:4px">{len(unpaid_summary)}</span>' if unpaid_summary else ''
    unanswered_events_count = sum(1 for ev in events if get_no_answer(ev['id']))
    unanswered_badge = f'<span style="background:#dc2626;color:#fff;border-radius:10px;font-size:10px;padding:1px 6px;margin-left:4px">{unanswered_events_count}</span>' if unanswered_events_count else ''

    name_prompt = ''
    if not (team['name'] or '').strip():
        name_prompt = f'''
  <a href="/t/{code}/admin/settings" style="display:block;text-decoration:none;background:#fffbeb;border:1.5px solid #d97706;border-radius:12px;padding:14px 16px;margin-bottom:14px">
    <div style="font-size:13px;font-weight:700;color:#111">📝 チーム名を設定しましょう</div>
    <div style="font-size:12px;color:#92400e;margin-top:2px">メンバーに共有する画面に表示されます（タップして設定）</div>
  </a>'''

    body = f'''
<div class="container">
  {'<div class="msg-ok">' + _CHK + ' 作成しました！まずはチーム名を設定して、予定を追加しましょう。</div>' if created else ''}
  {name_prompt}
  {(lambda: (
    '<div style="background:#fff;border:1.5px solid #d97706;border-radius:12px;padding:16px 18px;margin-bottom:14px">'
    '<div style="font-size:13px;font-weight:700;margin-bottom:10px;color:#111">🚀 はじめの3ステップ</div>'
    '<div style="display:flex;flex-direction:column;gap:8px;font-size:13px">'
    + (f'<a href="/t/{code}/admin/members" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;background:#22c55e;color:#fff;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">✓</span><span style="color:#6b7280;text-decoration:line-through">メンバー名簿を登録する</span></a>' if member_names else f'<a href="/t/{code}/admin/members" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;border:2px solid #d97706;color:#d97706;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">1</span><span style="font-weight:600">メンバー名簿を登録する →</span></a>')
    + (f'<a href="/t/{code}/admin/events/new" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;background:#22c55e;color:#fff;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">✓</span><span style="color:#6b7280;text-decoration:line-through">最初の予定を追加する</span></a>' if has_any_event else f'<a href="/t/{code}/admin/events/new" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;border:2px solid #d97706;color:#d97706;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">2</span><span style="font-weight:600">最初の予定を追加する →</span></a>')
    + (f'<a href="/t/{code}/admin/notices/new" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;background:#22c55e;color:#fff;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">✓</span><span style="color:#6b7280;text-decoration:line-through">お知らせを送る</span></a>' if has_any_notice else f'<a href="/t/{code}/admin/notices/new" style="display:flex;align-items:center;gap:10px;color:inherit;text-decoration:none"><span style="width:22px;height:22px;border-radius:50%;border:2px solid #d97706;color:#d97706;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center">3</span><span style="font-weight:600">お知らせを送る →</span></a>')
    + '</div></div>'
  ) if (not member_names or not has_any_event or not has_any_notice) else '')()}


  <style>
    .admin-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:4px}}
    .atile{{background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;cursor:pointer}}
    .atile summary{{list-style:none;padding:14px 8px 12px;text-align:center;display:flex;flex-direction:column;align-items:center;gap:5px;font-size:11px;font-weight:500;color:#6b7280;user-select:none}}
    .atile summary::-webkit-details-marker{{display:none}}
    .atile[open]{{border-color:#d97706;grid-column:1/-1}}
    .atile[open] summary{{background:#f9fafb;border-bottom:1px solid #e5e7eb;color:#111827;flex-direction:row;justify-content:flex-start;gap:8px;padding:12px 14px;font-size:13px}}
    .atile[open] .atile-icon{{color:#d97706}}
    .atile-icon{{width:28px;height:28px;display:flex;align-items:center;justify-content:center;color:#9ca3af}}
    .atile-body{{padding:10px 14px;display:flex;gap:8px;align-items:center;font-size:13px}}
    .atile-body .btn{{flex:1;font-size:13px;padding:9px 10px;text-align:center}}
  </style>

  <div class="admin-grid">

    <details class="atile" open data-default-open>
      <summary><span class="atile-icon">{_ICO_CALENDAR}</span>予定・締切{unanswered_badge}</summary>
      <div class="atile-body" style="flex-direction:column;align-items:stretch;padding:0;gap:0">
        <div id="dash-list-view">
          {timeline_rows}
        </div>
        <div id="dash-cal-view" style="display:none;padding:12px 10px 4px">
          {dash_calendar_html}
        </div>
        <div style="display:flex;gap:8px;padding:10px 14px;border-top:1px solid #f3f4f6;flex-wrap:wrap">
          <a href="/t/{code}/admin/events/new" class="btn btn-blue" style="flex:1;font-size:13px;padding:9px 10px;min-width:80px">＋ 予定追加</a>
          {'<a href="/t/' + code + '/admin/ai-schedule" class="btn" style="flex:1;font-size:13px;padding:9px 10px;min-width:80px;background:#d97706;color:#fff">✦ AI作成</a>' if is_pro(team) else '<a href="/t/' + code + '/upgrade" class="btn" style="flex:1;font-size:13px;padding:9px 10px;min-width:80px;background:#fff;color:#d97706;border:1.5px solid #d97706">✦ AI作成</a>'}
          <button onclick="dashToggleCal(this)" class="btn btn-outline" id="dash-cal-btn" style="flex:1;font-size:13px;padding:9px 10px;min-width:80px">カレンダー</button>
        </div>
      </div>
    </details>
    <script>
    function dashToggleCal(btn){{
      var lv=document.getElementById('dash-list-view');
      var cv=document.getElementById('dash-cal-view');
      if(cv.style.display==='none'){{
        cv.style.display='block';lv.style.display='none';
        btn.textContent='リスト';btn.style.background='#f9fafb';
      }}else{{
        cv.style.display='none';lv.style.display='block';
        btn.textContent='カレンダー';btn.style.background='';
      }}
    }}
    </script>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_BELL_SM}</span>お知らせ</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/notices/new" class="btn btn-blue">＋ 作成</a>
        <a href="/t/{code}/notices" class="btn btn-outline">すべて見る</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_PEOPLE}</span>メンバー</summary>
      <div class="atile-body">
        <span style="font-size:12px;color:#888;white-space:nowrap">{member_count_label(team, len(member_names))}</span>
        <a href="/t/{code}/admin/members" class="btn btn-outline">一覧・追加</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_MONEY_SM}</span>集金{unpaid_badge}{''+_PRO_BADGE if not is_pro(team) else ''}</summary>
      <div class="atile-body">
        <span style="font-size:12px;color:#888;white-space:nowrap">未払い {len(unpaid_summary)}件</span>
        <a href="/t/{code}/admin/fees" class="btn btn-outline">管理</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_CLIPBOARD}</span>注文フォーム{''+_PRO_BADGE if not is_pro(team) else ''}</summary>
      <div class="atile-body">
        <a href="/t/{code}/orders" class="btn btn-outline">フォーム一覧</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_UNIFORM}</span>ユニフォーム{''+_PRO_BADGE if not is_pro(team) else ''}</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/uniforms" class="btn btn-outline">管理</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_CHART_SM}</span>AI文章{''+_PRO_BADGE if not is_pro(team) else ''}</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/ai" class="btn btn-outline">文章を作成</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_LEDGER}</span>会計{''+_PRO_BADGE if not is_pro(team) else ''}</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/ledger" class="btn btn-outline">収支記録</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_MEMO}</span>メモ</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/memos" class="btn btn-outline">メモを開く</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_MAIL}</span>問い合わせ</summary>
      <div class="atile-body">
        <a href="/feedback?from={code}" class="btn btn-outline">送る</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_HELP}</span>使い方</summary>
      <div class="atile-body">
        <a href="/t/{code}/help" class="btn btn-outline">ガイドを見る</a>
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{_ICO_CROWN}</span>プラン</summary>
      <div class="atile-body">
        {f'<span style="font-size:13px;font-weight:600;color:#d97706">Pro ✦ トライアル中（残{trial_days}日）</span><br><a href="/t/{code}/upgrade" style="font-size:12px;color:#888">続けるには課金が必要です →</a>' if trial_days is not None else ('<span style="font-size:13px;font-weight:600;color:#d97706">Pro ✦ 利用中</span>' if is_pro(team) else f'<a href="/t/{code}/upgrade" class="btn btn-blue">Proへアップグレード</a>')}
      </div>
    </details>

    <details class="atile">
      <summary><span class="atile-icon">{ICONS['admin']}</span>設定</summary>
      <div class="atile-body">
        <a href="/t/{code}/admin/settings" class="btn btn-outline">チーム名・パスワード</a>
      </div>
    </details>

  </div>

  <div style="text-align:right;margin-top:16px">
    <a href="/t/{code}/admin/logout" style="font-size:12px;color:#aaa">ログアウト</a>
  </div>
</div>
<script>
window.addEventListener('pageshow', function(e) {{
  if (e.persisted) {{
    document.querySelectorAll('.atile[open]').forEach(function(d) {{
      if (!d.dataset.defaultOpen) d.removeAttribute('open');
    }});
  }}
}});
</script>'''
    return page('管理ダッシュボード', body, code, active='home')


# ── Admin: memo ───────────────────────────────────────────────────

@app.route('/t/<code>/admin/memo', methods=['POST'])
def admin_memo_save(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    memo = request.form.get('memo', '')
    conn = get_db()
    conn.execute('UPDATE teams SET admin_memo=? WHERE id=?', (memo, team['id']))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_dash', code=code))


# ── Admin: memos ──────────────────────────────────────────────────

MEMO_FILE_DIR = os.path.join(os.path.dirname(os.path.abspath(os.environ.get('DATABASE', 'rak.db'))), 'memo_files')

@app.route('/t/<code>/admin/memos')
def admin_memos(code):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    memos = conn.execute('SELECT * FROM admin_memos WHERE team_id=? ORDER BY updated_at DESC', (team['id'],)).fetchall()
    conn.close()
    rows = ''.join(f'''
    <a href="/t/{code}/admin/memos/{m['id']}" style="display:block;text-decoration:none;color:inherit">
      <div class="card-sm" style="display:flex;justify-content:space-between;align-items:center">
        <div>
          <div style="font-weight:700;font-size:14px">{m['title']}</div>
          <div style="font-size:11px;color:#aaa;margin-top:2px">{m['updated_at']}</div>
        </div>
        <span style="color:#d97706;font-size:18px">›</span>
      </div>
    </a>''' for m in memos) or '<div class="empty card">メモはまだありません</div>'
    body = f'''
<div class="container">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
    <h1 style="margin:0">{_ICO_MEMO} メモ</h1>
    <a href="/t/{code}/admin/memos/new" class="btn btn-blue">＋ 新規作成</a>
  </div>
  {rows}
  <div style="margin-top:12px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('メモ', body, code, active='memo')

@app.route('/t/<code>/admin/memos/new', methods=['GET', 'POST'])
def admin_memo_new(code):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    error = ''
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        if not title:
            error = 'タイトルを入力してください'
        else:
            mid = new_id()
            now = now_str()
            conn = get_db()
            conn.execute('INSERT INTO admin_memos (id,team_id,title,content,created_at,updated_at) VALUES (?,?,?,?,?,?)',
                         (mid, team['id'], title, content, now, now))
            conn.commit()
            conn.close()
            return redirect(f'/t/{code}/admin/memos/{mid}')
    body = f'''
<div class="container">
  <h1>{_ICO_MEMO} 新規メモ</h1>
  {'<div class="msg-err">' + error + '</div>' if error else ''}
  <div class="card">
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" placeholder="例：5月練習日程メモ" required>
      <label>内容</label>
      <textarea name="content" rows="8" placeholder="メモの内容を入力..."></textarea>
      <button type="submit" class="btn btn-blue btn-block" style="margin-top:16px">保存</button>
    </form>
  </div>
  <div><a href="/t/{code}/admin/memos" style="font-size:13px;color:#888">← メモ一覧に戻る</a></div>
</div>'''
    return page('新規メモ', body, code, active='memo')

@app.route('/t/<code>/admin/memos/<memo_id>')
def admin_memo_detail(code, memo_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    memo = conn.execute('SELECT * FROM admin_memos WHERE id=? AND team_id=?', (memo_id, team['id'])).fetchone()
    if not memo: conn.close(); return redirect(f'/t/{code}/admin/memos')
    files = conn.execute('SELECT * FROM memo_files WHERE memo_id=? ORDER BY created_at DESC', (memo_id,)).fetchall()
    conn.close()
    content_html = memo['content'].replace('\n', '<br>') if memo['content'] else '<span style="color:#aaa">内容なし</span>'
    _img_exts = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic')
    file_rows = ''
    for f in files:
        is_img = os.path.splitext(f['stored_name'])[1].lower() in _img_exts
        preview = (f'<a href="/t/{code}/admin/memos/file/{f["id"]}?inline=1" target="_blank">'
                   f'<img src="/t/{code}/admin/memos/file/{f["id"]}?inline=1" '
                   f'style="width:100%;max-height:220px;object-fit:cover;border-radius:8px;margin-bottom:8px" loading="lazy"></a>') if is_img else ''
        file_rows += f'''
    <div class="card-sm">
      {preview}
      <div style="display:flex;justify-content:space-between;align-items:center">
        <div style="font-size:13px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1">{f['original_name']}</div>
        <div style="display:flex;gap:8px;flex-shrink:0;margin-left:8px">
          <a href="/t/{code}/admin/memos/file/{f['id']}" class="btn btn-sm btn-outline">DL</a>
          <form method="post" action="/t/{code}/admin/memos/file/{f['id']}/delete" style="margin:0">
            <button class="btn btn-sm" style="background:#fee2e2;color:#dc2626;border:none">削除</button>
          </form>
        </div>
      </div>
    </div>'''
    file_rows = file_rows or '<div style="font-size:13px;color:#aaa">添付ファイルなし</div>'
    body = f'''
<div class="container">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px">
    <h1 style="margin:0;flex:1">{memo['title']}</h1>
    <a href="/t/{code}/admin/memos/{memo_id}/edit" class="btn btn-sm btn-outline" style="margin-left:8px;flex-shrink:0">編集</a>
  </div>
  <div style="font-size:11px;color:#aaa;margin-bottom:16px">更新: {memo['updated_at']}</div>

  <div class="card">
    <div style="font-size:14px;line-height:1.8">{content_html}</div>
  </div>

  <div class="card">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <h2 style="margin:0">添付ファイル</h2>
    </div>
    {file_rows}
    <form method="POST" action="/t/{code}/admin/memos/{memo_id}/file" enctype="multipart/form-data" style="margin-top:14px">
      <input type="file" name="file" multiple style="font-size:13px;margin-bottom:8px">
      <div style="font-size:11px;color:#aaa;margin-bottom:8px">写真・PDF・Excelなど何でも添付できます（複数選択可）</div>
      <button type="submit" class="btn btn-outline btn-block">アップロード</button>
    </form>
  </div>

  <div style="display:flex;justify-content:space-between;align-items:center;margin-top:4px">
    <a href="/t/{code}/admin/memos" style="font-size:13px;color:#888">← メモ一覧</a>
    <form method="post" action="/t/{code}/admin/memos/{memo_id}/delete" style="margin:0">
      <button class="btn btn-sm" style="background:#fee2e2;color:#dc2626;border:none;font-size:12px">このメモを削除</button>
    </form>
  </div>
</div>'''
    return page(memo['title'], body, code, active='memo')

@app.route('/t/<code>/admin/memos/<memo_id>/edit', methods=['GET', 'POST'])
def admin_memo_edit(code, memo_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    memo = conn.execute('SELECT * FROM admin_memos WHERE id=? AND team_id=?', (memo_id, team['id'])).fetchone()
    if not memo: conn.close(); return redirect(f'/t/{code}/admin/memos')
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        if title:
            conn.execute('UPDATE admin_memos SET title=?,content=?,updated_at=? WHERE id=?',
                         (title, content, now_str(), memo_id))
            conn.commit()
        conn.close()
        return redirect(f'/t/{code}/admin/memos/{memo_id}')
    conn.close()
    body = f'''
<div class="container">
  <h1>{_ICO_MEMO} メモを編集</h1>
  <div class="card">
    <form method="POST">
      <label>タイトル</label>
      <input type="text" name="title" value="{memo['title']}" required>
      <label>内容</label>
      <textarea name="content" rows="10">{memo['content'] or ''}</textarea>
      <button type="submit" class="btn btn-blue btn-block" style="margin-top:16px">保存</button>
    </form>
  </div>
  <div><a href="/t/{code}/admin/memos/{memo_id}" style="font-size:13px;color:#888">← キャンセル</a></div>
</div>'''
    return page('メモ編集', body, code, active='memo')

@app.route('/t/<code>/admin/memos/<memo_id>/delete', methods=['POST'])
def admin_memo_delete(code, memo_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    files = conn.execute('SELECT stored_name FROM memo_files WHERE memo_id=?', (memo_id,)).fetchall()
    for f in files:
        try: os.remove(os.path.join(MEMO_FILE_DIR, f['stored_name']))
        except: pass
    conn.execute('DELETE FROM memo_files WHERE memo_id=?', (memo_id,))
    conn.execute('DELETE FROM admin_memos WHERE id=? AND team_id=?', (memo_id, team['id']))
    conn.commit()
    conn.close()
    return redirect(f'/t/{code}/admin/memos')

@app.route('/t/<code>/admin/memos/<memo_id>/file', methods=['POST'])
def admin_memo_file_upload(code, memo_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    memo = conn.execute('SELECT id FROM admin_memos WHERE id=? AND team_id=?', (memo_id, team['id'])).fetchone()
    if not memo: conn.close(); return redirect(f'/t/{code}/admin/memos')
    for f in request.files.getlist('file'):
        if not f or not f.filename:
            continue
        os.makedirs(MEMO_FILE_DIR, exist_ok=True)
        fid = new_id()
        ext = os.path.splitext(f.filename)[1]
        stored = f'{fid}{ext}'
        f.save(os.path.join(MEMO_FILE_DIR, stored))
        conn.execute('INSERT INTO memo_files (id,memo_id,original_name,stored_name,created_at) VALUES (?,?,?,?,?)',
                     (fid, memo_id, f.filename, stored, now_str()))
        conn.commit()
    conn.close()
    return redirect(f'/t/{code}/admin/memos/{memo_id}')

@app.route('/t/<code>/admin/memos/file/<file_id>')
def admin_memo_file_download(code, file_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    conn = get_db()
    f = conn.execute('SELECT * FROM memo_files WHERE id=?', (file_id,)).fetchone()
    conn.close()
    if not f: return 'Not found', 404
    inline = request.args.get('inline') == '1'
    return send_file(os.path.join(MEMO_FILE_DIR, f['stored_name']),
                     download_name=f['original_name'], as_attachment=not inline)

@app.route('/t/<code>/admin/memos/file/<file_id>/delete', methods=['POST'])
def admin_memo_file_delete(code, file_id):
    if not is_admin(code): return redirect(url_for('admin_login', code=code))
    conn = get_db()
    f = conn.execute('SELECT * FROM memo_files WHERE id=?', (file_id,)).fetchone()
    if f:
        try: os.remove(os.path.join(MEMO_FILE_DIR, f['stored_name']))
        except: pass
        memo_id = f['memo_id']
        conn.execute('DELETE FROM memo_files WHERE id=?', (file_id,))
        conn.commit()
        conn.close()
        return redirect(f'/t/{code}/admin/memos/{memo_id}')
    conn.close()
    return redirect(f'/t/{code}/admin/memos')


# ── Help ──────────────────────────────────────────────────────────

@app.route('/t/<code>/help')
def team_help(code):
    team = get_team(code)
    if not team: return redirect('/')
    admin = is_admin(code)

    steps_member = [
        ('届いたリンクを開く', '連絡アプリ（LINE等）に管理者から届いた出欠リンクをタップするだけ。アプリのインストールも登録も不要です。'),
        ('自分の名前を選ぶ', 'リンクを開くと名簿が表示されます。自分の名前をタップして選びます。次回からは自動で記憶されます。'),
        ('出欠をタップで回答する', '予定ごとに「出席・欠席」をタップするだけ。回答はそのまま管理者に届き、自動で集計されます。'),
        ('予定表リンクで日程を確認', '管理者から予定表リンクが届いたら、開くだけでチームの練習・試合スケジュールを確認できます。'),
    ]

    steps_admin = [
        ('メンバー名簿を登録する', '「メンバー」タイルから名前を追加します。出欠や集金の集計で「誰が答えたか」を表示するために使います。'),
        ('予定を追加する', '「予定」タイルから練習や試合を登録。出欠の取り方（出席・欠席／欠席のみ／取らない）も選べます。'),
        ('共有リンクをメンバーに送る', '「予定」画面の共有リンクをコピーして、ふだんの連絡アプリに貼るだけ。メンバーは登録不要でリンクを開き、出欠をタップで回答できます。'),
        ('お知らせを作成する', '「お知らせ」タイルから連絡文を作成。AI文章作成で下書きを自動生成し、コピーして連絡アプリに貼れます。'),
        ('集金・会計を管理する', '「集金」で誰が払ったかを管理、「会計」でチームの収支を記録できます。'),
        ('メモ・ファイルを管理する', '「メモ」タイルで管理者専用のメモを作成。PDFや写真などのファイルも添付できます。'),
    ]

    def step_html(steps, color):
        html = ''
        for i, (title, desc) in enumerate(steps, 1):
            html += f'''
      <div style="display:flex;gap:14px;margin-bottom:20px;align-items:flex-start">
        <div style="width:28px;height:28px;border-radius:50%;background:{color};color:#fff;font-weight:900;font-size:13px;display:flex;align-items:center;justify-content:center;flex-shrink:0">{i}</div>
        <div>
          <div style="font-weight:700;font-size:14px;margin-bottom:4px">{title}</div>
          <div style="font-size:13px;color:#666;line-height:1.6">{desc}</div>
        </div>
      </div>'''
        return html

    body = f'''
<div class="container">
  <h1 style="margin-bottom:4px">使い方ガイド</h1>
  <p style="font-size:13px;color:#888;margin-bottom:20px">Rakの基本的な使い方をまとめました</p>

  <div class="card">
    <h2 style="margin-bottom:16px">{_ICO_PEOPLE} メンバーの方へ</h2>
    {step_html(steps_member, '#d97706')}
  </div>

  {'<div class="card"><h2 style="margin-bottom:16px">⚙ 管理者の方へ</h2>' + step_html(steps_admin, '#111') + '</div>' if admin else ''}

  <div class="card" style="background:#fef3c7;border-color:#fde68a">
    <div style="font-weight:700;margin-bottom:6px">困ったときは</div>
    <div style="font-size:13px;color:#666;line-height:1.7">
      ご不明な点・機能のご要望は<a href="/feedback" style="color:#d97706">お問い合わせ</a>からお気軽にご連絡ください。
    </div>
  </div>

  <div style="margin-top:4px"><a href="/t/{code}/{'admin/dash' if admin else 'schedule'}" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('使い方ガイド', body, code, active='help' if admin else None)


# ── Admin: events ─────────────────────────────────────────────────

@app.route('/t/<code>/admin/events/new', methods=['GET', 'POST'])
def admin_new_event(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    error = ''
    prefill_date = request.args.get('date', '')

    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        date        = request.form.get('event_date', '').strip()
        end_date    = request.form.get('end_date', '').strip()
        time        = request.form.get('event_time', '').strip()
        end_time    = request.form.get('end_time', '').strip()
        location    = request.form.get('location', '').strip()
        note        = request.form.get('note', '').strip()
        event_color = request.form.get('event_color', '').strip()
        rsvp_mode   = request.form.get('rsvp_mode', 'both').strip()
        if rsvp_mode not in ('both', 'absent_only', 'none'):
            rsvp_mode = 'both'
        if end_date and end_date < date:
            end_date = date
        if not title or not date:
            error = 'タイトルと日付を入力してください'
        else:
            conn = get_db()
            eid = new_id()
            conn.execute(
                'INSERT INTO events (id,team_id,title,event_date,event_time,location,note,created_at,end_date,end_time,event_color,rsvp_mode) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                (eid, team['id'], title, date, time, location, note, now_str(), end_date, end_time, event_color, rsvp_mode)
            )
            conn.commit()
            conn.close()
            return redirect(url_for('schedule', code=code))

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>予定を追加</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" placeholder="例：練習試合 vs FCさくら" required>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>開始日 *</label>
          <input type="date" name="event_date" id="ev-date" value="{prefill_date}" required oninput="rakDateJa(this,'ev-date-ja')">
          <div id="ev-date-ja" style="font-size:11px;color:#6b7280;margin-top:3px;min-height:16px"></div>
        </div>
        <div>
          <label>終了日（複数日）</label>
          <input type="date" name="end_date" id="ev-end-date" oninput="rakDateJa(this,'ev-end-date-ja')">
          <div id="ev-end-date-ja" style="font-size:11px;color:#6b7280;margin-top:3px;min-height:16px"></div>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>開始時刻</label>
          <input type="time" name="event_time">
        </div>
        <div>
          <label>終了時刻</label>
          <input type="time" name="end_time">
        </div>
      </div>
      <label>場所</label>
      <input type="text" name="location" placeholder="例：市営グラウンドA面">
      <label>備考・詳細</label>
      <textarea name="note" placeholder="持ち物・集合場所など"></textarea>
      <label>出欠の取り方</label>
      <select name="rsvp_mode">
        <option value="both">出席・欠席を聞く</option>
        <option value="absent_only">欠席のみ連絡（基本は出席扱い）</option>
        <option value="none">出欠は取らない（連絡だけ）</option>
      </select>
      <div style="font-size:11px;color:#888;margin-top:4px">回答リンクでメンバーに表示するボタンが変わります</div>
      <label>カラー（カレンダー表示色）</label>
      {color_picker_html()}
      <button class="btn btn-blue btn-block" type="submit">追加する</button>
    </form>
  </div>
  <div style="display:flex;justify-content:space-between">
    <a href="/t/{code}/schedule" style="font-size:13px;color:#888">← 予定一覧</a>
    <a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>
  </div>
</div>
<script>
function rakDateJa(inp,tid){{
  var v=inp.value;
  if(!v){{document.getElementById(tid).textContent='';return;}}
  var p=v.split('-'),d=new Date(v+'T00:00:00');
  var wd=['日','月','火','水','木','金','土'][d.getDay()];
  document.getElementById(tid).textContent=p[0]+'年'+parseInt(p[1])+'月'+parseInt(p[2])+'日（'+wd+'）';
}}
</script>'''
    return page('予定を追加', body, code, active='schedule')

@app.route('/t/<code>/admin/events/<event_id>')
def admin_event_detail(code, event_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    ev = conn.execute('SELECT * FROM events WHERE id=? AND team_id=?', (event_id, team['id'])).fetchone()
    if not ev:
        conn.close()
        return redirect(url_for('schedule', code=code))
    rsvps = conn.execute('SELECT * FROM rsvps WHERE event_id=? ORDER BY status,member_name', (event_id,)).fetchall()
    members_all = conn.execute('SELECT name FROM members WHERE team_id=? ORDER BY name', (team['id'],)).fetchall()
    conn.close()

    attending = [r for r in rsvps if r['status'] == 'attending']
    absent = [r for r in rsvps if r['status'] == 'absent']
    answered_names = set(r['member_name'] for r in rsvps)
    no_answer = [m['name'] for m in members_all if m['name'] not in answered_names]

    def names(lst):
        return ''.join(f'<div style="font-size:14px;padding:5px 0;border-bottom:1px solid #f0f0f0">{r["member_name"]}</div>' for r in lst) or '<div style="font-size:13px;color:#aaa">なし</div>'

    no_answer_card = ''
    if no_answer:
        no_answer_card = f'''
  <div class="card" style="border-color:#fca5a5">
    <div style="margin-bottom:10px"><span class="badge badge-red">未回答 {len(no_answer)}名</span></div>
    {''.join(f'<div style="font-size:14px;padding:5px 0;border-bottom:1px solid #f0f0f0">{n}</div>' for n in no_answer)}
  </div>'''

    body = f'''
<div class="container">
  <div class="card">
    <div class="row" style="justify-content:space-between;align-items:flex-start;margin-bottom:10px">
      <h1 style="margin:0">{ev['title']}</h1>
      <div style="display:flex;gap:8px;flex-shrink:0">
        <a href="/t/{code}/admin/events/{event_id}/edit" class="btn btn-sm btn-outline">編集</a>
        <form method="POST" action="/t/{code}/admin/events/{event_id}/delete" onsubmit="return confirm('この予定を削除しますか？')">
          <button class="btn btn-sm btn-gray" type="submit" style="color:#dc2626">削除</button>
        </form>
      </div>
    </div>
    <div style="font-size:14px;color:#555">
      {fmt_date_range(ev['event_date'], ev['end_date'])}{' ' + ev['event_time'] if ev['event_time'] else ''}
      {('　' + ev['location']) if ev['location'] else ''}
    </div>
    {f'<div style="margin-top:12px;background:#f8faff;padding:12px;border-radius:10px;font-size:14px">{ev["note"]}</div>' if ev['note'] else ''}
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
    <div class="card">
      <div class="row" style="margin-bottom:10px">
        <span class="badge badge-green">出席 {len(attending)}</span>
      </div>
      {names(attending)}
    </div>
    <div class="card">
      <div class="row" style="margin-bottom:10px">
        <span class="badge badge-red">欠席 {len(absent)}</span>
      </div>
      {names(absent)}
    </div>
  </div>
  {no_answer_card}
  <div style="margin-top:4px">
  </div>
  <div style="text-align:center;margin-top:12px"><a href="/t/{code}/schedule" style="font-size:13px;color:#888">← スケジュール</a></div>
</div>'''
    return page(ev['title'], body, code, active='schedule')


@app.route('/t/<code>/admin/events/<event_id>/edit', methods=['GET', 'POST'])
def admin_edit_event(code, event_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    ev = conn.execute('SELECT * FROM events WHERE id=? AND team_id=?', (event_id, team['id'])).fetchone()
    if not ev:
        conn.close()
        return redirect(url_for('schedule', code=code))

    ev_end_time = ''
    try:
        ev_end_time = ev['end_time'] or ''
    except Exception:
        pass

    error = ''
    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        date        = request.form.get('event_date', '').strip()
        end_date    = request.form.get('end_date', '').strip()
        time        = request.form.get('event_time', '').strip()
        end_time    = request.form.get('end_time', '').strip()
        location    = request.form.get('location', '').strip()
        note        = request.form.get('note', '').strip()
        event_color = request.form.get('event_color', '').strip()
        rsvp_mode   = request.form.get('rsvp_mode', 'both').strip()
        if rsvp_mode not in ('both', 'absent_only', 'none'):
            rsvp_mode = 'both'
        if end_date and end_date < date:
            end_date = date
        if not title or not date:
            error = 'タイトルと日付を入力してください'
        else:
            conn.execute(
                'UPDATE events SET title=?,event_date=?,end_date=?,event_time=?,end_time=?,location=?,note=?,event_color=?,rsvp_mode=? WHERE id=?',
                (title, date, end_date, time, end_time, location, note, event_color, rsvp_mode, event_id)
            )
            conn.commit()
            conn.close()
            return redirect(url_for('admin_event_detail', code=code, event_id=event_id))

    conn.close()

    ev_date_ja = ''
    try:
        import datetime as _dt
        d = _dt.date.fromisoformat(ev['event_date'])
        wd = ['月','火','水','木','金','土','日'][d.weekday()]
        ev_date_ja = f"{d.year}年{d.month}月{d.day}日（{wd}）"
    except Exception:
        pass

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>予定を編集</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" value="{ev['title']}" required>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>開始日 *</label>
          <input type="date" name="event_date" id="ev-date" value="{ev['event_date']}" required oninput="rakDateJa(this,'ev-date-ja')">
          <div id="ev-date-ja" style="font-size:11px;color:#6b7280;margin-top:3px">{ev_date_ja}</div>
        </div>
        <div>
          <label>終了日（複数日）</label>
          <input type="date" name="end_date" id="ev-end-date" value="{ev['end_date'] or ''}" oninput="rakDateJa(this,'ev-end-date-ja')">
          <div id="ev-end-date-ja" style="font-size:11px;color:#6b7280;margin-top:3px"></div>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>開始時刻</label>
          <input type="time" name="event_time" value="{ev['event_time'] or ''}">
        </div>
        <div>
          <label>終了時刻</label>
          <input type="time" name="end_time" value="{ev_end_time}">
        </div>
      </div>
      <label>場所</label>
      <input type="text" name="location" value="{ev['location'] or ''}">
      <label>備考・詳細</label>
      <textarea name="note">{ev['note'] or ''}</textarea>
      <label>出欠の取り方</label>
      <select name="rsvp_mode">
        <option value="both" {'selected' if (ev['rsvp_mode'] or 'both')=='both' else ''}>出席・欠席を聞く</option>
        <option value="absent_only" {'selected' if ev['rsvp_mode']=='absent_only' else ''}>欠席のみ連絡（基本は出席扱い）</option>
        <option value="none" {'selected' if ev['rsvp_mode']=='none' else ''}>出欠は取らない（連絡だけ）</option>
      </select>
      <div style="font-size:11px;color:#888;margin-top:4px">回答リンクでメンバーに表示するボタンが変わります</div>
      <label>カラー（カレンダー表示色）</label>
      {color_picker_html(ev['event_color'] or '')}
      <button class="btn btn-blue btn-block" type="submit">保存する</button>
    </form>
  </div>
  <div style="display:flex;justify-content:space-between">
    <a href="/t/{code}/admin/events/{event_id}" style="font-size:13px;color:#888">← 予定詳細</a>
    <a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>
  </div>
</div>
<script>
function rakDateJa(inp,tid){{
  var v=inp.value;
  if(!v){{document.getElementById(tid).textContent='';return;}}
  var p=v.split('-'),d=new Date(v+'T00:00:00');
  var wd=['日','月','火','水','木','金','土'][d.getDay()];
  document.getElementById(tid).textContent=p[0]+'年'+parseInt(p[1])+'月'+parseInt(p[2])+'日（'+wd+'）';
}}
</script>'''
    return page('予定を編集', body, code, active='schedule')


@app.route('/t/<code>/admin/events/<event_id>/delete', methods=['POST'])
def admin_delete_event(code, event_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    conn.execute('DELETE FROM events WHERE id=? AND team_id=?', (event_id, team['id']))
    conn.execute('DELETE FROM rsvps WHERE event_id=?', (event_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('schedule', code=code))


# ── Admin: notices ────────────────────────────────────────────────

@app.route('/t/<code>/admin/notices/new', methods=['GET', 'POST'])
def admin_new_notice(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    error = ''
    prefill_title = request.args.get('title', '')
    prefill_body = request.args.get('body', '')

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        body_text = request.form.get('body', '').strip()
        if not title or not body_text:
            error = 'タイトルと本文を入力してください'
        else:
            conn = get_db()
            conn.execute('INSERT INTO notices VALUES (?,?,?,?,?)',
                         (new_id(), team['id'], title, body_text, now_str()))
            conn.commit()
            conn.close()
            send_push_to_team(team['id'], f'📢 {team["name"]}', title, f'/t/{code}/notices')
            return redirect(url_for('notices', code=code, sent='1'))

    conn = get_db()
    team2 = get_team(code)
    templates = conn.execute('SELECT * FROM ai_templates WHERE team_id=? ORDER BY created_at DESC', (team2['id'],)).fetchall() if team2 else []
    conn.close()

    tmpl_opts = ''.join(f'<option value="{t["id"]}" data-title="{t["title"]}" data-body="{t["body"]}">{t["title"]}</option>' for t in templates)
    tmpl_select = f'''
    <div style="margin-bottom:12px">
      <select id="tmpl-select" class="btn btn-sm btn-outline" style="width:auto;border:2px solid #dde6ff;padding:7px 12px" onchange="applyTemplate(this)">
        <option value="">📌 テンプレートから選ぶ</option>
        {tmpl_opts}
      </select>
    </div>
    <script>
    function applyTemplate(sel) {{
      var opt = sel.options[sel.selectedIndex];
      if(opt.value) {{
        document.querySelector('[name=title]').value = opt.dataset.title;
        document.querySelector('[name=body]').value = opt.dataset.body;
      }}
    }}
    </script>''' if templates else ''

    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <h1>お知らせを作成</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <div style="margin-bottom:16px;display:flex;gap:8px;flex-wrap:wrap">
      <a href="/t/{code}/admin/ai?redirect=notice" class="btn btn-sm btn-outline">✦ AIで下書きを作る</a>
    </div>
    {tmpl_select}
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" placeholder="例：明日の練習について" required value="{prefill_title}">
      <label>本文 *</label>
      <textarea name="body" rows="8" placeholder="メンバーへのお知らせ内容を入力してください" required>{prefill_body}</textarea>
      <button class="btn btn-blue btn-block" type="submit">送信する</button>
    </form>
  </div>
  <div style="display:flex;justify-content:space-between">
    <a href="/t/{code}/notices" style="font-size:13px;color:#888">← お知らせ一覧</a>
    <a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>
  </div>
</div>'''
    return page('お知らせ作成', body, code, active='notices')


# ── Admin: AI ─────────────────────────────────────────────────────

@app.route('/t/<code>/admin/ai', methods=['GET', 'POST'])
def admin_ai(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='ai')

    redirect_to = request.args.get('redirect', '')
    result_title = ''
    result_body = ''
    error = ''
    memo = ''

    doc_type = ''
    if request.method == 'POST':
        memo = request.form.get('memo', '').strip()
        tone = request.form.get('tone', 'formal')
        doc_type = request.form.get('doc_type', 'notice')
        if not memo:
            error = 'メモを入力してください'
        elif not ANTHROPIC_API_KEY:
            error = 'ANTHROPIC_API_KEYが設定されていません'
        elif HAS_ANTHROPIC:
            try:
                client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
                tone_desc = '丁寧でやわらかい' if tone == 'formal' else 'シンプルで簡潔な'

                if doc_type == 'monthly':
                    prompt = f'''あなたはスポーツチームの運営をサポートするAIです。
コーチが書いたメモをもとに、保護者・関係者向けの月次活動報告書を作成してください。

メモ：{memo}

以下のJSON形式で返してください：
{{"title": "◯月度 活動報告（25字以内）", "body": "月次報告書本文（500〜700字、以下の構成で改行して記載：①今月の主な活動・試合結果 ②選手の成長・良かった点 ③課題と来月の目標 ④保護者へのお礼と協力依頼）"}}

JSONのみ返してください。説明不要です。'''
                    max_tokens = 1400
                elif doc_type == 'report':
                    prompt = f'''あなたはスポーツチームの運営をサポートするAIです。
コーチが書いたメモをもとに、保護者・関係者向けの活動報告書を作成してください。

メモ：{memo}

以下のJSON形式で返してください：
{{"title": "報告書タイトル（25字以内）", "body": "報告書本文（400〜600字、段落ごとに改行。活動の概要・成果・感謝・今後の活動について記載）"}}

JSONのみ返してください。説明不要です。'''
                    max_tokens = 1200
                elif doc_type == 'match':
                    prompt = f'''あなたはスポーツチームの運営をサポートするAIです。
コーチが書いたメモをもとに、試合結果レポートを作成してください。

メモ：{memo}

以下のJSON形式で返してください：
{{"title": "試合レポートタイトル（25字以内）", "body": "本文（300〜500字、改行あり。試合結果・内容の振り返り・選手の頑張り・今後に向けた抱負を含む）"}}

JSONのみ返してください。説明不要です。'''
                    max_tokens = 1000
                elif doc_type == 'practice':
                    prompt = f'''あなたはスポーツチームの運営をサポートするAIです。
コーチが書いたメモをもとに、練習報告文を作成してください。

メモ：{memo}

以下のJSON形式で返してください：
{{"title": "練習報告タイトル（20字以内）", "body": "本文（200〜350字、改行あり。練習内容・取り組み・成果・次回への課題を含む）"}}

JSONのみ返してください。説明不要です。'''
                    max_tokens = 800
                else:
                    prompt = f'''あなたはスポーツチームの運営をサポートするAIです。
コーチが書いた短いメモをもとに、保護者・メンバー向けの{tone_desc}連絡文を作成してください。

メモ：{memo}

以下のJSON形式で返してください：
{{"title": "お知らせのタイトル（20字以内）", "body": "本文（200字程度、改行あり）"}}

JSONのみ返してください。説明不要です。'''
                    max_tokens = 800

                message = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=max_tokens,
                    messages=[{'role': 'user', 'content': prompt}]
                )
                import json
                text = message.content[0].text.strip()
                if text.startswith('```'):
                    text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
                data = json.loads(text)
                result_title = data.get('title', '')
                result_body = data.get('body', '')
            except Exception as e:
                error = f'AI生成に失敗しました: {str(e)}'
        else:
            error = 'anthropicライブラリがインストールされていません'

    # テンプレート保存
    saved_msg = ''
    if request.method == 'POST' and request.form.get('action') == 'save_template':
        t_title = request.form.get('t_title', '').strip()
        t_body = request.form.get('t_body', '').strip()
        if t_title and t_body:
            team = get_team(code)
            conn = get_db()
            conn.execute('INSERT INTO ai_templates VALUES (?,?,?,?,?)',
                         (new_id(), team['id'], t_title, t_body, now_str()))
            conn.commit()
            conn.close()
            saved_msg = 'テンプレートに保存しました'
            result_title = t_title
            result_body = t_body

    result_block = ''
    if result_title and result_body:
        import html as _html
        esc_title = _html.escape(result_title, quote=True)
        esc_body = _html.escape(result_body)
        result_block = f'''
  <div class="card" style="border-color:#d97706">
    <div class="section-label">生成結果</div>
    <input id="ai-res-title" value="{esc_title}" style="width:100%;font-weight:700;font-size:16px;margin-bottom:8px">
    <textarea id="ai-res-body" rows="14" style="width:100%;font-size:14px;color:#333;line-height:1.8;background:#f8faff;padding:14px;border-radius:10px;border:1px solid #e5e7eb;box-sizing:border-box">{esc_body}</textarea>
    <div style="font-size:12px;color:#888;margin-top:4px">✎ タイトル・本文はこのまま書き換えできます</div>
    <button type="button" id="ai-copy-btn" class="btn btn-outline btn-sm" style="width:100%;margin-top:10px" onclick="rakAiCopy()">📋 本文をコピー</button>
    <a id="ai-use-btn" href="#" class="btn btn-blue" style="display:block;text-align:center;margin-top:8px">このままお知らせとして送信 →</a>
    <form method="POST" style="margin-top:8px" onsubmit="rakAiSyncForm()">
      <input type="hidden" name="action" value="save_template">
      <input type="hidden" name="t_title" id="ai-save-title" value="">
      <input type="hidden" name="t_body" id="ai-save-body" value="">
      <button class="btn btn-outline btn-sm" type="submit" style="width:100%">📌 テンプレートとして保存</button>
    </form>
  </div>
  <script>
  function rakAiTitle(){{return document.getElementById('ai-res-title').value;}}
  function rakAiBody(){{return document.getElementById('ai-res-body').value;}}
  function rakAiSyncUse(){{
    document.getElementById('ai-use-btn').href='/t/{code}/admin/notices/new?title='+encodeURIComponent(rakAiTitle())+'&body='+encodeURIComponent(rakAiBody());
  }}
  function rakAiSyncForm(){{
    document.getElementById('ai-save-title').value=rakAiTitle();
    document.getElementById('ai-save-body').value=rakAiBody();
  }}
  function rakAiCopy(){{
    const ta=document.getElementById('ai-res-body');
    ta.focus();ta.select();
    try{{document.execCommand('copy');}}catch(e){{}}
    if(navigator.clipboard&&navigator.clipboard.writeText){{navigator.clipboard.writeText(ta.value).catch(()=>{{}});}}
    ta.setSelectionRange(0,0);ta.blur();
    const b=document.getElementById('ai-copy-btn');
    b.textContent='✓ コピーしました';
    setTimeout(()=>{{b.textContent='📋 本文をコピー';}},1500);
  }}
  document.getElementById('ai-res-title').addEventListener('input',rakAiSyncUse);
  document.getElementById('ai-res-body').addEventListener('input',rakAiSyncUse);
  rakAiSyncUse();
  </script>'''

    # 保存済みテンプレート一覧
    team = get_team(code)
    conn = get_db()
    templates = conn.execute('SELECT * FROM ai_templates WHERE team_id=? ORDER BY created_at DESC', (team['id'],)).fetchall()
    conn.close()

    tmpl_rows = ''
    for t in templates:
        import urllib.parse
        params = urllib.parse.urlencode({'title': t['title'], 'body': t['body']})
        tmpl_rows += f'''
        <div class="card-sm row" style="justify-content:space-between;align-items:center">
          <div style="flex:1;min-width:0">
            <div style="font-weight:700;font-size:14px">{t['title']}</div>
            <div style="font-size:12px;color:#888;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{t['body'][:40]}…</div>
          </div>
          <div style="display:flex;gap:6px;margin-left:12px">
            <a href="/t/{code}/admin/notices/new?{params}" class="btn btn-sm btn-outline">使う</a>
            <a href="/t/{code}/admin/ai/template/{t['id']}/delete" class="btn btn-sm btn-gray">削除</a>
          </div>
        </div>'''

    tmpl_section = f'''
    <div class="card" style="margin-top:16px">
      <h2 style="margin-bottom:12px">📌 保存済みテンプレート</h2>
      {tmpl_rows if templates else '<div class="empty" style="padding:20px">保存されたテンプレートはありません</div>'}
    </div>''' if templates else ''

    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div class="section-label">✦ AI文章作成</div>
    <h1>AIで下書きを作る</h1>
    <p style="color:#666;font-size:13px;margin-bottom:16px">一言メモを入力するだけで、さまざまな文書を自動生成します</p>
    {f'<div class="msg-ok">{saved_msg}</div>' if saved_msg else ''}
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>文書タイプ</label>
      <select name="doc_type">
        <option value="notice" {'selected' if doc_type in ('','notice') else ''}>お知らせ・連絡文</option>
        <option value="monthly" {'selected' if doc_type=='monthly' else ''}>月次活動報告書</option>
        <option value="report" {'selected' if doc_type=='report' else ''}>活動報告書（単発）</option>
        <option value="match" {'selected' if doc_type=='match' else ''}>試合レポート</option>
        <option value="practice" {'selected' if doc_type=='practice' else ''}>練習報告</option>
      </select>
      <label>メモ・キーワード</label>
      <textarea name="memo" placeholder="例（活動報告書）：県大会 準優勝&#10;例（お知らせ）：明日の練習、雨で中止" rows="4">{memo}</textarea>
      <label>文体</label>
      <select name="tone">
        <option value="formal">丁寧・やわらか（保護者向け）</option>
        <option value="simple">シンプル・簡潔（メンバー向け）</option>
      </select>
      <button class="btn btn-blue btn-block" type="submit">✦ AI生成する</button>
    </form>
  </div>

  {result_block}

  {tmpl_section}

  <div style="margin-top:8px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('AI文章作成', body, code, active='ai')


@app.route('/t/<code>/admin/ai/template/<tmpl_id>/delete')
def admin_delete_template(code, tmpl_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    conn.execute('DELETE FROM ai_templates WHERE id=? AND team_id=?', (tmpl_id, team['id']))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_ai', code=code))


# ── Admin: members ───────────────────────────────────────────────

@app.route('/t/<code>/admin/members', methods=['GET', 'POST'])
def admin_members(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    msg = ''

    if request.method == 'POST':
        action = request.form.get('action')
        conn = get_db()
        if action == 'add':
            last_name = request.form.get('last_name', '').strip()
            first_name = request.form.get('first_name', '').strip()
            name = f'{last_name} {first_name}'.strip()
            number = request.form.get('number', '').strip()
            position = request.form.get('position', '').strip()
            if name:
                if can_add_team_member(team, conn):
                    conn.execute('INSERT INTO members VALUES (?,?,?,?,?,?)',
                                 (new_id(), team['id'], name, number, position, now_str()))
                    conn.commit()
                    msg = f'「{name}」を追加しました'
                else:
                    msg = 'メンバーの追加に失敗しました。'
        elif action == 'delete':
            mid = request.form.get('member_id')
            conn.execute('DELETE FROM members WHERE id=? AND team_id=?', (mid, team['id']))
            conn.commit()
            msg = 'メンバーを削除しました'
        conn.close()

    conn = get_db()
    members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY joined_at', (team['id'],)).fetchall()
    conn.close()

    rows = ''
    for m in members:
        rows += f'''
        <div class="card-sm">
          <div class="row" style="justify-content:space-between;align-items:center">
            <div>
              <span style="font-weight:700">{m['name']}</span>
              {f'<span style="font-size:12px;color:#888;margin-left:8px">#{m["number"]}</span>' if m['number'] else ''}
              {f'<span style="font-size:12px;color:#888;margin-left:6px">{m["position"]}</span>' if m['position'] else ''}
            </div>
            <div style="display:flex;gap:6px">
              <a href="/t/{code}/admin/members/{m['id']}/edit" class="btn btn-sm btn-outline">編集</a>
              <form method="POST" onsubmit="return confirm('{m['name']}を削除しますか？')" style="margin:0">
                <input type="hidden" name="action" value="delete">
                <input type="hidden" name="member_id" value="{m['id']}">
                <button class="btn btn-sm btn-gray" type="submit">削除</button>
              </form>
            </div>
          </div>
        </div>'''

    member_n = len(members)
    msg_cls = 'msg-ok' if msg and '追加しました' in msg else 'msg-err'
    add_section = f'''
  <div class="card" id="add-member">
    <h2>メンバーを追加</h2>
    <form method="POST">
      <input type="hidden" name="action" value="add">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>苗字 *</label>
          <input type="text" name="last_name" placeholder="田中" required>
        </div>
        <div>
          <label>名前 *</label>
          <input type="text" name="first_name" placeholder="花子" required>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>背番号</label>
          <input type="text" name="number" placeholder="例：10">
        </div>
        <div>
          <label>ポジション</label>
          <input type="text" name="position" placeholder="例：FW">
        </div>
      </div>
      <button class="btn btn-blue btn-block" type="submit">追加する</button>
    </form>
  </div>'''

    empty_invite = f'''
    <div style="text-align:center;padding:24px 16px">
      <div style="font-size:32px;margin-bottom:10px">👥</div>
      <div style="font-weight:700;font-size:15px;margin-bottom:6px">まだメンバーがいません</div>
      <div style="font-size:13px;color:#888;margin-bottom:20px;line-height:1.7">名簿を登録しておくと、出欠リンクでメンバーが<br>自分の名前を選んで回答できます。<br>メンバー側の登録・アプリは不要です。</div>
      <a href="#add-member" class="btn btn-blue" style="display:inline-block">＋ メンバーを追加する</a>
    </div>'''

    body = f'''
<div class="container" style="max-width:540px">
  {f'<div class="{msg_cls}">{msg}</div>' if msg else ''}
  <div class="card">
    <div class="row" style="margin-bottom:16px">
      <h1 style="margin:0">{_ICO_PEOPLE} メンバー名簿</h1>
      <div style="margin-left:auto;display:flex;align-items:center;gap:8px">
        <span class="badge badge-blue">{member_count_label(team, member_n)}</span>
        <a href="#add-member" class="btn btn-blue btn-sm">＋ 追加</a>
      </div>
    </div>
    {rows if members else empty_invite}
  </div>
  {add_section}
  <div style="margin-top:8px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('メンバー管理', body, code, active='members')


@app.route('/t/<code>/admin/members/<member_id>/edit', methods=['GET', 'POST'])
def admin_edit_member(code, member_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    m = conn.execute('SELECT * FROM members WHERE id=? AND team_id=?', (member_id, team['id'])).fetchone()
    if not m:
        conn.close()
        return redirect(url_for('admin_members', code=code))

    msg = ''
    if request.method == 'POST':
        last_name = request.form.get('last_name', '').strip()
        first_name = request.form.get('first_name', '').strip()
        name = f'{last_name} {first_name}'.strip()
        number = request.form.get('number', '').strip()
        position = request.form.get('position', '').strip()
        if name:
            conn.execute('UPDATE members SET name=?, number=?, position=? WHERE id=?',
                         (name, number, position, member_id))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_members', code=code))

    conn.close()
    _parts = m['name'].split(' ', 1)
    _existing_last = _parts[0]
    _existing_first = _parts[1] if len(_parts) > 1 else ''
    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>メンバーを編集</h1>
    <form method="POST">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>苗字 *</label>
          <input type="text" name="last_name" value="{_existing_last}" required>
        </div>
        <div>
          <label>名前 *</label>
          <input type="text" name="first_name" value="{_existing_first}" required>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label>背番号</label>
          <input type="text" name="number" value="{m['number']}">
        </div>
        <div>
          <label>ポジション</label>
          <input type="text" name="position" value="{m['position']}">
        </div>
      </div>
      <button class="btn btn-blue btn-block" type="submit">保存する</button>
    </form>
  </div>
  <div style="text-align:center"><a href="/t/{code}/admin/members" style="font-size:13px;color:#888">← メンバー一覧</a></div>
</div>'''
    return page('メンバー編集', body, code, active='members')


# ── Members (all users view) ──────────────────────────────────────

@app.route('/t/<code>/members')
def member_list(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY CAST(number AS INTEGER), name', (team['id'],)).fetchall()
    conn.close()

    rows = ''
    for m in members:
        rows += f'''
        <div class="card-sm row" style="align-items:center;gap:14px">
          <div style="width:32px;height:32px;border-radius:50%;background:#fef3c7;color:#d97706;font-weight:900;font-size:13px;display:flex;align-items:center;justify-content:center;flex-shrink:0">
            {m['number'] if m['number'] else '—'}
          </div>
          <div>
            <div style="font-weight:700">{m['name']}</div>
            {f'<div style="font-size:12px;color:#888">{m["position"]}</div>' if m['position'] else ''}
          </div>
        </div>'''

    edit_btn = f'<a href="/t/{code}/admin/members" class="btn btn-sm btn-blue">＋ 管理・追加</a>' if admin else ''
    home_url = f'/t/{code}/admin/dash' if (admin and not member) else f'/t/{code}/home'
    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div class="row" style="margin-bottom:16px">
      <div>
        <span class="section-label">{_ICO_PEOPLE} メンバー</span>
      </div>
      <div style="display:flex;align-items:center;gap:10px;margin-left:auto">
        <span class="badge badge-blue">{member_count_label(team, len(members))}</span>
        {edit_btn}
      </div>
    </div>
    {rows if members else '<div class="empty">まだメンバーがいません</div>'}
  </div>
  <div style="margin-top:8px"><a href="{home_url}" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('メンバー', body, code, active='members')


# ── Fees (member view) ────────────────────────────────────────────

@app.route('/t/<code>/fees')
def member_fees(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    fees = conn.execute('SELECT * FROM fees WHERE team_id=? ORDER BY due_date,created_at', (team['id'],)).fetchall()

    cards = ''
    for f in fees:
        if member:
            paid_row = conn.execute('SELECT paid FROM fee_payments WHERE fee_id=? AND member_name=?',
                                    (f['id'], member)).fetchone()
            paid = paid_row['paid'] if paid_row else 0
            badge = '<span class="badge badge-green">支払済</span>' if paid else '<span class="badge badge-red">未払い</span>'
        else:
            total = conn.execute('SELECT COUNT(*) FROM fee_payments WHERE fee_id=?', (f['id'],)).fetchone()[0]
            paid_count = conn.execute('SELECT COUNT(*) FROM fee_payments WHERE fee_id=? AND paid=1', (f['id'],)).fetchone()[0]
            badge = f'<span class="badge badge-blue">支払済 {paid_count}/{total}</span>'

        cards += f'''
        <div class="card-sm">
          <div class="row" style="justify-content:space-between;align-items:center">
            <div>
              <div style="font-weight:700">{f['title']}</div>
              <div style="font-size:13px;color:#555;margin-top:2px">
                ¥{f['amount']:,}{'　期限：' + fmt_date(f['due_date']) if f['due_date'] else ''}
              </div>
              {f'<div style="font-size:12px;color:#888;margin-top:4px">{f["note"]}</div>' if f['note'] else ''}
            </div>
            {badge}
          </div>
        </div>'''
    conn.close()

    new_btn = f'<a href="/t/{code}/admin/fees/new" class="btn btn-blue btn-sm">＋ 集金項目を追加</a>' if admin else ''
    home_url = f'/t/{code}/admin/dash' if (admin and not member) else f'/t/{code}/home'
    body = f'''
<div class="container">
  <div class="row" style="margin-bottom:16px">
    <span class="section-label">{_ICO_MONEY_SM} 集金</span>
    <div style="margin-left:auto">{new_btn}</div>
  </div>
  {cards if fees else (f'<div class="empty card"><div style="margin-bottom:8px">{_SVG_EMPTY_COIN}</div><div style="font-weight:700;margin-bottom:4px">集金項目はまだありません</div><div style="font-size:12px;color:#aaa;margin-bottom:16px">月会費や遠征費を作成して、誰が払ったか一覧管理できます。</div><a href="/t/{code}/admin/fees/new" class="btn btn-blue btn-sm">＋ 集金項目を追加</a></div>' if admin else f'<div class="empty card"><div style="margin-bottom:8px">{_SVG_EMPTY_COIN}</div>集金項目はまだありません</div>')}
  <div style="margin-top:8px"><a href="{home_url}" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('集金', body, code, active='fees')


# ── Admin: fees ───────────────────────────────────────────────────

@app.route('/t/<code>/admin/fees')
def admin_fees(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='fees')
    conn = get_db()
    fees = conn.execute('SELECT * FROM fees WHERE team_id=? ORDER BY due_date,created_at', (team['id'],)).fetchall()

    rows = ''
    for f in fees:
        paid_count = conn.execute('SELECT COUNT(*) FROM fee_payments WHERE fee_id=? AND paid=1', (f['id'],)).fetchone()[0]
        total = conn.execute('SELECT COUNT(*) FROM fee_payments WHERE fee_id=?', (f['id'],)).fetchone()[0]
        rows += f'''
        <div class="card-sm row" style="justify-content:space-between;align-items:center">
          <div>
            <div style="font-weight:700">{f['title']}</div>
            <div style="font-size:12px;color:#888">¥{f['amount']:,}{'　期限：' + fmt_date(f['due_date']) if f['due_date'] else ''}　支払済 {paid_count}/{total}名</div>
          </div>
          <a href="/t/{code}/admin/fees/{f['id']}" class="btn btn-sm btn-outline">管理</a>
        </div>'''
    conn.close()

    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div class="row" style="margin-bottom:16px">
      <h1 style="margin:0">{_ICO_MONEY_SM} 集金管理</h1>
      <a href="/t/{code}/admin/fees/new" class="btn btn-sm btn-blue" style="margin-left:auto">＋ 追加</a>
    </div>
    {rows if fees else '<div class="empty">集金項目がありません</div>'}
  </div>
  <div style="margin-top:8px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('集金管理', body, code, active='fees')


@app.route('/t/<code>/admin/fees/new', methods=['GET', 'POST'])
def admin_new_fee(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    error = ''
    conn = get_db()
    all_members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY number,name', (team['id'],)).fetchall()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        amount = request.form.get('amount', '0').strip().replace(',', '')
        due_date = request.form.get('due_date', '').strip()
        note = request.form.get('note', '').strip()
        selected = request.form.getlist('members')
        if not title:
            error = 'タイトルを入力してください'
        elif not selected:
            error = '対象メンバーを1人以上選択してください'
        else:
            fee_id = new_id()
            conn.execute('INSERT INTO fees VALUES (?,?,?,?,?,?,?)',
                         (fee_id, team['id'], title, int(amount or 0), due_date, note, now_str()))
            for name in selected:
                conn.execute('INSERT OR IGNORE INTO fee_payments VALUES (?,?,?,0,?)',
                             (new_id(), fee_id, name, ''))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_fee_detail', code=code, fee_id=fee_id))

    conn.close()

    member_checks = ''
    for m in all_members:
        num = f'<span style="font-size:11px;color:#aaa;margin-right:4px">#{m["number"]}</span>' if m['number'] else ''
        member_checks += f'''
<label style="display:flex;align-items:center;gap:10px;padding:9px 0;border-bottom:1px solid #f3f4f6;cursor:pointer;margin:0">
  <input type="checkbox" name="members" value="{m['name']}" checked style="width:16px;height:16px;accent-color:#d97706;flex-shrink:0">
  <span style="font-size:14px">{num}{m['name']}</span>
</label>'''

    all_check = f'''
<label style="display:flex;align-items:center;gap:10px;padding:9px 0;border-bottom:2px solid #e5e7eb;cursor:pointer;margin:0;font-weight:600">
  <input type="checkbox" id="chk-all" checked style="width:16px;height:16px;accent-color:#d97706;flex-shrink:0"
    onchange="document.querySelectorAll('[name=members]').forEach(c=>c.checked=this.checked)">
  <span style="font-size:13px;color:#6b7280">全員選択 / 全員解除</span>
</label>''' if all_members else ''

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>集金項目を追加</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" placeholder="例：月会費5月分、合宿費" required>
      <label>金額（円）</label>
      <input type="text" name="amount" placeholder="例：3000">
      <label>支払い期限</label>
      <input type="date" name="due_date">
      <label>備考</label>
      <textarea name="note" placeholder="振込先など" rows="3"></textarea>
      <div style="margin-top:18px">
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:8px">対象メンバー（{len(all_members)}名）</div>
        <div style="border:1px solid #e5e7eb;border-radius:8px;padding:0 12px;max-height:260px;overflow-y:auto">
          {all_check}
          {member_checks if member_checks else '<p style="padding:12px 0;color:#aaa;font-size:13px">メンバーがいません</p>'}
        </div>
      </div>
      <button class="btn btn-blue btn-block" type="submit" style="margin-top:20px">作成する</button>
    </form>
  </div>
  <div style="text-align:center"><a href="/t/{code}/admin/fees" style="font-size:13px;color:#888">← 集金一覧</a></div>
</div>'''
    return page('集金項目を追加', body, code, active='fees')


@app.route('/t/<code>/admin/fees/<fee_id>', methods=['GET', 'POST'])
def admin_fee_detail(code, fee_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    f = conn.execute('SELECT * FROM fees WHERE id=? AND team_id=?', (fee_id, team['id'])).fetchone()
    if not f:
        conn.close()
        return redirect(url_for('admin_fees', code=code))

    if request.method == 'POST':
        member_name = request.form.get('member_name')
        paid = int(request.form.get('paid', 0))
        paid_at = now_str() if paid else ''
        conn.execute('''
            INSERT INTO fee_payments (id,fee_id,member_name,paid,paid_at)
            VALUES (?,?,?,?,?)
            ON CONFLICT(fee_id,member_name) DO UPDATE SET paid=excluded.paid, paid_at=excluded.paid_at
        ''', (new_id(), fee_id, member_name, paid, paid_at))
        conn.commit()

    members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY name', (team['id'],)).fetchall()
    payments = conn.execute('SELECT * FROM fee_payments WHERE fee_id=?', (fee_id,)).fetchall()
    pay_map = {p['member_name']: p for p in payments}
    conn.close()

    def _rep(p):
        try: return p['reported'] if p else 0
        except (KeyError, IndexError): return 0

    rows = ''
    for m in members:
        p = pay_map.get(m['name'])
        paid = p['paid'] if p else 0
        paid_at = p['paid_at'] if p else ''
        reported = _rep(p) and not paid
        toggle_val = 0 if paid else 1
        if paid:
            btn_label = _CHK + ' 支払済'
        elif reported:
            btn_label = '申告→確定する'
        else:
            btn_label = '未払い'
        rep_badge = '<span style="font-size:11px;font-weight:700;color:#b45309;background:#fffbeb;border:1px solid #fde68a;border-radius:6px;padding:2px 7px;margin-left:8px">⏳ 申告あり</span>' if reported else ''
        rows += f'''
        <div class="card-sm row" style="justify-content:space-between;align-items:center;{'background:#fffdf5' if reported else ''}">
          <div>
            <span style="font-weight:700">{m['name']}</span>{rep_badge}
            {f'<span style="font-size:11px;color:#aaa;margin-left:8px">{paid_at}</span>' if paid_at else ''}
          </div>
          <form method="POST">
            <input type="hidden" name="member_name" value="{m['name']}">
            <input type="hidden" name="paid" value="{toggle_val}">
            <button class="btn btn-sm {'btn-blue' if paid else ('btn-amber' if reported else 'btn-outline')}" type="submit">{btn_label}</button>
          </form>
        </div>'''

    paid_count = sum(1 for m in members if pay_map.get(m['name']) and pay_map[m['name']]['paid'])
    report_count = sum(1 for m in members if _rep(pay_map.get(m['name'])) and not (pay_map.get(m['name']) and pay_map[m['name']]['paid']))

    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div class="row" style="justify-content:space-between;align-items:flex-start;margin-bottom:8px">
      <h1 style="margin:0">{f['title']}</h1>
      <div style="display:flex;gap:8px;flex-shrink:0">
        <a href="/t/{code}/admin/fees/{fee_id}/edit" class="btn btn-sm btn-outline">編集</a>
        <form method="POST" action="/t/{code}/admin/fees/{fee_id}/delete" onsubmit="return confirm('削除しますか？集金データもすべて消えます。')" style="margin:0">
          <button class="btn btn-sm btn-gray" type="submit" style="color:#dc2626">削除</button>
        </form>
      </div>
    </div>
    <div style="font-size:14px;color:#555">
      ¥{f['amount']:,}{'　期限：' + fmt_date(f['due_date']) if f['due_date'] else ''}
    </div>
    {f'<div style="font-size:13px;color:#666;margin-top:8px">{f["note"]}</div>' if f['note'] else ''}
    <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
      <span class="badge badge-green">支払済 {paid_count}名</span>
      <span class="badge badge-red">未払い {len(members)-paid_count}名</span>
      {f'<span class="badge" style="background:#fffbeb;color:#b45309;border:1px solid #fde68a">⏳ 申告 {report_count}名</span>' if report_count else ''}
    </div>
  </div>

  <div class="card" style="background:#fffdf7;border:1.5px solid #fde68a">
    <div style="font-weight:700;font-size:14px;margin-bottom:4px">📣 メンバーに支払いを申告してもらう</div>
    <div style="font-size:12px;color:#888;margin-bottom:10px;line-height:1.6">下のリンクを連絡アプリ（LINE等）に貼るだけ。メンバーは登録不要で「支払いました」をタップ→ここに「⏳ 申告あり」が出るので、確認して確定できます。</div>
    <div style="background:#fff;border:1px solid #eee;border-radius:8px;padding:8px 10px;font-size:11px;color:#374151;word-break:break-all;font-family:monospace;margin-bottom:8px">{base_url()}t/{code}/pay-answer/{team['viewer_token']}</div>
    <button onclick="navigator.clipboard.writeText('{base_url()}t/{code}/pay-answer/{team['viewer_token']}').then(()=>{{this.textContent='✓ コピーしました';setTimeout(()=>this.textContent='🔗 支払いリンクをコピー',1500)}})" class="btn btn-blue btn-sm" style="width:100%">🔗 支払いリンクをコピー</button>
  </div>

  <div class="card">
    <h2 style="margin-bottom:12px">支払い状況</h2>
    {rows if members else '<div class="empty">メンバーがいません。先にメンバー名簿を登録してください。</div>'}
  </div>
  <div style="text-align:center"><a href="/t/{code}/admin/fees" style="font-size:13px;color:#888">← 集金一覧</a></div>
</div>'''
    return page(f['title'], body, code, active='fees')


@app.route('/t/<code>/admin/fees/<fee_id>/edit', methods=['GET', 'POST'])
def admin_edit_fee(code, fee_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    f = conn.execute('SELECT * FROM fees WHERE id=? AND team_id=?', (fee_id, team['id'])).fetchone()
    if not f:
        conn.close()
        return redirect(url_for('admin_fees', code=code))

    error = ''
    if request.method == 'POST':
        title    = request.form.get('title', '').strip()
        amount   = request.form.get('amount', '0').strip().replace(',', '')
        due_date = request.form.get('due_date', '').strip()
        note     = request.form.get('note', '').strip()
        if not title:
            error = 'タイトルを入力してください'
        else:
            conn.execute('UPDATE fees SET title=?,amount=?,due_date=?,note=? WHERE id=?',
                         (title, int(amount or 0), due_date, note, fee_id))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_fee_detail', code=code, fee_id=fee_id))

    conn.close()
    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>集金項目を編集</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>タイトル *</label>
      <input type="text" name="title" value="{f['title']}" required>
      <label>金額（円）</label>
      <input type="text" name="amount" value="{f['amount']}">
      <label>支払い期限</label>
      <input type="date" name="due_date" value="{f['due_date'] or ''}">
      <label>備考</label>
      <textarea name="note" rows="3">{f['note'] or ''}</textarea>
      <button class="btn btn-blue btn-block" type="submit">保存する</button>
    </form>
  </div>
  <div style="text-align:center"><a href="/t/{code}/admin/fees/{fee_id}" style="font-size:13px;color:#888">← 集金詳細に戻る</a></div>
</div>'''
    return page('集金項目を編集', body, code, active='fees')


@app.route('/t/<code>/admin/fees/<fee_id>/delete', methods=['POST'])
def admin_delete_fee(code, fee_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    conn.execute('DELETE FROM fees WHERE id=? AND team_id=?', (fee_id, team['id']))
    conn.execute('DELETE FROM fee_payments WHERE fee_id=?', (fee_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_fees', code=code))


# ── Admin: CSV export ─────────────────────────────────────────────

@app.route('/t/<code>/admin/events/<event_id>/csv')
def admin_event_csv(code, event_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    ev = conn.execute('SELECT * FROM events WHERE id=? AND team_id=?', (event_id, team['id'])).fetchone()
    if not ev:
        conn.close()
        return redirect(url_for('schedule', code=code))

    members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY name', (team['id'],)).fetchall()
    rsvps = conn.execute('SELECT * FROM rsvps WHERE event_id=?', (event_id,)).fetchall()
    conn.close()

    rsvp_map = {r['member_name']: r['status'] for r in rsvps}
    status_label = {'attending': '出席', 'absent': '欠席'}

    rows = [['名前', '背番号', 'ポジション', '出欠', '更新日時']]
    if members:
        for m in members:
            status = rsvp_map.get(m['name'], '未回答')
            rows.append([m['name'], m['number'], m['position'],
                         status_label.get(status, status),
                         next((r['updated_at'] for r in rsvps if r['member_name'] == m['name']), '')])
    else:
        for name, status in rsvp_map.items():
            rows.append([name, '', '', status_label.get(status, status),
                         next((r['updated_at'] for r in rsvps if r['member_name'] == name), '')])

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return csv_response(output.getvalue(), f"出欠_{ev['title']}_{ev['event_date']}.csv")
    return excel_response(rows, f"出欠_{ev['title']}_{ev['event_date']}.xlsx")


# ── Order Forms ──────────────────────────────────────────────────

@app.route('/t/<code>/orders')
def orders_list(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))
    if not is_pro(team):
        return pro_gate(code, team, active='orders')

    conn = get_db()
    forms = conn.execute(
        'SELECT * FROM order_forms WHERE team_id=? ORDER BY created_at DESC',
        (team['id'],)
    ).fetchall()

    cards = ''
    for f in forms:
        field_count = conn.execute('SELECT COUNT(*) FROM order_form_fields WHERE form_id=?', (f['id'],)).fetchone()[0]
        response_count = conn.execute('SELECT COUNT(*) FROM order_responses WHERE form_id=?', (f['id'],)).fetchone()[0]
        has_responded = bool(conn.execute(
            'SELECT 1 FROM order_responses WHERE form_id=? AND member_name=?', (f['id'], member)
        ).fetchone()) if member else False

        if admin:
            badge = f'<span class="badge badge-blue">{response_count}件の回答</span>'
        elif has_responded:
            badge = '<span class="badge badge-green">回答済</span>'
        else:
            badge = '<span class="badge badge-red">未回答</span>'

        deadline_html = f'　期限：{fmt_date(f["deadline"])}' if f['deadline'] else ''
        cards += f'''
        <a href="/t/{code}/orders/{f['id']}" style="text-decoration:none;display:block">
          <div class="card-sm">
            <div class="row" style="justify-content:space-between">
              <div>
                <div style="font-weight:700;color:#1a1a1a">{f['title']}</div>
                <div style="font-size:12px;color:#aaa;margin-top:4px">{fmt_datetime(f['created_at'])}{deadline_html}　項目 {field_count}件</div>
              </div>
              {badge}
            </div>
          </div>
        </a>'''
    conn.close()

    new_btn = f'<a href="/t/{code}/admin/orders/new" class="btn btn-blue btn-sm">＋ フォーム作成</a>' if admin else ''
    home_url = f'/t/{code}/admin/dash' if (admin and not member) else f'/t/{code}/home'
    body = f'''
<div class="container">
  <div class="row" style="margin-bottom:16px">
    <span class="section-label">{_ICO_CLIPBOARD} 注文フォーム</span>
    <div style="margin-left:auto">{new_btn}</div>
  </div>
  {cards if forms else '<div class="empty card"><div style="margin-bottom:8px">' + _SVG_EMPTY_FORM + '</div>注文フォームはまだありません</div>'}
  <div style="margin-top:8px"><a href="{home_url}" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('注文フォーム', body, code, active='orders')


@app.route('/t/<code>/orders/<form_id>', methods=['GET', 'POST'])
def order_form_view(code, form_id):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    form = conn.execute(
        'SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])
    ).fetchone()
    if not form:
        conn.close()
        return redirect(url_for('orders_list', code=code))

    fields = conn.execute(
        'SELECT * FROM order_form_fields WHERE form_id=? ORDER BY sort_order', (form_id,)
    ).fetchall()

    photos = conn.execute(
        'SELECT * FROM order_form_photos WHERE form_id=? ORDER BY uploaded_at', (form_id,)
    ).fetchall()

    if not admin and request.method == 'POST' and member:
        resp = conn.execute(
            'SELECT id FROM order_responses WHERE form_id=? AND member_name=?', (form_id, member)
        ).fetchone()
        if resp:
            resp_id = resp['id']
            conn.execute('UPDATE order_responses SET submitted_at=? WHERE id=?', (now_str(), resp_id))
            conn.execute('DELETE FROM order_response_values WHERE response_id=?', (resp_id,))
        else:
            resp_id = new_id()
            conn.execute('INSERT INTO order_responses VALUES (?,?,?,?)',
                         (resp_id, form_id, member, now_str()))
        for field in fields:
            value = request.form.get(f'field_{field["id"]}', '').strip()
            conn.execute('INSERT INTO order_response_values VALUES (?,?,?,?)',
                         (new_id(), resp_id, field['id'], value))
        conn.commit()
        conn.close()
        return redirect(url_for('orders_list', code=code))

    if admin:
        responses = conn.execute(
            'SELECT * FROM order_responses WHERE form_id=? ORDER BY submitted_at', (form_id,)
        ).fetchall()

        resp_rows = ''
        for r in responses:
            vals = conn.execute(
                'SELECT * FROM order_response_values WHERE response_id=?', (r['id'],)
            ).fetchall()
            val_map = {v['field_id']: v['value'] for v in vals}
            cells = ''.join(
                f'<td style="padding:8px 12px;border-bottom:1px solid #e0e8ff">{val_map.get(f["id"], "")}</td>'
                for f in fields
            )
            resp_rows += (
                f'<tr><td style="padding:8px 12px;border-bottom:1px solid #e0e8ff;font-weight:700">{r["member_name"]}</td>'
                f'{cells}'
                f'<td style="padding:8px 12px;border-bottom:1px solid #e0e8ff;font-size:12px;color:#aaa">{fmt_datetime(r["submitted_at"])}</td></tr>'
            )

        headers = ''.join(
            f'<th style="padding:8px 12px;text-align:left;font-size:13px;color:#d97706">{f["label"]}</th>'
            for f in fields
        )

        field_rows = ''
        for f in fields:
            f_opts = f'（{f["options"]}）' if f['options'] else ''
            f_type_label = '選択' if f['field_type'] == 'select' else 'テキスト'
            field_rows += f'''
            <div class="card-sm row" style="justify-content:space-between;align-items:center">
              <div>
                <span style="font-weight:700">{f["label"]}</span>
                <span style="font-size:12px;color:#888;margin-left:8px">{f_type_label}{f_opts}</span>
              </div>
              <a href="/t/{code}/admin/orders/{form_id}/field/{f['id']}/delete"
                 class="btn btn-sm btn-gray"
                 onclick="return confirm('削除しますか？')">削除</a>
            </div>'''

        conn.close()

        deadline_html = f'<div style="font-size:13px;color:#f59e0b">期限：{fmt_date(form["deadline"])}</div>' if form['deadline'] else ''
        desc_html = f'<div style="font-size:13px;color:#666;margin-bottom:8px">{form["description"]}</div>' if form['description'] else ''

        photo_thumbs = ''.join(
            f'<div style="position:relative;display:inline-block">'
            f'<img src="/uploads/{p["id"]}" style="width:120px;height:90px;object-fit:cover;border-radius:10px;border:1.5px solid #e0e8ff">'
            f'<a href="/t/{code}/admin/orders/{form_id}/photo/{p["id"]}/delete"'
            f' style="position:absolute;top:-7px;right:-7px;background:#ef4444;color:#fff;border-radius:50%;width:22px;height:22px;font-size:13px;font-weight:900;display:flex;align-items:center;justify-content:center;text-decoration:none"'
            f' onclick="return confirm(\'削除しますか？\')">×</a>'
            f'</div>'
            for p in photos
        )
        photo_grid = f'<div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:16px">{photo_thumbs}</div>' if photos else '<div class="empty" style="padding:12px">まだ写真がありません</div>'
        photo_card = f'''
  <div class="card">
    <h2 style="margin-bottom:12px">📸 写真（メンバーにも表示されます）</h2>
    {photo_grid}
    <form method="POST" action="/t/{code}/admin/orders/{form_id}/photo" enctype="multipart/form-data">
      <label>写真を追加（複数選択可・JPG/PNG/GIF/WebP）</label>
      <input type="file" name="photos" accept="image/*" multiple style="padding:8px;background:#fafcff">
      <button class="btn btn-outline btn-sm" type="submit" style="margin-top:8px">📤 アップロード</button>
    </form>
  </div>'''

        if responses:
            table_html = (
                f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">'
                f'<thead><tr>'
                f'<th style="padding:8px 12px;text-align:left;font-size:13px;color:#d97706">名前</th>'
                f'{headers}'
                f'<th style="padding:8px 12px;text-align:left;font-size:13px;color:#d97706">回答日時</th>'
                f'</tr></thead><tbody>{resp_rows}</tbody></table></div>'
            )
        else:
            table_html = '<div class="empty">まだ回答がありません</div>'

        body = f'''
<div class="container" style="max-width:680px">
  <div class="card">
    <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:4px">
      <div style="font-weight:700;font-size:20px">{form["title"]}</div>
      <div style="display:flex;gap:6px;flex-shrink:0">
        <a href="/t/{code}/admin/orders/{form_id}/edit" class="btn btn-sm btn-outline">編集</a>
        <form method="POST" action="/t/{code}/admin/orders/{form_id}/delete"
              onsubmit="return confirm('このフォームを削除しますか？回答データもすべて消えます。')" style="margin:0">
          <button class="btn btn-sm btn-gray" type="submit" style="color:#dc2626">削除</button>
        </form>
      </div>
    </div>
    {desc_html}
    {deadline_html}
    <div style="margin-top:12px">
      <a href="/t/{code}/admin/orders/{form_id}/csv" class="btn btn-gray btn-sm">📥 Excel</a>
    </div>
  </div>

  <div class="card" style="background:#fffdf7;border:1.5px solid #fde68a">
    <div style="font-weight:700;font-size:14px;margin-bottom:4px">📣 メンバーに回答してもらう</div>
    <div style="font-size:12px;color:#888;margin-bottom:10px;line-height:1.6">下のリンクをコピーして、ふだんの連絡アプリ（LINE等）に貼るだけ。メンバーは登録不要で、名前を選んでタップ回答できます。</div>
    <div style="background:#fff;border:1px solid #eee;border-radius:8px;padding:8px 10px;font-size:11px;color:#374151;word-break:break-all;font-family:monospace;margin-bottom:8px">{base_url()}t/{code}/order-answer/{team['viewer_token']}/{form_id}</div>
    <button onclick="navigator.clipboard.writeText('{base_url()}t/{code}/order-answer/{team['viewer_token']}/{form_id}').then(()=>{{this.textContent='✓ コピーしました';setTimeout(()=>this.textContent='🔗 回答リンクをコピー',1500)}})" class="btn btn-blue btn-sm" style="width:100%">🔗 回答リンクをコピー</button>
  </div>

  {photo_card}

  <div class="card">
    <div class="row" style="margin-bottom:12px">
      <h2 style="margin:0">回答一覧 <span class="badge badge-blue" style="font-size:13px">{len(responses)}件</span></h2>
    </div>
    {table_html}
  </div>

  <div class="card">
    <h2 style="margin-bottom:12px">項目管理</h2>
    {field_rows if fields else '<div class="empty" style="padding:16px">まだ項目がありません</div>'}
    <form method="POST" action="/t/{code}/admin/orders/{form_id}/field" style="margin-top:16px">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
        <div>
          <label>項目名 *</label>
          <input type="text" name="label" placeholder="例：お弁当の種類" required>
        </div>
        <div>
          <label>種類</label>
          <select name="field_type" id="ftype" onchange="toggleOpts()">
            <option value="text">テキスト入力</option>
            <option value="select">選択肢から選ぶ</option>
          </select>
        </div>
      </div>
      <div id="opts-area" style="margin-top:10px">
        <label>選択肢（「選択肢から選ぶ」を選んだ場合のみ。カンマ区切り）</label>
        <input type="text" name="options" placeholder="例：のり弁,唐揚げ弁当,幕の内">
      </div>
      <button class="btn btn-outline btn-block" type="submit" style="margin-top:12px">＋ 項目を追加</button>
    </form>
  </div>

  <div style="text-align:center"><a href="/t/{code}/orders" style="font-size:13px;color:#888">← フォーム一覧</a></div>
</div>'''
        return page(form['title'], body, code, active='orders')

    # Member: fill form
    my_resp = conn.execute(
        'SELECT * FROM order_responses WHERE form_id=? AND member_name=?', (form_id, member)
    ).fetchone()
    my_values = {}
    if my_resp:
        vals = conn.execute(
            'SELECT * FROM order_response_values WHERE response_id=?', (my_resp['id'],)
        ).fetchall()
        my_values = {v['field_id']: v['value'] for v in vals}

    # 各フィールドの既存回答値を取得（候補として表示するため）
    existing_vals = {}
    for field in fields:
        rows = conn.execute(
            '''SELECT DISTINCT orv.value FROM order_response_values orv
               JOIN order_responses orr ON orv.response_id=orr.id
               WHERE orv.field_id=? AND orv.value!='' ORDER BY orv.value''',
            (field['id'],)
        ).fetchall()
        existing_vals[field['id']] = [r['value'] for r in rows]

    conn.close()

    field_inputs = ''
    for field in fields:
        current_val = my_values.get(field['id'], '')
        if field['field_type'] == 'select' and field['options']:
            opts_list = [o.strip() for o in field['options'].split(',') if o.strip()]
            options_html = '<option value="">選択してください</option>'
            options_html += ''.join(
                f'<option value="{o}" {"selected" if current_val == o else ""}>{o}</option>'
                for o in opts_list
            )
            field_inputs += f'<label>{field["label"]}</label><select name="field_{field["id"]}">{options_html}</select>'
        else:
            dl_id = f'dl_{field["id"]}'
            datalist_html = ''
            prev = existing_vals.get(field['id'], [])
            if prev:
                opts = ''.join(f'<option value="{o}">' for o in prev)
                datalist_html = f'<datalist id="{dl_id}">{opts}</datalist>'
            list_attr = f' list="{dl_id}"' if datalist_html else ''
            field_inputs += (
                f'<label>{field["label"]}</label>'
                f'{datalist_html}'
                f'<input type="text" name="field_{field["id"]}" value="{current_val}"'
                f' placeholder="入力してください"{list_attr}>'
            )

    submit_label = '更新する' if my_resp else '送信する'
    deadline_html = f'<div style="font-size:13px;color:#f59e0b;margin-bottom:12px">期限：{fmt_date(form["deadline"])}</div>' if form['deadline'] else ''
    desc_html = f'<div style="font-size:13px;color:#666;margin-bottom:12px">{form["description"]}</div>' if form['description'] else ''
    already_html = '<div class="msg-ok">' + _CHK + ' 回答済みです。修正して再送信できます。</div>' if my_resp else ''
    photos_html = ''.join(
        f'<img src="/uploads/{p["id"]}" style="width:100%;border-radius:10px;border:1.5px solid #e0e8ff;margin-bottom:10px;display:block">'
        for p in photos
    )

    no_fields_html = '<div class="empty" style="padding:20px">まだ項目が設定されていません</div>' if not fields else ''

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <div style="font-weight:700;font-size:20px;margin-bottom:8px">{form["title"]}</div>
    {desc_html}
    {deadline_html}
    {photos_html}
    {already_html}
    {no_fields_html}
    {'<form method="POST">' + field_inputs + f'<button class="btn btn-blue btn-block" type="submit">{submit_label}</button></form>' if fields else ''}
  </div>
  <div style="text-align:center"><a href="/t/{code}/orders" style="font-size:13px;color:#888">← 一覧に戻る</a></div>
</div>'''
    return page(form['title'], body, code, active='orders')


@app.route('/t/<code>/admin/orders/new', methods=['GET', 'POST'])
def admin_new_order_form(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    error = ''

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        deadline = request.form.get('deadline', '').strip()
        if not title:
            error = 'フォーム名を入力してください'
        else:
            conn = get_db()
            form_id = new_id()
            conn.execute('INSERT INTO order_forms VALUES (?,?,?,?,?,?)',
                         (form_id, team['id'], title, description, deadline, now_str()))
            conn.commit()
            conn.close()
            return redirect(url_for('order_form_view', code=code, form_id=form_id))

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>注文フォームを作成</h1>
    <p style="color:#666;font-size:13px;margin-bottom:16px">フォームを作成後、項目（質問）を追加できます</p>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>フォーム名 *</label>
      <input type="text" name="title" placeholder="例：遠征弁当注文、ユニフォームサイズ確認" required>
      <label>説明（任意）</label>
      <textarea name="description" placeholder="例：5/25（日）遠征分のお弁当を注文してください" rows="3"></textarea>
      <label>回答期限（任意）</label>
      <input type="date" name="deadline">
      <button class="btn btn-blue btn-block" type="submit">作成する →</button>
    </form>
  </div>
  <div style="text-align:center"><a href="/t/{code}/orders" style="font-size:13px;color:#888">← フォーム一覧</a></div>
</div>'''
    return page('フォーム作成', body, code, active='orders')


@app.route('/t/<code>/admin/orders/<form_id>/edit', methods=['GET', 'POST'])
def admin_edit_order_form(code, form_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    form = conn.execute('SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])).fetchone()
    if not form:
        conn.close()
        return redirect(url_for('orders_list', code=code))

    error = ''
    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        deadline    = request.form.get('deadline', '').strip()
        if not title:
            error = 'フォーム名を入力してください'
        else:
            conn.execute('UPDATE order_forms SET title=?,description=?,deadline=? WHERE id=?',
                         (title, description, deadline, form_id))
            conn.commit()
            conn.close()
            return redirect(url_for('order_form_view', code=code, form_id=form_id))

    conn.close()
    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>フォームを編集</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>フォーム名 *</label>
      <input type="text" name="title" value="{form['title']}" required>
      <label>説明（任意）</label>
      <textarea name="description" rows="3">{form['description'] or ''}</textarea>
      <label>回答期限（任意）</label>
      <input type="date" name="deadline" value="{form['deadline'] or ''}">
      <button class="btn btn-blue btn-block" type="submit">保存する</button>
    </form>
  </div>
  <div style="text-align:center">
    <a href="/t/{code}/orders/{form_id}" style="font-size:13px;color:#888">← フォームに戻る</a>
  </div>
</div>'''
    return page('フォームを編集', body, code, active='orders')


@app.route('/t/<code>/admin/orders/<form_id>/delete', methods=['POST'])
def admin_delete_order_form(code, form_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    form = conn.execute('SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])).fetchone()
    if form:
        resp_ids = [r['id'] for r in conn.execute('SELECT id FROM order_responses WHERE form_id=?', (form_id,)).fetchall()]
        for rid in resp_ids:
            conn.execute('DELETE FROM order_response_values WHERE response_id=?', (rid,))
        conn.execute('DELETE FROM order_responses WHERE form_id=?', (form_id,))
        conn.execute('DELETE FROM order_form_fields WHERE form_id=?', (form_id,))
        conn.execute('DELETE FROM order_form_photos WHERE form_id=?', (form_id,))
        conn.execute('DELETE FROM order_forms WHERE id=?', (form_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('orders_list', code=code))


@app.route('/uploads/<photo_id>')
def serve_photo(photo_id):
    if not photo_id.replace('-', '').isalnum():
        return 'Not found', 404
    conn = get_db()
    photo = conn.execute('SELECT * FROM order_form_photos WHERE id=?', (photo_id,)).fetchone()
    conn.close()
    if not photo:
        return 'Not found', 404
    path = os.path.join(UPLOAD_DIR, photo_id)
    if not os.path.exists(path):
        return 'Not found', 404
    return send_file(path, mimetype=photo['mime_type'])


@app.route('/t/<code>/admin/orders/<form_id>/photo', methods=['POST'])
def admin_upload_order_photo(code, form_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    form = conn.execute(
        'SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])
    ).fetchone()
    if not form:
        conn.close()
        return redirect(url_for('orders_list', code=code))

    allowed = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    for f in request.files.getlist('photos'):
        if f and f.mimetype in allowed:
            photo_id = new_id()
            f.save(os.path.join(UPLOAD_DIR, photo_id))
            conn.execute('INSERT INTO order_form_photos VALUES (?,?,?,?,?)',
                         (photo_id, form_id, f.filename or '', f.mimetype, now_str()))
    conn.commit()
    conn.close()
    return redirect(url_for('order_form_view', code=code, form_id=form_id))


@app.route('/t/<code>/admin/orders/<form_id>/photo/<photo_id>/delete')
def admin_delete_order_photo(code, form_id, photo_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    photo = conn.execute(
        'SELECT * FROM order_form_photos WHERE id=? AND form_id=?', (photo_id, form_id)
    ).fetchone()
    if photo:
        path = os.path.join(UPLOAD_DIR, photo_id)
        if os.path.exists(path):
            os.remove(path)
        conn.execute('DELETE FROM order_form_photos WHERE id=?', (photo_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('order_form_view', code=code, form_id=form_id))


@app.route('/t/<code>/admin/orders/<form_id>/field', methods=['POST'])
def admin_add_order_field(code, form_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    form = conn.execute(
        'SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])
    ).fetchone()
    if not form:
        conn.close()
        return redirect(url_for('orders_list', code=code))
    label = request.form.get('label', '').strip()
    field_type = request.form.get('field_type', 'text')
    options = request.form.get('options', '').strip()
    if label:
        sort_order = conn.execute(
            'SELECT COUNT(*) FROM order_form_fields WHERE form_id=?', (form_id,)
        ).fetchone()[0]
        conn.execute('INSERT INTO order_form_fields VALUES (?,?,?,?,?,?)',
                     (new_id(), form_id, label, field_type, options, sort_order))
        conn.commit()
    conn.close()
    return redirect(url_for('order_form_view', code=code, form_id=form_id))


@app.route('/t/<code>/admin/orders/<form_id>/field/<field_id>/delete')
def admin_delete_order_field(code, form_id, field_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    conn = get_db()
    conn.execute('DELETE FROM order_form_fields WHERE id=? AND form_id=?', (field_id, form_id))
    conn.commit()
    conn.close()
    return redirect(url_for('order_form_view', code=code, form_id=form_id))


@app.route('/t/<code>/admin/orders/<form_id>/csv')
def admin_order_form_csv(code, form_id):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    conn = get_db()
    form = conn.execute(
        'SELECT * FROM order_forms WHERE id=? AND team_id=?', (form_id, team['id'])
    ).fetchone()
    if not form:
        conn.close()
        return redirect(url_for('orders_list', code=code))
    fields = conn.execute(
        'SELECT * FROM order_form_fields WHERE form_id=? ORDER BY sort_order', (form_id,)
    ).fetchall()
    responses = conn.execute(
        'SELECT * FROM order_responses WHERE form_id=? ORDER BY submitted_at', (form_id,)
    ).fetchall()

    rows = [['名前', '回答日時'] + [f['label'] for f in fields]]
    for r in responses:
        vals = conn.execute(
            'SELECT * FROM order_response_values WHERE response_id=?', (r['id'],)
        ).fetchall()
        val_map = {v['field_id']: v['value'] for v in vals}
        rows.append([r['member_name'], r['submitted_at']] + [val_map.get(f['id'], '') for f in fields])
    conn.close()

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return csv_response(output.getvalue(), f"注文_{form['title']}_{now_str()[:10]}.csv")
    return excel_response(rows, f"注文_{form['title']}_{now_str()[:10]}.xlsx")


# ── Survey ───────────────────────────────────────────────────────

@app.route('/t/<code>/survey')
def survey_list(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))
    if not is_pro(team):
        return pro_gate(code, team, active='orders')

    conn = get_db()
    surveys = conn.execute('SELECT * FROM surveys WHERE team_id=? ORDER BY created_at DESC', (team['id'],)).fetchall()

    cards = ''
    for sv in surveys:
        options = conn.execute('SELECT * FROM survey_options WHERE survey_id=? ORDER BY sort_order', (sv['id'],)).fetchall()
        total = conn.execute('SELECT COUNT(*) FROM survey_answers WHERE survey_id=?', (sv['id'],)).fetchone()[0]
        answered = bool(conn.execute('SELECT 1 FROM survey_answers WHERE survey_id=? AND member_name=?', (sv['id'], member)).fetchone()) if member else True

        cards += f'''
        <a href="/t/{code}/survey/{sv['id']}" style="text-decoration:none;display:block">
          <div class="card-sm">
            <div class="row" style="justify-content:space-between">
              <div style="font-weight:700;color:#1a1a1a">{'📌 ' if not answered else ''}{sv['title']}</div>
              {'<span class="badge badge-red">未回答</span>' if not answered else f'<span class="badge badge-gray">回答済 {total}名</span>'}
            </div>
            <div style="font-size:12px;color:#aaa;margin-top:4px">{fmt_datetime(sv['created_at'])}　選択肢 {len(options)}件</div>
          </div>
        </a>'''
    conn.close()

    new_btn = f'<a href="/t/{code}/admin/survey/new" class="btn btn-blue btn-sm">＋ 作成</a>' if admin else ''
    body = f'''
<div class="container">
  <div class="row" style="margin-bottom:16px">
    <span class="section-label">{_ICO_CHART_SM} アンケート</span>
    <div style="margin-left:auto">{new_btn}</div>
  </div>
  {cards if surveys else '<div class="empty card"><div style="margin-bottom:8px">' + _SVG_EMPTY_CHART + '</div>アンケートはまだありません</div>'}
</div>'''
    return page('アンケート', body, code, active='survey')


@app.route('/t/<code>/survey/<survey_id>', methods=['GET', 'POST'])
def survey_detail(code, survey_id):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))

    conn = get_db()
    sv = conn.execute('SELECT * FROM surveys WHERE id=? AND team_id=?', (survey_id, team['id'])).fetchone()
    if not sv:
        conn.close()
        return redirect(url_for('survey_list', code=code))

    options = conn.execute('SELECT * FROM survey_options WHERE survey_id=? ORDER BY sort_order', (survey_id,)).fetchall()

    if request.method == 'POST' and member:
        option_id = request.form.get('option_id')
        if option_id:
            conn.execute('''
                INSERT INTO survey_answers (id,survey_id,option_id,member_name,answered_at)
                VALUES (?,?,?,?,?)
                ON CONFLICT(survey_id,member_name) DO UPDATE SET option_id=excluded.option_id, answered_at=excluded.answered_at
            ''', (new_id(), survey_id, option_id, member, now_str()))
            conn.commit()

    my_answer = conn.execute('SELECT option_id FROM survey_answers WHERE survey_id=? AND member_name=?', (survey_id, member)).fetchone() if member else None
    my_option_id = my_answer['option_id'] if my_answer else None

    results = {}
    for opt in options:
        count = conn.execute('SELECT COUNT(*) FROM survey_answers WHERE survey_id=? AND option_id=?', (survey_id, opt['id'])).fetchone()[0]
        results[opt['id']] = count
    total = sum(results.values())

    option_btns = ''
    for opt in options:
        selected = my_option_id == opt['id']
        count = results[opt['id']]
        pct = int(count / total * 100) if total > 0 else 0
        if member:
            option_btns += f'''
            <form method="POST" style="margin-bottom:8px">
              <input type="hidden" name="option_id" value="{opt['id']}">
              <button type="submit" class="btn {'btn-blue' if selected else 'btn-outline'}" style="width:100%;text-align:left;padding:12px 16px">
                <div style="display:flex;justify-content:space-between;align-items:center">
                  <span>{'✓ ' if selected else ''}{opt['label']}</span>
                  <span style="font-size:13px;opacity:.8">{count}票 ({pct}%)</span>
                </div>
                <div style="margin-top:6px;height:4px;border-radius:4px;background:{'rgba(255,255,255,.3)' if selected else '#fde68a'}">
                  <div style="height:4px;border-radius:4px;background:{'rgba(255,255,255,.8)' if selected else '#d97706'};width:{pct}%"></div>
                </div>
              </button>
            </form>'''
        else:
            option_btns += f'''
            <div style="margin-bottom:8px;padding:12px 16px;border:1.5px solid #e0e8ff;border-radius:10px;background:#fff">
              <div style="display:flex;justify-content:space-between">
                <span>{opt['label']}</span>
                <span style="font-size:13px;color:#888">{count}票 ({pct}%)</span>
              </div>
              <div style="margin-top:6px;height:4px;border-radius:4px;background:#e0e8ff">
                <div style="height:4px;border-radius:4px;background:#111;width:{pct}%"></div>
              </div>
            </div>'''

    conn.close()
    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div style="font-size:12px;color:#888;margin-bottom:8px">{fmt_datetime(sv['created_at'])}　回答 {total}名</div>
    <h1 style="margin-bottom:20px">{sv['title']}</h1>
    {option_btns}
  </div>
  <div style="text-align:center"><a href="/t/{code}/survey" style="font-size:13px;color:#888">← アンケート一覧</a></div>
</div>'''
    return page(sv['title'], body, code, active='survey')


@app.route('/t/<code>/admin/survey/new', methods=['GET', 'POST'])
def admin_new_survey(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not team:
        return redirect('/')
    error = ''

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        options = [v.strip() for v in request.form.getlist('option') if v.strip()]
        if not title:
            error = '質問を入力してください'
        elif len(options) < 2:
            error = '選択肢を2つ以上入力してください'
        else:
            conn = get_db()
            sid = new_id()
            conn.execute('INSERT INTO surveys VALUES (?,?,?,?)', (sid, team['id'], title, now_str()))
            for i, label in enumerate(options):
                conn.execute('INSERT INTO survey_options VALUES (?,?,?,?)', (new_id(), sid, label, i))
            conn.commit()
            conn.close()
            return redirect(url_for('survey_list', code=code))

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1>アンケートを作成</h1>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>質問 *</label>
      <input type="text" name="title" placeholder="例：来週の練習に参加できますか？" required>
      <label>選択肢（最大6つ）</label>
      <input type="text" name="option" placeholder="例：参加できる" style="margin-bottom:8px">
      <input type="text" name="option" placeholder="例：参加できない" style="margin-bottom:8px">
      <input type="text" name="option" placeholder="例：未定" style="margin-bottom:8px">
      <input type="text" name="option" placeholder="（任意）" style="margin-bottom:8px">
      <input type="text" name="option" placeholder="（任意）" style="margin-bottom:8px">
      <input type="text" name="option" placeholder="（任意）" style="margin-bottom:8px">
      <button class="btn btn-blue btn-block" type="submit">作成する</button>
    </form>
  </div>
  <div style="text-align:center"><a href="/t/{code}/survey" style="font-size:13px;color:#888">← アンケート一覧</a></div>
</div>'''
    return page('アンケート作成', body, code, active='survey')


# ── App Feedback ─────────────────────────────────────────────────

@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    sent = request.args.get('sent') == '1'
    from_code = request.args.get('from', '').strip().upper()
    back_url = f'/t/{from_code}/admin/dash' if from_code else '/'
    back_label = 'ホームに戻る' if from_code else 'トップに戻る'
    error = ''
    prefill_team = ''
    prefill_email = ''
    if from_code:
        _ft = get_team(from_code)
        if _ft:
            prefill_team = _ft['name']
            prefill_email = _ft['admin_email'] or ''
    if request.method == 'POST':
        team_name = request.form.get('team_name', '').strip()
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        subject = request.form.get('subject', '').strip()
        message = request.form.get('message', '').strip()
        if not team_name or not name or not email or not message:
            error = 'チーム名・お名前・メールアドレス・メッセージは必須です'
        else:
            conn = get_db()
            conn.execute(
                'INSERT INTO app_feedback (id,name,message,created_at,team_name,email,subject) VALUES (?,?,?,?,?,?,?)',
                (new_id(), name, message, now_str(), team_name, email, subject)
            )
            conn.commit()
            conn.close()
            send_inquiry_email(team_name, name, email, subject, message)
            return redirect('/feedback?sent=1')

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1 style="margin-bottom:4px">お問い合わせ</h1>
    <p style="font-size:13px;color:#666;margin-bottom:20px">いただいた内容をご確認の上、返信させていただきます。</p>
    {'<div class="msg-ok">送信しました！内容を確認の上、ご連絡いたします。</div>' if sent else ''}
    {'<div class="msg-err">' + error + '</div>' if error else ''}
    <form method="POST">
      <label>チーム名 <span style="font-size:11px;color:#9ca3af;font-weight:400">（任意）</span></label>
      <input type="text" name="team_name" value="{prefill_team}" placeholder="例：FC東京U-15（検討中の方は「検討中」とご記入ください）">
      <label>お名前 *</label>
      <input type="text" name="name" placeholder="例：田中 太郎" required>
      <label>メールアドレス *</label>
      <input type="text" name="email" value="{prefill_email}" placeholder="例：tanaka@example.com" required>
      <label>表題 *</label>
      <select name="subject" required>
        <option value="">選択してください</option>
        <option value="機能の要望">機能の要望</option>
        <option value="使い方について">使い方について</option>
        <option value="料金・プランについて">料金・プランについて</option>
        <option value="導入のご相談">導入のご相談</option>
        <option value="その他">その他</option>
      </select>
      <label>メッセージ *</label>
      <textarea name="message" rows="5" placeholder="ご要望・ご質問など、詳しくお聞かせください" required></textarea>
      <button class="btn btn-blue btn-block" type="submit">送信する</button>
    </form>
  </div>
  <div style="text-align:center"><a href="{back_url}" style="font-size:13px;color:#aaa">← {back_label}</a></div>
</div>'''
    return page('お問い合わせ', body, from_code or None, active='contact')


@app.route('/rak/lp-stats')
def rak_lp_stats():
    pw = request.args.get('pw', '')
    admin_pw = os.environ.get('RAK_ADMIN_PW', 'rakadmin2026')
    if pw != admin_pw:
        return redirect('/rak/feedback')
    conn = get_db()
    days = conn.execute('''
        SELECT substr(created_at,1,10) AS d,
               SUM(event='lp_view') AS lp,
               SUM(event='create_view') AS cv,
               SUM(event='team_created') AS tc,
               SUM(is_mobile) AS mob,
               COUNT(*) AS total
        FROM lp_events
        WHERE created_at >= datetime('now', '-7 days')
        GROUP BY d ORDER BY d DESC
    ''').fetchall()
    srcs = conn.execute('''
        SELECT src, event, COUNT(*) AS c FROM lp_events
        WHERE event IN ('create_view','team_created') AND created_at >= datetime('now', '-7 days')
        GROUP BY src, event ORDER BY c DESC
    ''').fetchall()
    conn.close()
    day_rows = ''.join(
        f"<tr><td>{r['d']}</td><td>{r['lp']}</td><td>{r['cv']}</td><td>{r['tc']}</td>"
        f"<td>{(r['cv'] / r['lp'] * 100) if r['lp'] else 0:.1f}%</td>"
        f"<td>{(r['tc'] / r['cv'] * 100) if r['cv'] else 0:.1f}%</td>"
        f"<td>{(r['mob'] / r['total'] * 100) if r['total'] else 0:.0f}%</td></tr>"
        for r in days) or '<tr><td colspan="7">データなし</td></tr>'
    src_rows = ''.join(
        f"<tr><td>{r['src'] or '(直接/不明)'}</td><td>{r['event']}</td><td>{r['c']}</td></tr>"
        for r in srcs) or '<tr><td colspan="3">データなし</td></tr>'
    body = f'''
<div class="container" style="max-width:720px">
  <h1>LPファネル（直近7日）</h1>
  <div class="card" style="overflow-x:auto">
    <table style="width:100%;font-size:13px;border-collapse:collapse" border="1" cellpadding="6">
      <tr><th>日付</th><th>LP表示</th><th>登録P表示</th><th>登録完了</th><th>LP→登録P</th><th>登録P→完了</th><th>モバイル率</th></tr>
      {day_rows}
    </table>
  </div>
  <div class="card" style="overflow-x:auto">
    <h2>CTA別（src）</h2>
    <table style="width:100%;font-size:13px;border-collapse:collapse" border="1" cellpadding="6">
      <tr><th>src</th><th>イベント</th><th>件数</th></tr>
      {src_rows}
    </table>
  </div>
</div>'''
    return page('LPファネル', body)


@app.route('/rak/feedback')
def rak_feedback_admin():
    pw = request.args.get('pw', '')
    admin_pw = os.environ.get('RAK_ADMIN_PW', 'rakadmin2026')
    if pw != admin_pw:
        body = '''
<div class="container" style="max-width:400px;padding-top:60px">
  <div class="card">
    <h1>管理者ログイン</h1>
    <form method="GET">
      <label>パスワード</label>
      <input type="password" name="pw" autofocus>
      <button class="btn btn-blue btn-block" type="submit">確認する</button>
    </form>
  </div>
</div>'''
        return page('Rak Admin', body)

    conn = get_db()
    items = conn.execute('SELECT * FROM app_feedback ORDER BY created_at DESC').fetchall()
    conn.close()

    def fb_val(row, key):
        try:
            return row[key] or ''
        except Exception:
            return ''

    rows = ''.join(f'''
    <div class="card-sm">
      <div style="font-size:12px;color:#aaa;margin-bottom:6px">{f["created_at"]}</div>
      <div style="font-weight:700;margin-bottom:4px">{fb_val(f,"subject") or "（表題なし）"}</div>
      <div style="font-size:13px;color:#555;margin-bottom:6px">
        {fb_val(f,"team_name")}　{f["name"] or "匿名"}　{fb_val(f,"email")}
      </div>
      <div style="white-space:pre-wrap;font-size:15px">{f["message"]}</div>
    </div>''' for f in items)

    body = f'''
<div class="container">
  <h1 style="margin-bottom:20px">フィードバック <span class="badge badge-blue">{len(items)}件</span></h1>
  {rows or '<div class="empty card">まだフィードバックはありません</div>'}
</div>'''
    return page('Rak フィードバック', body)


# ── Stripe / Upgrade ─────────────────────────────────────────────

@app.route('/t/<code>/upgrade')
def upgrade_page(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    already_pro = is_pro(team)
    trial_days = get_trial_days_left(team)
    # 正規課金のProかつStripe設定済み → 管理画面へ
    if already_pro and not trial_days and STRIPE_SECRET_KEY and team['stripe_customer_id']:
        manage_btn = f'<div style="margin-top:12px"><a href="/t/{code}/billing-portal" style="font-size:13px;color:#888;text-decoration:underline">サブスクリプションを管理・解約</a></div>'
        body = f'''
<div class="container" style="max-width:480px;padding-top:40px">
  <div class="card" style="text-align:center;padding:40px 24px">
    <div style="margin-bottom:16px">{_ICO_CELEBRATE}</div>
    <h1 style="font-size:22px;margin-bottom:8px">Proプラン利用中</h1>
    <p style="color:#666;font-size:14px">すべての機能をご利用いただけます。</p>
    <div style="margin-top:24px"><a href="/t/{code}/admin/dash" class="btn btn-blue btn-block" style="margin-top:0">ホームに戻る</a></div>
    {manage_btn}
  </div>
</div>'''
        return page('プラン', body, code, active='plan')
    # トライアル中 → 残り日数を表示してUpgrade促す
    if trial_days is not None:
        trial_banner = f'<div style="background:#fffbeb;border:1.5px solid #f59e0b;border-radius:12px;padding:16px;margin-bottom:24px;text-align:center"><div style="font-size:14px;font-weight:700;color:#d97706">トライアル中 — 残り{trial_days}日</div><div style="font-size:12px;color:#888;margin-top:4px">トライアル終了後、課金しないとFreeプランに戻ります</div></div>'
    else:
        trial_banner = ''

    stripe_ready = bool(STRIPE_SECRET_KEY and STRIPE_PRICE_ID_PRO)
    if stripe_ready:
        yearly_btn = ''
        if STRIPE_PRICE_ID_PRO_YEARLY:
            yearly_btn = f'''
    <form method="POST" action="/t/{code}/upgrade/checkout" style="margin-top:8px">
      <input type="hidden" name="plan" value="yearly">
      <button class="btn btn-block" type="submit" style="font-size:14px;padding:13px;background:#fff;color:#d97706;border:1.5px solid #f59e0b">年額プラン ¥9,800/年　<span style="font-size:11px;opacity:.8">（2ヶ月分お得）</span></button>
    </form>'''
        checkout_btn = f'''
    <form method="POST" action="/t/{code}/upgrade/checkout">
      <input type="hidden" name="plan" value="monthly">
      <button class="btn btn-blue btn-block" type="submit" style="font-size:16px;padding:15px;font-weight:700">14日間無料で試す</button>
    </form>{yearly_btn}
    '''
    else:
        checkout_btn = '<div class="msg-err">現在オンライン決済の準備中です。しばらくお待ちください。</div>'

    body = f'''
<div class="container" style="max-width:420px;padding-top:40px">
  {trial_banner}
  <div class="card" style="text-align:center;padding:32px 24px 28px">

    <div style="font-size:11px;font-weight:700;color:#d97706;letter-spacing:.1em;margin-bottom:12px">RAK PRO</div>
    <div style="font-size:40px;font-weight:900;color:#111;margin-bottom:2px">¥980<span style="font-size:15px;font-weight:400;color:#888">/月</span></div>
    <div style="font-size:12px;color:#aaa;margin-bottom:6px">年払い ¥9,800（2ヶ月分お得）</div>
    <div style="display:inline-block;background:#fef3c7;color:#92400e;font-size:11px;font-weight:700;padding:4px 12px;border-radius:999px;margin-bottom:24px">14日間無料トライアル付き</div>

    <div style="background:#f8f9fb;border-radius:10px;padding:16px 20px;margin-bottom:16px;text-align:left">
      <div style="font-size:13px;color:#444;line-height:2.2">
        {_CHK} 集金・支払い管理<br>
        {_CHK} 注文フォーム<br>
        {_CHK} 会計・収支記録<br>
        {_CHK} ユニフォーム管理（サイズ・枚数割当）<br>
        {_CHK} アンケート<br>
        {_CHK} AI文章生成<br>
        {_CHK} AIスケジュール自動生成<br>
        {_CHK} Excel出力<br>
        {_CHK} 優先サポート
      </div>
    </div>
    <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:10px 14px;margin-bottom:16px;font-size:12px;color:#166534;text-align:left">
      🔒 トライアル期間中に解約すれば料金はかかりません。<br>
      課金開始日：<strong>{(datetime.now(JST) + timedelta(days=14)).strftime('%Y年%m月%d日')}</strong>
    </div>

    {checkout_btn}

    <div style="font-size:11px;color:#bbb;margin-top:14px">いつでもキャンセル可能 · クレジットカード払い</div>

    <div style="margin-top:20px;border-top:1px solid #f0f0f0;padding-top:16px">
      <p style="font-size:12px;color:#bbb;margin-bottom:8px">プロモコードをお持ちの方</p>
      <form method="POST" action="/t/{code}/upgrade/promo" style="display:flex;gap:8px">
        <input type="text" name="promo" placeholder="プロモコード" style="flex:1;padding:9px 12px;border:1px solid #e5e7eb;border-radius:8px;font-size:13px;outline:none">
        <button type="submit" class="btn btn-outline" style="white-space:nowrap;padding:9px 14px;font-size:13px">適用</button>
      </form>
    </div>
    <div style="margin-top:14px"><a href="/t/{code}/admin/dash" style="font-size:12px;color:#bbb">← ホームに戻る</a></div>
  </div>
</div>'''
    return page('Proプランへアップグレード', body, code, active='plan')


@app.route('/t/<code>/billing-portal')
def billing_portal(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not STRIPE_SECRET_KEY or not team['stripe_customer_id']:
        return redirect(url_for('upgrade_page', code=code))
    import stripe
    stripe.api_key = STRIPE_SECRET_KEY
    portal = stripe.billing_portal.Session.create(
        customer=team['stripe_customer_id'],
        return_url=f"{base_url()}t/{code}/upgrade"
    )
    return redirect(portal.url)


@app.route('/t/<code>/upgrade/checkout', methods=['POST'])
def upgrade_checkout(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    if not STRIPE_SECRET_KEY or not STRIPE_PRICE_ID_PRO:
        return redirect(url_for('upgrade_page', code=code))
    import stripe
    stripe.api_key = STRIPE_SECRET_KEY
    team = get_team(code)
    plan = request.form.get('plan', 'monthly')
    price_id = STRIPE_PRICE_ID_PRO
    if plan == 'yearly' and STRIPE_PRICE_ID_PRO_YEARLY:
        price_id = STRIPE_PRICE_ID_PRO_YEARLY
    base = base_url().rstrip('/')
    checkout = stripe.checkout.Session.create(
        payment_method_types=['card'],
        line_items=[{'price': price_id, 'quantity': 1}],
        mode='subscription',
        subscription_data={'trial_period_days': 14},
        success_url=f'{base}/t/{code}/upgrade/success?session_id={{CHECKOUT_SESSION_ID}}',
        cancel_url=f'{base}/t/{code}/upgrade',
        metadata={'team_code': code},
        customer_email=team['admin_email'] or None,
    )
    return redirect(checkout.url)


@app.route('/t/<code>/upgrade/promo', methods=['POST'])
def upgrade_promo(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    promo = request.form.get('promo', '').strip().upper()
    valid_codes = [c.upper() for c in PROMO_CODES]
    if promo and valid_codes and promo in valid_codes:
        conn = get_db()
        conn.execute("UPDATE teams SET plan='pro' WHERE id=?", (team['id'],))
        conn.commit()
        conn.close()
        body = f'''
<div class="container" style="max-width:480px;padding-top:40px">
  <div class="card" style="text-align:center;padding:40px 24px">
    <div style="margin-bottom:16px">{_ICO_CELEBRATE}</div>
    <h1 style="font-size:22px;margin-bottom:8px">プロモコード適用完了！</h1>
    <p style="color:#666;font-size:14px;margin-bottom:24px">Rak Proへようこそ。すべての機能が使えるようになりました。</p>
    <a href="/t/{code}/admin/dash" class="btn btn-blue btn-block" style="margin-top:0">ホームに戻る</a>
  </div>
</div>'''
        return page('アップグレード完了', body, code, active='plan')
    body = f'''
<div class="container" style="max-width:480px;padding-top:40px">
  <div class="card" style="text-align:center;padding:40px 24px">
    <h1 style="font-size:20px;margin-bottom:12px">コードが無効です</h1>
    <p style="color:#666;font-size:14px;margin-bottom:24px">プロモコードが正しくないか、有効期限が切れています。</p>
    <a href="/t/{code}/upgrade" class="btn btn-outline btn-block" style="margin-top:0">← 戻る</a>
  </div>
</div>'''
    return page('コードエラー', body, code, active='home')


@app.route('/t/<code>/upgrade/success')
def upgrade_success(code):
    team = get_team(code)
    body = f'''
<div class="container" style="max-width:480px;padding-top:40px">
  <div class="card" style="text-align:center;padding:40px 24px">
    <div style="margin-bottom:16px">{_ICO_CELEBRATE}</div>
    <h1 style="font-size:22px;margin-bottom:8px">アップグレード完了！</h1>
    <p style="color:#666;font-size:14px;margin-bottom:24px">Rak Proへようこそ。すべての機能が使えるようになりました。</p>
    <a href="/t/{code}/admin/dash" class="btn btn-blue btn-block" style="margin-top:0">ホームに戻る</a>
  </div>
</div>'''
    return page('アップグレード完了', body, code, active='plan')


# ── Admin: uniforms ───────────────────────────────────────────────

@app.route('/t/<code>/admin/uniforms', methods=['GET', 'POST'])
def admin_uniforms(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='uniforms')
    error = ''

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        if not name:
            error = 'ユニフォーム名を入力してください'
        else:
            uid = new_id()
            conn = get_db()
            conn.execute('INSERT INTO uniforms VALUES (?,?,?,?,?)',
                         (uid, team['id'], name, description, now_str()))
            members = conn.execute('SELECT name FROM members WHERE team_id=?', (team['id'],)).fetchall()
            for m in members:
                conn.execute('INSERT OR IGNORE INTO uniform_assignments VALUES (?,?,?,?,?,?,?)',
                             (new_id(), uid, m['name'], '', '', 0, ''))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_uniform_detail', code=code, uid=uid))

    conn = get_db()
    uniforms = conn.execute('SELECT * FROM uniforms WHERE team_id=? ORDER BY created_at DESC', (team['id'],)).fetchall()
    rows = ''
    for u in uniforms:
        received = conn.execute('SELECT COUNT(*) FROM uniform_assignments WHERE uniform_id=? AND received=1', (u['id'],)).fetchone()[0]
        total = conn.execute('SELECT COUNT(*) FROM uniform_assignments WHERE uniform_id=?', (u['id'],)).fetchone()[0]
        rows += f'''
        <div class="card-sm row" style="justify-content:space-between;align-items:center">
          <div>
            <div style="font-weight:600">{u['name']}</div>
            <div style="font-size:12px;color:#888">受取済 {received}/{total}名{('　' + u['description']) if u['description'] else ''}</div>
          </div>
          <a href="/t/{code}/admin/uniforms/{u['id']}" class="btn btn-sm btn-outline">管理</a>
        </div>'''
    conn.close()

    body = f'''
<div class="container" style="max-width:540px">
  <div class="card">
    <div class="row" style="margin-bottom:16px">
      <h1 style="margin:0">{_ICO_UNIFORM} ユニフォーム管理</h1>
    </div>
    {rows if uniforms else '<div class="empty">ユニフォームがまだ登録されていません</div>'}
  </div>
  <div class="card">
    <h2 style="margin-bottom:12px">新しいユニフォームを追加</h2>
    {f'<div class="msg-err">{error}</div>' if error else ''}
    <form method="POST">
      <label>ユニフォーム名 *</label>
      <input type="text" name="name" placeholder="例：ホームユニ2026、アウェイユニ" required>
      <label>備考（任意）</label>
      <input type="text" name="description" placeholder="例：10月配布予定">
      <button class="btn btn-blue btn-block" type="submit">追加する</button>
    </form>
  </div>
  <div style="margin-top:8px"><a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a></div>
</div>'''
    return page('ユニフォーム管理', body, code, active='uniforms')


@app.route('/t/<code>/admin/uniforms/<uid>', methods=['GET', 'POST'])
def admin_uniform_detail(code, uid):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='uniforms')
    conn = get_db()
    u = conn.execute('SELECT * FROM uniforms WHERE id=? AND team_id=?', (uid, team['id'])).fetchone()
    if not u:
        conn.close()
        return redirect(url_for('admin_uniforms', code=code))

    if request.method == 'POST':
        members = conn.execute('SELECT name FROM members WHERE team_id=?', (team['id'],)).fetchall()
        for m in members:
            mname = m['name']
            size = request.form.get(f'size_{mname}', '').strip()
            number = request.form.get(f'number_{mname}', '').strip()
            quantity_s = request.form.get(f'qty_{mname}', '1').strip()
            quantity = int(quantity_s) if quantity_s.isdigit() and int(quantity_s) > 0 else 1
            received = 1 if request.form.get(f'received_{mname}') else 0
            notes = request.form.get(f'notes_{mname}', '').strip()
            conn.execute('''
                INSERT INTO uniform_assignments (id,uniform_id,member_name,size,number,quantity,received,notes)
                VALUES (?,?,?,?,?,?,?,?)
                ON CONFLICT(uniform_id,member_name) DO UPDATE SET
                  size=excluded.size, number=excluded.number,
                  quantity=excluded.quantity,
                  received=excluded.received, notes=excluded.notes
            ''', (new_id(), uid, mname, size, number, quantity, received, notes))
        conn.commit()

    members = conn.execute('SELECT * FROM members WHERE team_id=? ORDER BY CAST(number AS INTEGER), name', (team['id'],)).fetchall()
    assignments = conn.execute('SELECT * FROM uniform_assignments WHERE uniform_id=?', (uid,)).fetchall()
    assign_map = {a['member_name']: a for a in assignments}
    for m in members:
        if m['name'] not in assign_map:
            conn.execute('INSERT OR IGNORE INTO uniform_assignments VALUES (?,?,?,?,?,?,?)',
                         (new_id(), uid, m['name'], '', '', 0, ''))
    conn.commit()
    assignments = conn.execute('SELECT * FROM uniform_assignments WHERE uniform_id=?', (uid,)).fetchall()
    assign_map = {a['member_name']: a for a in assignments}
    conn.close()

    received_count = sum(1 for a in assign_map.values() if a['received'])

    rows = ''
    for m in members:
        a = assign_map.get(m['name'], {})
        size_val = a['size'] if a else ''
        num_val = a['number'] if a else ''
        qty_val = a['quantity'] if a and a['quantity'] else 1
        recv_val = a['received'] if a else 0
        notes_val = a['notes'] if a else ''
        checked = 'checked' if recv_val else ''
        rows += f'''
    <tr style="border-bottom:1px solid #f3f4f6">
      <td style="padding:10px 8px;font-weight:500;white-space:nowrap">{m['name']}</td>
      <td style="padding:10px 4px">
        <select name="size_{m['name']}" style="width:100%;padding:5px 6px;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;background:#fff">
          <option value="">-</option>
          {''.join(f'<option value="{s}" {"selected" if size_val==s else ""}>{s}</option>' for s in ['120','130','140','150','160','SS','XS','S','M','L','XL','XXL','3XL'])}
          <option value="{size_val}" {'selected' if size_val and size_val not in ['120','130','140','150','160','SS','XS','S','M','L','XL','XXL','3XL'] else ''}>{size_val if size_val and size_val not in ['120','130','140','150','160','SS','XS','S','M','L','XL','XXL','3XL'] else 'その他'}</option>
        </select>
      </td>
      <td style="padding:10px 4px">
        <input type="number" name="qty_{m['name']}" value="{qty_val}" min="1" max="99" style="width:54px;padding:5px 6px;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;text-align:center">
      </td>
      <td style="padding:10px 4px">
        <input type="text" name="number_{m['name']}" value="{num_val}" placeholder="#" style="width:54px;padding:5px 6px;border:1px solid #e5e7eb;border-radius:6px;font-size:13px;text-align:center">
      </td>
      <td style="padding:10px 8px;text-align:center">
        <input type="checkbox" name="received_{m['name']}" {checked} style="width:16px;height:16px;accent-color:#d97706">
      </td>
    </tr>'''

    desc_html = f'<div style="font-size:13px;color:#888;margin-bottom:8px">{u["description"]}</div>' if u['description'] else ''
    if not members:
        table_html = '<div class="empty">メンバーがいません。先にメンバー名簿を登録してください。</div>'
    else:
        table_html = f'''
    <form method="POST">
      <div style="overflow-x:auto">
        <table style="width:100%;border-collapse:collapse">
          <thead>
            <tr style="border-bottom:1px solid #e5e7eb">
              <th style="text-align:left;padding:8px;font-size:12px;color:#6b7280;font-weight:500">名前</th>
              <th style="text-align:left;padding:8px;font-size:12px;color:#6b7280;font-weight:500">サイズ</th>
              <th style="text-align:center;padding:8px;font-size:12px;color:#6b7280;font-weight:500">枚数</th>
              <th style="text-align:left;padding:8px;font-size:12px;color:#6b7280;font-weight:500">背番号</th>
              <th style="text-align:center;padding:8px;font-size:12px;color:#6b7280;font-weight:500">受取</th>
            </tr>
          </thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
      <div style="margin-top:14px">
        <button class="btn btn-blue btn-block" type="submit">保存する</button>
      </div>
    </form>'''

    body = f'''
<div class="container" style="max-width:600px">
  <div class="card" style="margin-bottom:12px">
    <div class="row" style="justify-content:space-between;align-items:flex-start;margin-bottom:4px">
      <h1 style="margin:0">{u['name']}</h1>
      <form method="POST" action="/t/{code}/admin/uniforms/{uid}/delete"
            onsubmit="return confirm('このユニフォームを削除しますか？')" style="margin:0">
        <button class="btn btn-sm btn-gray" type="submit" style="color:#dc2626">削除</button>
      </form>
    </div>
    {desc_html}
    <div style="display:flex;gap:10px;margin-top:8px">
      <span class="badge badge-green">受取済 {received_count}名</span>
      <span class="badge badge-red">未受取 {len(members)-received_count}名</span>
    </div>
  </div>
  <div class="card">
    {table_html}
  </div>
  <div style="text-align:center"><a href="/t/{code}/admin/uniforms" style="font-size:13px;color:#888">← 一覧に戻る</a></div>
</div>'''
    return page(u['name'], body, code, active='uniforms')


@app.route('/t/<code>/admin/uniforms/<uid>/delete', methods=['POST'])
def admin_delete_uniform(code, uid):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='home')
    conn = get_db()
    conn.execute('DELETE FROM uniform_assignments WHERE uniform_id=?', (uid,))
    conn.execute('DELETE FROM uniforms WHERE id=? AND team_id=?', (uid, team['id']))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_uniforms', code=code))


# ── Member: uniforms ──────────────────────────────────────────────

@app.route('/t/<code>/uniforms')
def member_uniforms(code):
    team = get_team(code)
    if not team:
        return redirect('/')
    member = get_member(code)
    admin = is_admin(code)
    if not member and not admin:
        return redirect(url_for('team_portal', code=code))
    if admin and not member:
        return redirect(url_for('admin_uniforms', code=code))

    conn = get_db()
    uniforms = conn.execute('SELECT * FROM uniforms WHERE team_id=? ORDER BY created_at', (team['id'],)).fetchall()
    rows = ''
    for u in uniforms:
        a = conn.execute('SELECT * FROM uniform_assignments WHERE uniform_id=? AND member_name=?', (u['id'], member)).fetchone()
        if not a:
            continue
        recv_badge = (
            f'<span style="font-size:11px;background:#f0fdf4;color:#16a34a;border-radius:4px;padding:2px 7px;font-weight:500">受取済</span>'
            if a['received'] else
            f'<span style="font-size:11px;background:#f9fafb;color:#6b7280;border-radius:4px;padding:2px 7px;font-weight:500">未受取</span>'
        )
        details = []
        if a['size']:   details.append(f'サイズ：{a["size"]}')
        if a['number']: details.append(f'番号：{a["number"]}')
        rows += f'''
    <div style="display:flex;align-items:center;gap:12px;padding:12px 0;border-bottom:1px solid var(--rak-line-soft)">
      <div style="flex:1;min-width:0">
        <div style="font-weight:500;font-size:14px">{u['name']}</div>
        <div style="font-size:12px;color:var(--rak-mute);margin-top:2px">{'　'.join(details) if details else '詳細未設定'}</div>
      </div>
      {recv_badge}
    </div>'''

    if not rows:
        rows = '<div style="padding:20px 0;text-align:center;color:var(--rak-mute);font-size:13px">ユニフォーム情報はまだ登録されていません</div>'

    conn.close()

    body = f'''
<div class="container" style="max-width:480px">
  <div class="card">
    <h1 style="margin-bottom:16px">{_ICO_UNIFORM} ユニフォーム</h1>
    {rows}
  </div>
</div>'''
    return page('ユニフォーム', body, code, active='uniforms')


# ── Admin: ledger (会計) ──────────────────────────────────────────

LEDGER_INCOME_CATS = ['会費', '寄付', 'スポンサー', 'その他収入']
LEDGER_EXPENSE_CATS = ['遠征費', '練習場代', 'ユニフォーム', '備品・消耗品', '飲食費', 'その他支出']

@app.route('/t/<code>/admin/ledger', methods=['GET', 'POST'])
def admin_ledger(code):
    if not is_admin(code):
        return redirect(url_for('admin_login', code=code))
    team = get_team(code)
    if not is_pro(team):
        return pro_gate(code, team, active='ledger')

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add':
            entry_type = request.form.get('type', 'expense')
            title      = request.form.get('title', '').strip()
            amount_s   = request.form.get('amount', '0').replace(',', '').strip()
            category   = request.form.get('category', '').strip()
            entry_date = request.form.get('entry_date', '').strip()
            memo       = request.form.get('memo', '').strip()
            if title and entry_date:
                conn = get_db()
                conn.execute('INSERT INTO ledger VALUES (?,?,?,?,?,?,?,?,?)',
                             (new_id(), team['id'], entry_type, title,
                              int(amount_s or 0), category, entry_date, memo, now_str()))
                conn.commit()
                conn.close()
        elif action == 'delete':
            entry_id = request.form.get('entry_id', '')
            conn = get_db()
            conn.execute('DELETE FROM ledger WHERE id=? AND team_id=?', (entry_id, team['id']))
            conn.commit()
            conn.close()
        return redirect(url_for('admin_ledger', code=code))

    filter_year = request.args.get('y', '')
    filter_month = request.args.get('m', '')
    conn = get_db()
    all_entries = conn.execute(
        'SELECT * FROM ledger WHERE team_id=? ORDER BY entry_date DESC, created_at DESC',
        (team['id'],)
    ).fetchall()
    conn.close()

    if filter_year and filter_month:
        entries = [e for e in all_entries if e['entry_date'].startswith(f'{filter_year}-{filter_month.zfill(2)}')]
    elif filter_year:
        entries = [e for e in all_entries if e['entry_date'].startswith(filter_year)]
    else:
        entries = all_entries

    total_income  = sum(e['amount'] for e in entries if e['type'] == 'income')
    total_expense = sum(e['amount'] for e in entries if e['type'] == 'expense')
    balance = total_income - total_expense

    bal_color = '#16a34a' if balance >= 0 else '#dc2626'

    rows = ''
    for e in entries:
        sign   = '+' if e['type'] == 'income' else '-'
        color  = '#16a34a' if e['type'] == 'income' else '#dc2626'
        cat_badge = f'<span style="font-size:10px;background:#f3f4f6;color:#6b7280;border-radius:4px;padding:1px 5px;margin-left:5px">{e["category"]}</span>' if e['category'] else ''
        memo_txt  = f'<div style="font-size:11px;color:#9ca3af;margin-top:1px">{e["memo"]}</div>' if e['memo'] else ''
        rows += f'''
    <div style="display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid #f3f4f6">
      <div style="min-width:52px;font-size:11px;color:#6b7280;flex-shrink:0">{fmt_date(e['entry_date'])}</div>
      <div style="flex:1;min-width:0">
        <div style="font-size:13px;font-weight:500">{e['title']}{cat_badge}</div>
        {memo_txt}
      </div>
      <div style="font-size:14px;font-weight:600;color:{color};flex-shrink:0">{sign}¥{e['amount']:,}</div>
      <form method="POST" onsubmit="return confirm('削除しますか？')" style="margin:0">
        <input type="hidden" name="action" value="delete">
        <input type="hidden" name="entry_id" value="{e['id']}">
        <button type="submit" style="background:none;border:none;color:#d1d5db;font-size:16px;cursor:pointer;padding:2px 4px">×</button>
      </form>
    </div>'''

    if not rows:
        rows = '<div style="padding:20px 0;text-align:center;color:#9ca3af;font-size:13px">収支記録がありません</div>'

    income_opts  = ''.join(f'<option value="{c}">{c}</option>' for c in LEDGER_INCOME_CATS)
    expense_opts = ''.join(f'<option value="{c}">{c}</option>' for c in LEDGER_EXPENSE_CATS)

    today_val = datetime.now(JST).strftime('%Y-%m-%d')

    years_available = sorted(set(e['entry_date'][:4] for e in all_entries if e['entry_date']), reverse=True)
    cur_year = filter_year or (years_available[0] if years_available else str(datetime.now(JST).year))
    year_opts = ''.join(f'<option value="{y}" {"selected" if y==filter_year else ""}>{y}年</option>' for y in years_available)
    month_opts = '<option value="">全月</option>' + ''.join(f'<option value="{str(i)}" {"selected" if str(i)==filter_month else ""}>{i}月</option>' for i in range(1,13))

    body = f'''
<div class="container" style="max-width:560px">

  <form method="GET" style="display:flex;gap:8px;margin-bottom:14px;align-items:center">
    <select name="y" onchange="this.form.submit()" style="flex:1;padding:8px 10px;border:1px solid #e5e7eb;border-radius:8px;font-size:14px;background:#fff">
      <option value="">全期間</option>
      {year_opts}
    </select>
    <select name="m" onchange="this.form.submit()" style="flex:1;padding:8px 10px;border:1px solid #e5e7eb;border-radius:8px;font-size:14px;background:#fff">
      {month_opts}
    </select>
    {'<a href="/t/' + code + '/admin/ledger" style="font-size:12px;color:#d97706;white-space:nowrap">リセット</a>' if (filter_year or filter_month) else ''}
  </form>

  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px">
    <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px;text-align:center">
      <div style="font-size:10px;color:#6b7280;margin-bottom:4px">収入合計</div>
      <div style="font-size:18px;font-weight:600;color:#16a34a">¥{total_income:,}</div>
    </div>
    <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px;text-align:center">
      <div style="font-size:10px;color:#6b7280;margin-bottom:4px">支出合計</div>
      <div style="font-size:18px;font-weight:600;color:#dc2626">¥{total_expense:,}</div>
    </div>
    <div style="background:#0a0a0a;border-radius:10px;padding:14px;text-align:center">
      <div style="font-size:10px;color:rgba(255,255,255,.5);margin-bottom:4px">残高</div>
      <div style="font-size:18px;font-weight:600;color:{bal_color if balance != 0 else '#fff'}">{"+" if balance > 0 else ""}¥{balance:,}</div>
    </div>
  </div>

  <div class="card" style="margin-bottom:12px">
    <h2 style="margin-bottom:12px">収支を追加</h2>
    <form method="POST">
      <input type="hidden" name="action" value="add">
      <input type="hidden" name="type" id="type-hidden" value="expense">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px">
        <button type="button" id="btn-income" onclick="rakSetType('income')"
          style="padding:10px 12px;border:1px solid #e5e7eb;border-radius:8px;background:#f9fafb;font-size:13px;font-weight:500;color:#6b7280;cursor:pointer;text-align:center;width:100%">収入</button>
        <button type="button" id="btn-expense" onclick="rakSetType('expense')"
          style="padding:10px 12px;border:1.5px solid #dc2626;border-radius:8px;background:#fff5f5;font-size:13px;font-weight:500;color:#dc2626;cursor:pointer;text-align:center;width:100%">支出</button>
      </div>
      <div style="display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:8px">
        <div>
          <label>タイトル *</label>
          <input type="text" name="title" placeholder="例：5月会費" required>
        </div>
        <div>
          <label>金額（円）*</label>
          <input type="text" name="amount" placeholder="例：5000" required>
        </div>
        <div>
          <label>日付 *</label>
          <input type="date" name="entry_date" value="{today_val}" required style="line-height:1.2;height:auto;padding:10px 12px">
        </div>
        <div>
          <label>カテゴリ</label>
          <select name="category" id="cat-select" style="width:100%;padding:9px 10px;border:1px solid #e5e7eb;border-radius:8px;font-size:14px;background:#fff">
            <option value="">選択しない</option>
            <optgroup label="収入" id="income-cats">{income_opts}</optgroup>
            <optgroup label="支出" id="expense-cats">{expense_opts}</optgroup>
          </select>
        </div>
      </div>
      <label style="margin-top:8px">メモ（任意）</label>
      <input type="text" name="memo" placeholder="補足など">
      <button class="btn btn-blue btn-block" type="submit" style="margin-top:4px">追加する</button>
    </form>
  </div>

  <div class="card">
    <h2 style="margin-bottom:12px">収支一覧</h2>
    {rows}
  </div>

  <div style="margin-top:8px">
    <a href="/t/{code}/admin/dash" style="font-size:13px;color:#888">← ホームに戻る</a>
  </div>
</div>
<script>
var _rakType = 'expense';
function rakSetType(t){{
  _rakType = t;
  document.getElementById('type-hidden').value = t;
  var bi = document.getElementById('btn-income');
  var be = document.getElementById('btn-expense');
  if(t==='income'){{
    bi.style.border='1.5px solid #16a34a'; bi.style.background='#f0fdf4'; bi.style.color='#16a34a';
    be.style.border='1px solid #e5e7eb'; be.style.background='#f9fafb'; be.style.color='#6b7280';
  }}else{{
    be.style.border='1.5px solid #dc2626'; be.style.background='#fff5f5'; be.style.color='#dc2626';
    bi.style.border='1px solid #e5e7eb'; bi.style.background='#f9fafb'; bi.style.color='#6b7280';
  }}
  var ig=document.getElementById('income-cats');
  var eg=document.getElementById('expense-cats');
  ig.style.display=t==='income'?'':'none';
  eg.style.display=t==='expense'?'':'none';
  document.getElementById('cat-select').value='';
}}
(function(){{
  var ig=document.getElementById('income-cats');
  var eg=document.getElementById('expense-cats');
  ig.style.display='none';
}})();
</script>'''
    return page('会計', body, code, active='ledger')


@app.route('/stripe/webhook', methods=['POST'])
def stripe_webhook():
    import json as _json
    if not STRIPE_SECRET_KEY:
        return jsonify(ok=True)
    payload = request.get_data()
    sig = request.headers.get('Stripe-Signature', '')
    # 署名検証
    try:
        import stripe
        stripe.api_key = STRIPE_SECRET_KEY
        stripe.Webhook.construct_event(payload, sig, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        print(f'[WEBHOOK SIG ERROR] {type(e).__name__}: {e}')
        return jsonify(error='invalid'), 400
    # ペイロードはJSONとして直接扱う（stripe objectの型問題を回避）
    try:
        ev = _json.loads(payload)
        event_type = ev.get('type', '')
        obj = ev.get('data', {}).get('object', {})
        print(f'[WEBHOOK] type={event_type}')

        if event_type == 'checkout.session.completed':
            team_code = str((obj.get('metadata') or {}).get('team_code', ''))
            customer = str(obj.get('customer') or '')
            subscription = str(obj.get('subscription') or '')
            print(f'[WEBHOOK] team_code={team_code!r} customer={customer} subscription={subscription}')
            if team_code:
                conn = get_db()
                conn.execute(
                    'UPDATE teams SET plan="pro", stripe_customer_id=?, stripe_subscription_id=? WHERE team_code=?',
                    (customer, subscription, team_code.upper())
                )
                conn.commit()
                conn.close()
                print(f'[WEBHOOK] plan→pro team_code={team_code}')

        elif event_type == 'customer.subscription.deleted':
            sub_id = str(obj.get('id') or '')
            print(f'[WEBHOOK] subscription.deleted sub_id={sub_id}')
            if sub_id:
                conn = get_db()
                conn.execute(
                    'UPDATE teams SET plan="free", stripe_subscription_id="" WHERE stripe_subscription_id=?',
                    (sub_id,)
                )
                conn.commit()
                conn.close()

    except Exception as e:
        print(f'[WEBHOOK ERROR] {type(e).__name__}: {e}')
        return jsonify(error='server error'), 500

    return jsonify(ok=True)


# ── Super Admin ──────────────────────────────────────────────────

@app.route('/superadmin/teams')
def superadmin_teams():
    import base64
    auth = request.headers.get('Authorization', '')
    if BASIC_AUTH_USER and BASIC_AUTH_PASS:
        expected = base64.b64encode(f'{BASIC_AUTH_USER}:{BASIC_AUTH_PASS}'.encode()).decode()
        if auth != f'Basic {expected}':
            return Response('Unauthorized', 401, {'WWW-Authenticate': 'Basic realm="Admin"'})
    conn = get_db()
    teams = conn.execute('SELECT name, sport, team_code, plan, created_at FROM teams ORDER BY created_at DESC').fetchall()
    members = conn.execute('SELECT team_id, COUNT(*) as cnt FROM members GROUP BY team_id').fetchall()
    conn.close()
    member_map = {m['team_id']: m['cnt'] for m in members}
    rows = ''.join(
        f'<tr><td>{t["created_at"][:16]}</td><td>{t["name"]}</td><td>{t["sport"]}</td>'
        f'<td>{t["team_code"]}</td><td>{t["plan"]}</td><td>{member_map.get(t["team_code"], 0)}</td></tr>'
        for t in teams
    )
    html = f'''<!DOCTYPE html><html><head><meta charset="UTF-8">
    <title>Rak Admin</title>
    <style>body{{font-family:sans-serif;padding:24px;}}table{{border-collapse:collapse;width:100%;}}
    th,td{{border:1px solid #ddd;padding:8px 12px;text-align:left;font-size:14px;}}
    th{{background:#f3f4f6;}}tr:hover{{background:#f9fafb;}}</style></head><body>
    <h2>Rak チーム一覧（{len(teams)}件）</h2>
    <table><tr><th>登録日時</th><th>チーム名</th><th>競技</th><th>コード</th><th>プラン</th><th>メンバー数</th></tr>
    {rows}</table></body></html>'''
    return html


# ── Run ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3004))
    print(f'Rak アプリ起動中: http://localhost:{port}')
    app.run(host='0.0.0.0', port=port, debug=False)
